from __future__ import annotations

import hashlib
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import text

from app.database import engine, init_database


PROJECT_ROOT = Path(__file__).resolve().parent
DATA_SOURCES_DIR = PROJECT_ROOT / "data_sources"

WORKBOOKS = [
    {
        "workbook_key": "MPPS_MAY_2026",
        "workbook_type": "MPPS",
        "file_name": "MPPS Ver-04  MAY 2026.xlsx",
    },
    {
        "workbook_key": "OVEN_PLAN_JUNE_2026",
        "workbook_type": "OVEN_PLAN",
        "file_name": "OVEN SHEET PLAN  JUNE 01-2026.xlsx",
    },
]


def calculate_file_hash(file_path: Path) -> str:
    sha256 = hashlib.sha256()

    with file_path.open("rb") as file:
        for block in iter(lambda: file.read(1024 * 1024), b""):
            sha256.update(block)

    return sha256.hexdigest()


def normalise_value(value: Any) -> str | None:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")

    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")

    return str(value)


def number_value(value: Any):
    if isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):
        return value

    return None


def date_value(value: Any):
    if isinstance(value, datetime):
        return value

    if isinstance(value, date):
        return datetime.combine(value, time.min)

    return None


def is_formula_value(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def get_or_create_workbook(connection, workbook_info: dict[str, str], file_hash: str) -> int:
    row = connection.execute(
        text(
            """
            INSERT INTO excel_workbooks (
                workbook_key,
                original_file_name,
                workbook_type,
                file_hash,
                imported_at,
                is_active,
                note
            )
            VALUES (
                :workbook_key,
                :original_file_name,
                :workbook_type,
                :file_hash,
                CURRENT_TIMESTAMP,
                TRUE,
                :note
            )
            ON CONFLICT (workbook_key)
            DO UPDATE SET
                original_file_name = EXCLUDED.original_file_name,
                workbook_type = EXCLUDED.workbook_type,
                file_hash = EXCLUDED.file_hash,
                imported_at = CURRENT_TIMESTAMP,
                is_active = TRUE,
                note = EXCLUDED.note,
                updated_at = CURRENT_TIMESTAMP
            RETURNING id;
            """
        ),
        {
            "workbook_key": workbook_info["workbook_key"],
            "original_file_name": workbook_info["file_name"],
            "workbook_type": workbook_info["workbook_type"],
            "file_hash": file_hash,
            "note": "Raw Excel workbook preserved into database.",
        },
    ).mappings().one()

    return int(row["id"])


def start_import_run(connection, workbook_id: int) -> int:
    row = connection.execute(
        text(
            """
            INSERT INTO excel_import_runs (
                workbook_id,
                import_type,
                status,
                message
            )
            VALUES (
                :workbook_id,
                'RAW_EXCEL_IMPORT',
                'STARTED',
                'Raw Excel import started.'
            )
            RETURNING id;
            """
        ),
        {
            "workbook_id": workbook_id,
        },
    ).mappings().one()

    return int(row["id"])


def finish_import_run(connection, import_run_id: int, status: str, message: str) -> None:
    connection.execute(
        text(
            """
            UPDATE excel_import_runs
            SET
                status = :status,
                finished_at = CURRENT_TIMESTAMP,
                message = :message
            WHERE id = :import_run_id;
            """
        ),
        {
            "import_run_id": import_run_id,
            "status": status,
            "message": message,
        },
    )


def upsert_sheet(
    connection,
    workbook_id: int,
    sheet_name: str,
    sheet_index: int,
    max_row: int,
    max_column: int,
) -> int:
    row = connection.execute(
        text(
            """
            INSERT INTO excel_sheets (
                workbook_id,
                sheet_name,
                sheet_index,
                max_row,
                max_column,
                is_active
            )
            VALUES (
                :workbook_id,
                :sheet_name,
                :sheet_index,
                :max_row,
                :max_column,
                TRUE
            )
            ON CONFLICT (workbook_id, sheet_name)
            DO UPDATE SET
                sheet_index = EXCLUDED.sheet_index,
                max_row = EXCLUDED.max_row,
                max_column = EXCLUDED.max_column,
                is_active = TRUE,
                updated_at = CURRENT_TIMESTAMP
            RETURNING id;
            """
        ),
        {
            "workbook_id": workbook_id,
            "sheet_name": sheet_name,
            "sheet_index": sheet_index,
            "max_row": max_row,
            "max_column": max_column,
        },
    ).mappings().one()

    return int(row["id"])


def upsert_rows(connection, row_batch: list[dict[str, Any]]) -> None:
    if not row_batch:
        return

    connection.execute(
        text(
            """
            INSERT INTO excel_raw_rows (
                sheet_id,
                row_number,
                is_empty,
                row_text
            )
            VALUES (
                :sheet_id,
                :row_number,
                :is_empty,
                :row_text
            )
            ON CONFLICT (sheet_id, row_number)
            DO UPDATE SET
                is_empty = EXCLUDED.is_empty,
                row_text = EXCLUDED.row_text;
            """
        ),
        row_batch,
    )


def upsert_cells(connection, cell_batch: list[dict[str, Any]]) -> None:
    if not cell_batch:
        return

    connection.execute(
        text(
            """
            INSERT INTO excel_raw_cells (
                sheet_id,
                row_number,
                column_number,
                column_letter,
                cell_address,
                raw_value,
                display_value,
                data_type,
                number_value,
                date_value,
                formula_value,
                is_formula,
                mapped_table,
                mapped_field,
                mapping_status
            )
            VALUES (
                :sheet_id,
                :row_number,
                :column_number,
                :column_letter,
                :cell_address,
                :raw_value,
                :display_value,
                :data_type,
                :number_value,
                :date_value,
                :formula_value,
                :is_formula,
                NULL,
                NULL,
                'RAW_ONLY'
            )
            ON CONFLICT (sheet_id, row_number, column_number)
            DO UPDATE SET
                column_letter = EXCLUDED.column_letter,
                cell_address = EXCLUDED.cell_address,
                raw_value = EXCLUDED.raw_value,
                display_value = EXCLUDED.display_value,
                data_type = EXCLUDED.data_type,
                number_value = EXCLUDED.number_value,
                date_value = EXCLUDED.date_value,
                formula_value = EXCLUDED.formula_value,
                is_formula = EXCLUDED.is_formula;
            """
        ),
        cell_batch,
    )


def import_sheet(connection, formula_ws, display_ws, sheet_id: int) -> tuple[int, int]:
    inserted_rows = 0
    inserted_cells = 0

    row_batch: list[dict[str, Any]] = []
    cell_batch: list[dict[str, Any]] = []

    max_row = formula_ws.max_row or 0
    max_column = formula_ws.max_column or 0

    for row_number in range(1, max_row + 1):
        row_text_parts: list[str] = []
        row_has_value = False

        for column_number in range(1, max_column + 1):
            formula_cell = formula_ws.cell(row=row_number, column=column_number)
            display_cell = display_ws.cell(row=row_number, column=column_number)

            raw = formula_cell.value
            display = display_cell.value

            raw_text = normalise_value(raw)
            display_text = normalise_value(display)

            if raw_text is None and display_text is None:
                continue

            row_has_value = True

            if display_text is not None:
                row_text_parts.append(display_text)
            elif raw_text is not None:
                row_text_parts.append(raw_text)

            column_letter = get_column_letter(column_number)
            cell_address = f"{column_letter}{row_number}"
            is_formula = is_formula_value(raw)

            cell_batch.append(
                {
                    "sheet_id": sheet_id,
                    "row_number": row_number,
                    "column_number": column_number,
                    "column_letter": column_letter,
                    "cell_address": cell_address,
                    "raw_value": raw_text,
                    "display_value": display_text,
                    "data_type": str(formula_cell.data_type or ""),
                    "number_value": number_value(display),
                    "date_value": date_value(display),
                    "formula_value": raw_text if is_formula else None,
                    "is_formula": is_formula,
                }
            )

            inserted_cells += 1

            if len(cell_batch) >= 1000:
                upsert_cells(connection, cell_batch)
                cell_batch.clear()

        row_batch.append(
            {
                "sheet_id": sheet_id,
                "row_number": row_number,
                "is_empty": not row_has_value,
                "row_text": " | ".join(row_text_parts) if row_text_parts else None,
            }
        )

        inserted_rows += 1

        if len(row_batch) >= 1000:
            upsert_rows(connection, row_batch)
            row_batch.clear()

    upsert_rows(connection, row_batch)
    upsert_cells(connection, cell_batch)

    return inserted_rows, inserted_cells


def import_workbook(workbook_info: dict[str, str]) -> None:
    file_path = DATA_SOURCES_DIR / workbook_info["file_name"]

    if not file_path.exists():
        raise FileNotFoundError(
            f"Excel file not found: {file_path}\n"
            "Please copy the Excel file into the data_sources folder."
        )

    print("")
    print(f"Importing workbook: {workbook_info['file_name']}")

    file_hash = calculate_file_hash(file_path)

    formula_wb = load_workbook(file_path, data_only=False, read_only=False)
    display_wb = load_workbook(file_path, data_only=True, read_only=False)

    total_rows = 0
    total_cells = 0

    with engine.begin() as connection:
        workbook_id = get_or_create_workbook(connection, workbook_info, file_hash)
        import_run_id = start_import_run(connection, workbook_id)

        try:
            for sheet_index, sheet_name in enumerate(formula_wb.sheetnames, start=1):
                formula_ws = formula_wb[sheet_name]
                display_ws = display_wb[sheet_name]

                sheet_id = upsert_sheet(
                    connection,
                    workbook_id=workbook_id,
                    sheet_name=sheet_name,
                    sheet_index=sheet_index,
                    max_row=formula_ws.max_row or 0,
                    max_column=formula_ws.max_column or 0,
                )

                rows_count, cells_count = import_sheet(
                    connection,
                    formula_ws=formula_ws,
                    display_ws=display_ws,
                    sheet_id=sheet_id,
                )

                total_rows += rows_count
                total_cells += cells_count

                print(
                    f"  Sheet imported: {sheet_name} | "
                    f"rows={rows_count} | cells={cells_count}"
                )

            finish_import_run(
                connection,
                import_run_id=import_run_id,
                status="SUCCESS",
                message=(
                    f"Raw import completed. "
                    f"Sheets={len(formula_wb.sheetnames)}, "
                    f"Rows={total_rows}, "
                    f"Cells={total_cells}"
                ),
            )

        except Exception as exc:
            finish_import_run(
                connection,
                import_run_id=import_run_id,
                status="FAILED",
                message=str(exc),
            )
            raise

    formula_wb.close()
    display_wb.close()

    print(
        f"Workbook completed: {workbook_info['file_name']} | "
        f"rows={total_rows} | cells={total_cells}"
    )


def main() -> None:
    print("Starting raw Excel to database import...")
    print(f"Data source folder: {DATA_SOURCES_DIR}")

    init_database()

    if not DATA_SOURCES_DIR.exists():
        raise FileNotFoundError(
            f"Data source folder not found: {DATA_SOURCES_DIR}"
        )

    for workbook_info in WORKBOOKS:
        import_workbook(workbook_info)

    print("")
    print("Raw Excel import completed successfully.")
    print("Excel data is now preserved in:")
    print("- excel_workbooks")
    print("- excel_sheets")
    print("- excel_raw_rows")
    print("- excel_raw_cells")
    print("- excel_import_runs")
    print("")
    print("Next step: map raw Excel data into clean MPPS factory tables.")


if __name__ == "__main__":
    main()