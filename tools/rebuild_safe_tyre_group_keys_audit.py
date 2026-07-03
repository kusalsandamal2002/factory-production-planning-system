from pathlib import Path
from datetime import datetime
import json
import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy import text

from app.database import engine


EXPORT_FILE = Path("exports/tyre_group_key_audit_safe.xlsx")
TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")


def clean(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def normalize_desc(description: str) -> str:
    desc = clean(description).upper()
    desc = desc.replace("“", '"').replace("”", '"').replace("″", '"')
    desc = desc.replace("–", "-").replace("—", "-")
    desc = desc.replace("STD -", "STD-")
    desc = re.sub(r"\s+", " ", desc)
    return desc


def normalize_size(raw: str) -> str:
    value = clean(raw).upper()
    value = re.sub(r"\s*-\s*", "-", value)
    value = value.replace(" ", "")

    # 10.0020 -> 10.00-20
    if re.fullmatch(r"\d{1,2}\.\d{2}\d{2}", value):
        value = value[:-2] + "-" + value[-2:]

    return value


def extract_tyre_size(description: str) -> str:
    desc = normalize_desc(description)

    patterns = [
        r"\b\d{1,2}\.\d{2}\s*-\s*\d{1,2}\b",
        r"\b\d{1,2}\.\d{2}\d{2}\b",
        r"\b\d{2,3}/\d{2,3}\s*-\s*\d{1,2}\b",
        r"\b\d{1,2}X\d{1,2}\s+\d/\d\s*-\s*\d{1,2}(?:\s+\d/\d)?\b",
        r"\b\d{1,2}X\d{1,2}\d/\d\s*-\s*\d{1,2}(?:\s*\d/\d)?\b",
        r"\b\d{1,2}X\d{1,2}(?:\s+\d/\d)?\s*-\s*\d{1,2}(?:\s+\d/\d)?\b",
        r"\b\d{1,2}X\d{1,2}(?:\s+\d/\d)?\b",
    ]

    for pattern in patterns:
        match = re.search(pattern, desc)
        if match:
            return normalize_size(match.group(0))

    first_token = desc.split()[0] if desc.split() else ""
    return normalize_size(first_token)


def extract_rim_width(description: str) -> str:
    desc = normalize_desc(description)

    # Usually 7.50" / 7.00" / 4.00"
    match = re.search(r'\b(\d{1,2}\.\d{2})\s*"', desc)
    if match:
        return match.group(1)

    return "NO_RIM"


def has_token(desc: str, token: str) -> bool:
    escaped = re.escape(token)
    return re.search(rf"(?<![A-Z0-9]){escaped}(?![A-Z0-9])", desc) is not None


def extract_product_family(description: str) -> str:
    desc = normalize_desc(description)

    families = []
    for token in ["NOR", "EF", "OPTIMA", "OPT", "ULTIMA", "ULT"]:
        if has_token(desc, token):
            if token == "OPTIMA":
                token = "OPT"
            if token == "ULTIMA":
                token = "ULT"
            if token not in families:
                families.append(token)

    return "+".join(families) if families else "NO_FAMILY"


def extract_construction_tokens(description: str) -> str:
    desc = normalize_desc(description)

    tokens = []

    # Order matters: XT+ before XT.
    for token in [
        "BB",
        "LA",
        "XT+",
        "XT",
        "SUP",
        "LGR",
        "AMS",
        "ROV",
        "TRX",
        "PON",
        "COT",
        "RES",
        "PRESS",
        "CURED",
        "O'RING",
        "ORING",
    ]:
        if has_token(desc, token):
            clean_token = token.replace("O'RING", "ORING")
            if clean_token not in tokens:
                tokens.append(clean_token)

    return "+".join(tokens) if tokens else "BASE"


def extract_tread(description: str) -> str:
    desc = normalize_desc(description)

    # TRX before TR.
    for token in ["TRX", "SM", "TR"]:
        if has_token(desc, token):
            return token

    return "NO_TREAD"


def extract_layer(description: str) -> str:
    desc = normalize_desc(description)

    if re.search(r"(?<![A-Z0-9])STD-2L(?![A-Z0-9])", desc) or has_token(desc, "2L"):
        return "2L"

    if re.search(r"(?<![A-Z0-9])STD-3L(?![A-Z0-9])", desc) or has_token(desc, "3L"):
        return "3L"

    return "NO_LAYER"


def extract_color(description: str) -> str:
    desc = normalize_desc(description)

    if has_token(desc, "GREY") or has_token(desc, "GRAY"):
        return "GREY"

    if has_token(desc, "BLACK"):
        return "BLACK"

    if has_token(desc, "NM") or desc.endswith(" NM"):
        return "NM"

    return "NO_COLOR"


def extract_std_flag(description: str) -> str:
    desc = normalize_desc(description)

    if has_token(desc, "STD") or "STD-" in desc or has_token(desc, "STANDARD"):
        return "STD"

    return "NON_STD"


def make_safe_group_key(description: str) -> dict:
    tyre_size = extract_tyre_size(description)
    rim_width = extract_rim_width(description)
    product_family = extract_product_family(description)
    construction_type = extract_construction_tokens(description)
    tread_pattern = extract_tread(description)
    layer = extract_layer(description)
    color = extract_color(description)
    standard_type = extract_std_flag(description)

    parts = {
        "tyre_size": tyre_size,
        "rim_width": rim_width,
        "product_family": product_family,
        "construction_type": construction_type,
        "tread_pattern": tread_pattern,
        "layer": layer,
        "color": color,
        "standard_type": standard_type,
    }

    group_key = "|".join([
        parts["tyre_size"],
        parts["rim_width"],
        parts["product_family"],
        parts["construction_type"],
        parts["tread_pattern"],
        parts["layer"],
        parts["color"],
        parts["standard_type"],
    ])

    parts["group_key"] = group_key
    return parts


def style_sheet(ws, freeze_cell="A2"):
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = freeze_cell
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 55)

    ws.row_dimensions[1].height = 30


with engine.begin() as conn:
    # Backup current generated mapping tables.
    conn.execute(text(f"""
        CREATE TABLE IF NOT EXISTS tyre_process_groups_backup_{TIMESTAMP}
        AS SELECT * FROM tyre_process_groups
    """))

    conn.execute(text(f"""
        CREATE TABLE IF NOT EXISTS tyre_process_group_items_backup_{TIMESTAMP}
        AS SELECT * FROM tyre_process_group_items
    """))

    conn.execute(text("""
        ALTER TABLE tyre_process_groups
        ADD COLUMN IF NOT EXISTS rim_width VARCHAR(64) NOT NULL DEFAULT ''
    """))

    conn.execute(text("""
        ALTER TABLE tyre_process_groups
        ADD COLUMN IF NOT EXISTS product_family VARCHAR(255) NOT NULL DEFAULT ''
    """))

    conn.execute(text("""
        ALTER TABLE tyre_process_groups
        ADD COLUMN IF NOT EXISTS construction_type VARCHAR(255) NOT NULL DEFAULT ''
    """))

    conn.execute(text("""
        ALTER TABLE tyre_process_groups
        ADD COLUMN IF NOT EXISTS standard_type VARCHAR(64) NOT NULL DEFAULT ''
    """))

    conn.execute(text("""
        ALTER TABLE tyre_process_groups
        ADD COLUMN IF NOT EXISTS key_method VARCHAR(64) NOT NULL DEFAULT ''
    """))

    conn.execute(text("""
        ALTER TABLE tyre_process_groups
        ADD COLUMN IF NOT EXISTS key_parts_json TEXT NOT NULL DEFAULT ''
    """))

    conn.execute(text("""
        ALTER TABLE tyre_process_group_items
        ADD COLUMN IF NOT EXISTS old_group_key VARCHAR(255) NOT NULL DEFAULT ''
    """))

    conn.execute(text("""
        ALTER TABLE tyre_process_group_items
        ADD COLUMN IF NOT EXISTS key_parts_json TEXT NOT NULL DEFAULT ''
    """))

    conn.execute(text("""
        ALTER TABLE tyre_item_master
        ADD COLUMN IF NOT EXISTS process_group_key VARCHAR(255) NOT NULL DEFAULT ''
    """))

    rows = conn.execute(text("""
        SELECT
            sap_code,
            description,
            process_group_key AS old_group_key
        FROM tyre_item_master
        WHERE sap_code IS NOT NULL
          AND sap_code <> ''
        ORDER BY sap_code
    """)).mappings().all()

    conn.execute(text("DELETE FROM tyre_process_group_items"))
    conn.execute(text("DELETE FROM tyre_process_groups"))

    group_cache = {}
    item_components = []
    audit_issues = []

    for row in rows:
        sap_code = clean(row["sap_code"])
        description = clean(row["description"])
        old_group_key = clean(row["old_group_key"])

        if not sap_code or not description:
            audit_issues.append({
                "severity": "HIGH",
                "sap_code": sap_code,
                "group_key": "",
                "issue": "Missing SAP code or description",
                "description": description,
            })
            continue

        parts = make_safe_group_key(description)
        group_key = parts["group_key"]

        key_parts_json = json.dumps(parts, ensure_ascii=False)

        if group_key not in group_cache:
            result = conn.execute(
                text("""
                    INSERT INTO tyre_process_groups (
                        group_key,
                        tyre_size,
                        rim_width,
                        product_family,
                        construction_type,
                        pattern,
                        layer,
                        color,
                        standard_type,
                        key_method,
                        key_parts_json,
                        remarks,
                        updated_at
                    )
                    VALUES (
                        :group_key,
                        :tyre_size,
                        :rim_width,
                        :product_family,
                        :construction_type,
                        :pattern,
                        :layer,
                        :color,
                        :standard_type,
                        'SAFE_DESCRIPTION_PARTS_V2',
                        :key_parts_json,
                        '',
                        CURRENT_TIMESTAMP
                    )
                    RETURNING id
                """),
                {
                    "group_key": group_key,
                    "tyre_size": parts["tyre_size"],
                    "rim_width": parts["rim_width"],
                    "product_family": parts["product_family"],
                    "construction_type": parts["construction_type"],
                    "pattern": parts["tread_pattern"],
                    "layer": parts["layer"],
                    "color": parts["color"],
                    "standard_type": parts["standard_type"],
                    "key_parts_json": key_parts_json,
                },
            )
            group_cache[group_key] = result.scalar_one()

        group_id = group_cache[group_key]

        conn.execute(
            text("""
                INSERT INTO tyre_process_group_items (
                    group_id,
                    sap_code,
                    description,
                    old_group_key,
                    key_parts_json
                )
                VALUES (
                    :group_id,
                    :sap_code,
                    :description,
                    :old_group_key,
                    :key_parts_json
                )
            """),
            {
                "group_id": group_id,
                "sap_code": sap_code,
                "description": description,
                "old_group_key": old_group_key,
                "key_parts_json": key_parts_json,
            },
        )

        conn.execute(
            text("""
                UPDATE tyre_item_master
                SET process_group_key = :group_key,
                    updated_at = CURRENT_TIMESTAMP
                WHERE sap_code = :sap_code
            """),
            {
                "group_key": group_key,
                "sap_code": sap_code,
            },
        )

        item_components.append({
            "sap_code": sap_code,
            "description": description,
            "old_group_key": old_group_key,
            **parts,
        })

        for field_name in ["tyre_size", "rim_width", "product_family", "construction_type", "tread_pattern", "layer", "color"]:
            value = parts[field_name]
            if value.startswith("NO_") or value in ("BASE",):
                audit_issues.append({
                    "severity": "REVIEW",
                    "sap_code": sap_code,
                    "group_key": group_key,
                    "issue": f"{field_name} not clearly identified: {value}",
                    "description": description,
                })

    groups = conn.execute(text("""
        SELECT
            g.id,
            g.group_key,
            g.tyre_size,
            g.rim_width,
            g.product_family,
            g.construction_type,
            g.pattern AS tread_pattern,
            g.layer,
            g.color,
            g.standard_type,
            COUNT(i.sap_code) AS sap_count
        FROM tyre_process_groups g
        LEFT JOIN tyre_process_group_items i ON i.group_id = g.id
        GROUP BY
            g.id,
            g.group_key,
            g.tyre_size,
            g.rim_width,
            g.product_family,
            g.construction_type,
            g.pattern,
            g.layer,
            g.color,
            g.standard_type
        ORDER BY COUNT(i.sap_code) DESC, g.group_key
    """)).mappings().all()

    # Add high-count groups for review only.
    for group in groups:
        if group["sap_count"] >= 25:
            audit_issues.append({
                "severity": "REVIEW",
                "sap_code": "",
                "group_key": group["group_key"],
                "issue": f"Large group: {group['sap_count']} SAP codes. Review whether all are truly same tyre.",
                "description": "",
            })

    old_to_new = conn.execute(text("""
        SELECT
            old_group_key,
            COUNT(DISTINCT group_id) AS new_group_count,
            COUNT(sap_code) AS sap_count
        FROM tyre_process_group_items
        WHERE old_group_key IS NOT NULL
          AND old_group_key <> ''
        GROUP BY old_group_key
        ORDER BY COUNT(DISTINCT group_id) DESC, COUNT(sap_code) DESC
    """)).mappings().all()

    group_items = conn.execute(text("""
        SELECT
            g.group_key,
            g.tyre_size,
            g.rim_width,
            g.product_family,
            g.construction_type,
            g.pattern AS tread_pattern,
            g.layer,
            g.color,
            g.standard_type,
            i.sap_code,
            i.description
        FROM tyre_process_group_items i
        JOIN tyre_process_groups g ON g.id = i.group_id
        ORDER BY g.group_key, i.sap_code
    """)).mappings().all()


wb = Workbook()

ws_summary = wb.active
ws_summary.title = "Summary"

ws_groups = wb.create_sheet("Safe Group Keys")
ws_items = wb.create_sheet("SAP Codes By Group")
ws_components = wb.create_sheet("Extracted Components")
ws_audit = wb.create_sheet("Audit Issues")
ws_split = wb.create_sheet("Old To New Split")

ws_summary.append(["Metric", "Value"])
ws_summary.append(["Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
ws_summary.append(["Key Method", "SAFE_DESCRIPTION_PARTS_V2"])
ws_summary.append(["SAP Codes", len(item_components)])
ws_summary.append(["Safe Group Keys", len(groups)])
ws_summary.append(["Audit Review Rows", len(audit_issues)])
ws_summary.append(["Old Backup Timestamp", TIMESTAMP])
ws_summary.append(["Curing/Handling Used In Key", "NO"])

style_sheet(ws_summary)

ws_groups.append([
    "Group Key",
    "Tyre Size",
    "Rim/Width",
    "Product Family",
    "Construction Tokens",
    "Tread",
    "Layer",
    "Color",
    "STD Flag",
    "SAP Count",
])

for row in groups:
    ws_groups.append([
        row["group_key"],
        row["tyre_size"],
        row["rim_width"],
        row["product_family"],
        row["construction_type"],
        row["tread_pattern"],
        row["layer"],
        row["color"],
        row["standard_type"],
        row["sap_count"],
    ])

style_sheet(ws_groups)

ws_items.append([
    "Group Key",
    "Tyre Size",
    "Rim/Width",
    "Product Family",
    "Construction Tokens",
    "Tread",
    "Layer",
    "Color",
    "STD Flag",
    "SAP Code",
    "Description",
])

for row in group_items:
    ws_items.append([
        row["group_key"],
        row["tyre_size"],
        row["rim_width"],
        row["product_family"],
        row["construction_type"],
        row["tread_pattern"],
        row["layer"],
        row["color"],
        row["standard_type"],
        row["sap_code"],
        row["description"],
    ])

style_sheet(ws_items)

ws_components.append([
    "SAP Code",
    "Description",
    "Old Group Key",
    "New Group Key",
    "Tyre Size",
    "Rim/Width",
    "Product Family",
    "Construction Tokens",
    "Tread",
    "Layer",
    "Color",
    "STD Flag",
])

for row in item_components:
    ws_components.append([
        row["sap_code"],
        row["description"],
        row["old_group_key"],
        row["group_key"],
        row["tyre_size"],
        row["rim_width"],
        row["product_family"],
        row["construction_type"],
        row["tread_pattern"],
        row["layer"],
        row["color"],
        row["standard_type"],
    ])

style_sheet(ws_components)

ws_audit.append(["Severity", "SAP Code", "Group Key", "Issue", "Description"])

for row in audit_issues:
    ws_audit.append([
        row["severity"],
        row["sap_code"],
        row["group_key"],
        row["issue"],
        row["description"],
    ])

style_sheet(ws_audit)

ws_split.append(["Old Group Key", "New Group Count", "SAP Count"])

for row in old_to_new:
    ws_split.append([
        row["old_group_key"],
        row["new_group_count"],
        row["sap_count"],
    ])

style_sheet(ws_split)

ws_groups.column_dimensions["A"].width = 55
ws_items.column_dimensions["A"].width = 55
ws_items.column_dimensions["K"].width = 55
ws_components.column_dimensions["B"].width = 55
ws_components.column_dimensions["D"].width = 55
ws_audit.column_dimensions["C"].width = 55
ws_audit.column_dimensions["D"].width = 45
ws_audit.column_dimensions["E"].width = 55
ws_split.column_dimensions["A"].width = 45

wb.save(EXPORT_FILE)

print("Safe tyre group key rebuild completed.")
print("SAP Codes:", len(item_components))
print("Safe Groups:", len(groups))
print("Audit Review Rows:", len(audit_issues))
print("Excel Audit:", EXPORT_FILE.resolve())
print("Backup timestamp:", TIMESTAMP)
