from pathlib import Path
from openpyxl import load_workbook
from sqlalchemy import text
import re

from app.database import engine


SOURCE_FILE = Path("data/local_master/Tire production time with curing cycle.xlsx")


def clean(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def to_number(value) -> float:
    if value is None:
        return 0.0

    if isinstance(value, (int, float)):
        return float(value)

    text_value = clean(value)

    if not text_value:
        return 0.0

    try:
        return float(text_value)
    except ValueError:
        return 0.0


def total_minutes(hours_value, minutes_value) -> float:
    hours = to_number(hours_value)
    minutes = to_number(minutes_value)
    return round((hours * 60) + minutes, 2)


def is_short_cycle(description: str) -> bool:
    desc = f" {clean(description).upper()} "
    return " STD" in desc or " STD-" in desc or "STANDARD" in desc


if not SOURCE_FILE.exists():
    raise SystemExit(f"Source file not found: {SOURCE_FILE}")

wb = load_workbook(SOURCE_FILE, read_only=True, data_only=True)
ws = wb["Sheet2"]

updates = []
skipped = 0

# Sheet2 header:
# Col B = SAP Code
# Col C = Material Description
# Col E = Curing Cycle Hours
# Col F = Curing Cycle Min
# Col G = Handling time
for row in ws.iter_rows(min_row=3, values_only=True):
    sap_code = clean(row[1] if len(row) > 1 else "")
    description = clean(row[2] if len(row) > 2 else "")

    if not re.fullmatch(r"\d{6,14}", sap_code):
        skipped += 1
        continue

    curing_minutes = total_minutes(
        row[4] if len(row) > 4 else None,
        row[5] if len(row) > 5 else None,
    )

    handling_minutes = to_number(row[6] if len(row) > 6 else None)

    if curing_minutes <= 0:
        skipped += 1
        continue

    if is_short_cycle(description):
        normal_curing_minutes = 0
        short_cycle_curing_minutes = curing_minutes
    else:
        normal_curing_minutes = curing_minutes
        short_cycle_curing_minutes = 0

    updates.append({
        "sap_code": sap_code,
        "description": description,
        "normal_curing_minutes": normal_curing_minutes,
        "short_cycle_curing_minutes": short_cycle_curing_minutes,
        "handling_minutes": handling_minutes,
    })

wb.close()

with engine.begin() as conn:
    conn.execute(text("""
        ALTER TABLE tyre_item_master
        ADD COLUMN IF NOT EXISTS normal_curing_minutes NUMERIC(12, 2) NOT NULL DEFAULT 0
    """))

    conn.execute(text("""
        ALTER TABLE tyre_item_master
        ADD COLUMN IF NOT EXISTS short_cycle_curing_minutes NUMERIC(12, 2) NOT NULL DEFAULT 0
    """))

    conn.execute(text("""
        ALTER TABLE tyre_item_master
        ADD COLUMN IF NOT EXISTS handling_minutes NUMERIC(12, 2) NOT NULL DEFAULT 0
    """))

    matched = 0

    for item in updates:
        result = conn.execute(
            text("""
                UPDATE tyre_item_master
                SET normal_curing_minutes = CASE
                        WHEN :normal_curing_minutes > 0 THEN :normal_curing_minutes
                        ELSE normal_curing_minutes
                    END,
                    short_cycle_curing_minutes = CASE
                        WHEN :short_cycle_curing_minutes > 0 THEN :short_cycle_curing_minutes
                        ELSE short_cycle_curing_minutes
                    END,
                    handling_minutes = CASE
                        WHEN :handling_minutes > 0 THEN :handling_minutes
                        ELSE handling_minutes
                    END,
                    updated_at = CURRENT_TIMESTAMP
                WHERE sap_code = :sap_code
            """),
            item,
        )

        matched += result.rowcount or 0

with engine.connect() as conn:
    total = conn.execute(text("SELECT COUNT(*) FROM tyre_item_master")).scalar_one()

    normal_count = conn.execute(text("""
        SELECT COUNT(*)
        FROM tyre_item_master
        WHERE normal_curing_minutes > 0
    """)).scalar_one()

    short_count = conn.execute(text("""
        SELECT COUNT(*)
        FROM tyre_item_master
        WHERE short_cycle_curing_minutes > 0
    """)).scalar_one()

    handling_count = conn.execute(text("""
        SELECT COUNT(*)
        FROM tyre_item_master
        WHERE handling_minutes > 0
    """)).scalar_one()

    samples = conn.execute(text("""
        SELECT
            sap_code,
            description,
            normal_curing_minutes,
            short_cycle_curing_minutes,
            handling_minutes
        FROM tyre_item_master
        WHERE normal_curing_minutes > 0
           OR short_cycle_curing_minutes > 0
        ORDER BY sap_code
        LIMIT 30
    """)).all()

print("Curing source rows read:", len(updates))
print("Skipped rows:", skipped)
print("DB matched / updated:", matched)
print("Total tyre items:", total)
print("Normal curing filled:", normal_count)
print("Short cycle curing filled:", short_count)
print("Handling filled:", handling_count)

print("")
print("Sample:")
for row in samples:
    print(
        row.sap_code,
        "| Normal:", row.normal_curing_minutes,
        "| Short:", row.short_cycle_curing_minutes,
        "| Handling:", row.handling_minutes,
        "|", row.description,
    )
