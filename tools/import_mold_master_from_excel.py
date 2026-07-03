from pathlib import Path
from datetime import datetime
from collections import defaultdict
import re

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy import text

from app.database import engine


SOURCE_DIR = Path("data/local_master")
EXPORT_FILE = Path("exports/mold_master_import_audit.xlsx")


def clean(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def to_int(value) -> int:
    if value is None:
        return 0

    if isinstance(value, (int, float)):
        return int(round(float(value)))

    value_text = clean(value)
    if not value_text:
        return 0

    try:
        return int(round(float(value_text)))
    except ValueError:
        return 0


def norm_header(value) -> str:
    return re.sub(r"[^a-z0-9]+", "", clean(value).lower())


def find_source_file() -> Path:
    candidates = []

    for pattern in [
        "Master File (3).xlsx",
        "Master File*.xlsx",
        "*Master*.xlsx",
    ]:
        candidates.extend(SOURCE_DIR.glob(pattern))

    candidates = sorted(set(candidates))

    if not candidates:
        print("No Master File workbook found in data/local_master")
        print("Copy Master File (3).xlsx into data/local_master and run again.")
        raise SystemExit(1)

    for file_path in candidates:
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            sheet_names = [s.lower().strip() for s in wb.sheetnames]
            wb.close()

            if any("mold" in name and "casing" in name for name in sheet_names):
                return file_path
        except Exception:
            continue

    return candidates[0]


def find_mold_sheet_and_header(wb):
    expected = {
        "keycode": "mold_key_code",
        "numberofmolds": "mold_count",
        "casingtype": "casing_type",
        "numofcasing": "casing_count",
        "numberofcasing": "casing_count",
        "noofcasing": "casing_count",
    }

    preferred_sheets = []
    for ws in wb.worksheets:
        title = ws.title.lower()
        if "mold" in title and "casing" in title:
            preferred_sheets.append(ws)

    worksheets = preferred_sheets + [ws for ws in wb.worksheets if ws not in preferred_sheets]

    for ws in worksheets:
        for row_no, row in enumerate(ws.iter_rows(min_row=1, max_row=40, values_only=True), start=1):
            mapped = {}
            for col_no, value in enumerate(row, start=1):
                key = norm_header(value)
                if key in expected:
                    mapped[expected[key]] = col_no

            if {"mold_key_code", "mold_count", "casing_type", "casing_count"}.issubset(mapped):
                return ws, row_no, mapped

    raise SystemExit("Could not find columns: Key Code, Number of Molds, Casing Type, Num of casing")


source_file = find_source_file()

wb = load_workbook(source_file, read_only=True, data_only=True)
ws, header_row, columns = find_mold_sheet_and_header(wb)

raw_rows = []
aggregated = {}

for excel_row_no, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
    mold_key_code = clean(row[columns["mold_key_code"] - 1] if len(row) >= columns["mold_key_code"] else "")
    mold_count = to_int(row[columns["mold_count"] - 1] if len(row) >= columns["mold_count"] else 0)
    casing_type = clean(row[columns["casing_type"] - 1] if len(row) >= columns["casing_type"] else "")
    casing_count = to_int(row[columns["casing_count"] - 1] if len(row) >= columns["casing_count"] else 0)

    if not mold_key_code:
        continue

    if not casing_type:
        casing_type = "No Casing"

    raw_rows.append({
        "workbook_name": source_file.name,
        "sheet_name": ws.title,
        "row_no": excel_row_no,
        "mold_key_code": mold_key_code,
        "mold_count": mold_count,
        "casing_type": casing_type,
        "casing_count": casing_count,
    })

    if mold_key_code not in aggregated:
        aggregated[mold_key_code] = {
            "mold_key_code": mold_key_code,
            "mold_count": 0,
            "casing_types": set(),
            "casing_count": 0,
            "source_rows": [],
        }

    aggregated[mold_key_code]["mold_count"] += mold_count
    aggregated[mold_key_code]["casing_types"].add(casing_type)
    aggregated[mold_key_code]["casing_count"] += casing_count
    aggregated[mold_key_code]["source_rows"].append(excel_row_no)

wb.close()

master_rows = []

for key_code, item in aggregated.items():
    casing_types = sorted(item["casing_types"])
    casing_type = casing_types[0] if len(casing_types) == 1 else "Mixed: " + ", ".join(casing_types)

    master_rows.append({
        "mold_key_code": key_code,
        "mold_count": item["mold_count"],
        "casing_type": casing_type,
        "casing_count": item["casing_count"],
        "source_row_count": len(item["source_rows"]),
        "source_rows": ", ".join(str(x) for x in item["source_rows"]),
    })

master_rows.sort(key=lambda x: x["mold_key_code"])

duplicate_rows = [row for row in master_rows if row["source_row_count"] > 1]

with engine.begin() as conn:
    conn.execute(text("""
        CREATE TABLE IF NOT EXISTS mold_master (
            id BIGSERIAL PRIMARY KEY,
            mold_key_code VARCHAR(255) NOT NULL UNIQUE,
            mold_count INTEGER NOT NULL DEFAULT 0,
            casing_type VARCHAR(255) NOT NULL DEFAULT '',
            casing_count INTEGER NOT NULL DEFAULT 0,
            status VARCHAR(32) NOT NULL DEFAULT 'Active',
            remarks TEXT NOT NULL DEFAULT '',
            source_file VARCHAR(255) NOT NULL DEFAULT '',
            source_sheet VARCHAR(255) NOT NULL DEFAULT '',
            source_rows TEXT NOT NULL DEFAULT '',
            created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
    """))

    conn.execute(text("""
        CREATE TABLE IF NOT EXISTS mold_master_source_rows (
            id BIGSERIAL PRIMARY KEY,
            import_batch_id VARCHAR(64) NOT NULL,
            workbook_name VARCHAR(255) NOT NULL,
            sheet_name VARCHAR(255) NOT NULL,
            row_no INTEGER NOT NULL,
            mold_key_code VARCHAR(255) NOT NULL,
            mold_count INTEGER NOT NULL DEFAULT 0,
            casing_type VARCHAR(255) NOT NULL DEFAULT '',
            casing_count INTEGER NOT NULL DEFAULT 0,
            created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
    """))

    import_batch_id = datetime.now().strftime("%Y%m%d_%H%M%S")

    conn.execute(text("DELETE FROM mold_master_source_rows"))

    for row in raw_rows:
        conn.execute(
            text("""
                INSERT INTO mold_master_source_rows (
                    import_batch_id,
                    workbook_name,
                    sheet_name,
                    row_no,
                    mold_key_code,
                    mold_count,
                    casing_type,
                    casing_count
                )
                VALUES (
                    :import_batch_id,
                    :workbook_name,
                    :sheet_name,
                    :row_no,
                    :mold_key_code,
                    :mold_count,
                    :casing_type,
                    :casing_count
                )
            """),
            {
                "import_batch_id": import_batch_id,
                **row,
            },
        )

    conn.execute(text("DELETE FROM mold_master"))

    for row in master_rows:
        conn.execute(
            text("""
                INSERT INTO mold_master (
                    mold_key_code,
                    mold_count,
                    casing_type,
                    casing_count,
                    status,
                    remarks,
                    source_file,
                    source_sheet,
                    source_rows,
                    updated_at
                )
                VALUES (
                    :mold_key_code,
                    :mold_count,
                    :casing_type,
                    :casing_count,
                    'Active',
                    '',
                    :source_file,
                    :source_sheet,
                    :source_rows,
                    CURRENT_TIMESTAMP
                )
            """),
            {
                "mold_key_code": row["mold_key_code"],
                "mold_count": row["mold_count"],
                "casing_type": row["casing_type"],
                "casing_count": row["casing_count"],
                "source_file": source_file.name,
                "source_sheet": ws.title,
                "source_rows": row["source_rows"],
            },
        )

    db_summary = conn.execute(text("""
        SELECT
            COUNT(*) AS mold_key_count,
            COALESCE(SUM(mold_count), 0) AS total_mold_count,
            COALESCE(SUM(casing_count), 0) AS total_casing_count
        FROM mold_master
    """)).mappings().one()

    casing_summary = conn.execute(text("""
        SELECT
            casing_type,
            COUNT(*) AS mold_key_count,
            COALESCE(SUM(mold_count), 0) AS mold_count,
            COALESCE(SUM(casing_count), 0) AS casing_count
        FROM mold_master
        GROUP BY casing_type
        ORDER BY mold_count DESC, casing_type
    """)).mappings().all()


# Export audit workbook.
out = Workbook()
ws_summary = out.active
ws_summary.title = "Summary"

ws_master = out.create_sheet("Mold Master")
ws_casing = out.create_sheet("Casing Summary")
ws_duplicates = out.create_sheet("Duplicate Mold Keys")
ws_raw = out.create_sheet("Raw Source Rows")


def style_sheet(sheet):
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col_idx, column_cells in enumerate(sheet.columns, start=1):
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 45)


ws_summary.append(["Metric", "Value"])
ws_summary.append(["Source File", source_file.name])
ws_summary.append(["Source Sheet", ws.title])
ws_summary.append(["Header Row", header_row])
ws_summary.append(["Raw Mold Rows", len(raw_rows)])
ws_summary.append(["Unique Mold Keys", db_summary["mold_key_count"]])
ws_summary.append(["Total Mold Count", db_summary["total_mold_count"]])
ws_summary.append(["Total Casing Count", db_summary["total_casing_count"]])
ws_summary.append(["Duplicate Key Count", len(duplicate_rows)])
style_sheet(ws_summary)

ws_master.append([
    "Mold Key Code",
    "Mold Count",
    "Casing Type",
    "Casing Count",
    "Source Row Count",
    "Source Rows",
])

for row in master_rows:
    ws_master.append([
        row["mold_key_code"],
        row["mold_count"],
        row["casing_type"],
        row["casing_count"],
        row["source_row_count"],
        row["source_rows"],
    ])

style_sheet(ws_master)

ws_casing.append(["Casing Type", "Mold Key Count", "Mold Count", "Casing Count"])

for row in casing_summary:
    ws_casing.append([
        row["casing_type"],
        row["mold_key_count"],
        row["mold_count"],
        row["casing_count"],
    ])

style_sheet(ws_casing)

ws_duplicates.append([
    "Mold Key Code",
    "Aggregated Mold Count",
    "Casing Type",
    "Aggregated Casing Count",
    "Source Row Count",
    "Source Rows",
])

for row in duplicate_rows:
    ws_duplicates.append([
        row["mold_key_code"],
        row["mold_count"],
        row["casing_type"],
        row["casing_count"],
        row["source_row_count"],
        row["source_rows"],
    ])

style_sheet(ws_duplicates)

ws_raw.append([
    "Workbook",
    "Sheet",
    "Row No",
    "Mold Key Code",
    "Mold Count",
    "Casing Type",
    "Casing Count",
])

for row in raw_rows:
    ws_raw.append([
        row["workbook_name"],
        row["sheet_name"],
        row["row_no"],
        row["mold_key_code"],
        row["mold_count"],
        row["casing_type"],
        row["casing_count"],
    ])

style_sheet(ws_raw)

ws_master.column_dimensions["A"].width = 34
ws_raw.column_dimensions["D"].width = 34
ws_duplicates.column_dimensions["A"].width = 34

out.save(EXPORT_FILE)

print("Mold Master import completed.")
print("Source:", source_file)
print("Sheet:", ws.title)
print("Raw mold rows:", len(raw_rows))
print("Unique mold keys:", db_summary["mold_key_count"])
print("Total mold count:", db_summary["total_mold_count"])
print("Total casing count:", db_summary["total_casing_count"])
print("Duplicate key count:", len(duplicate_rows))
print("Audit Excel:", EXPORT_FILE.resolve())

print("")
print("Casing Summary:")
for row in casing_summary:
    print(
        row["casing_type"],
        "| Keys:", row["mold_key_count"],
        "| Molds:", row["mold_count"],
        "| Casings:", row["casing_count"],
    )
