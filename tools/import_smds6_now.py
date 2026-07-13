import sys
import re
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook
from sqlalchemy import create_engine, text


DEFAULT_EXCEL = r"C:\MY\Laugfs\Laugfs_App\data_sources\SMDS6.xlsx"


def q(name):
    return '"' + str(name).replace('"', '""') + '"'


def load_db_url():
    env = Path(".env")
    if env.exists():
        for line in env.read_text(encoding="utf-8").splitlines():
            if line.strip().startswith("DATABASE_URL="):
                return line.split("=", 1)[1].strip()
    return "postgresql+psycopg://postgres:adminkusa@localhost:5434/factory_planner"


def norm(v):
    if v is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(v).strip().lower())


def clean(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def to_int(v, default=0):
    try:
        if v is None or v == "":
            return default
        return int(float(v))
    except Exception:
        return default


ALIASES = {
    "sap_code": ["sapcode", "sap", "materialcode", "itemcode", "code"],
    "description": ["materialdescription", "description", "itemdescription", "tyredescription"],
    "key_code": ["keycode", "moldkeycode", "mouldkeycode", "moldkey", "mouldkey", "groupkey"],
    "casing": ["casing", "casingtype", "casingcode"],
    "line": ["line", "linename", "machine", "machineline", "productionline", "press", "oven"],
    "day_plan": ["dayplan", "daycapacity", "dayshift"],
    "night_plan": ["nightplan", "nightcapacity", "nightshift"],
    "total_plan": ["totalplan", "dailycapacity", "totalcapacity", "capacity"],
    "status": ["approvalstatus", "planningmanagerapprovalstatus", "status", "approved"],
    "product_group": ["productgroup", "product", "category"],
    "tyre_size": ["tyresize", "size"],
    "weight": ["weight", "tyreweight"],
    "curing_time": ["curingtime", "normalcuringtime", "curetime"],
}


def get_val(row, key):
    for a in ALIASES[key]:
        if a in row:
            return clean(row[a])
    return None


def detect_header(rows):
    best = None
    best_score = 0

    for idx, row in enumerate(rows[:120]):
        headers = [norm(x) for x in row]
        score = 0
        for aliases in ALIASES.values():
            if any(a in headers for a in aliases):
                score += 1

        has_code = any(a in headers for a in ALIASES["sap_code"])
        if has_code and score > best_score:
            best = idx
            best_score = score

    return best


def read_excel(path):
    wb = load_workbook(path, data_only=True, read_only=True)
    output = {}

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        header_idx = detect_header(rows)
        if header_idx is None:
            continue

        headers = [norm(x) for x in rows[header_idx]]

        for row_no, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
            raw = {}
            for i, cell in enumerate(row):
                if i < len(headers) and headers[i]:
                    raw[headers[i]] = clean(cell)

            sap = get_val(raw, "sap_code")
            desc = get_val(raw, "description")

            if not sap:
                continue

            day_plan = to_int(get_val(raw, "day_plan"), 0)
            night_plan = to_int(get_val(raw, "night_plan"), 0)
            total_plan = get_val(raw, "total_plan")
            total_plan = to_int(total_plan, day_plan + night_plan)

            output[str(sap).strip()] = {
                "sap_code": str(sap).strip(),
                "material_description": desc,
                "key_code": get_val(raw, "key_code"),
                "casing_type": get_val(raw, "casing"),
                "line_name": get_val(raw, "line"),
                "day_plan": day_plan,
                "night_plan": night_plan,
                "total_plan": total_plan,
                "planning_manager_approval_status": get_val(raw, "status") or "APPROVED",
                "product_group": get_val(raw, "product_group"),
                "tyre_size": get_val(raw, "tyre_size"),
                "weight": get_val(raw, "weight"),
                "curing_time": get_val(raw, "curing_time"),
                "source_sheet_name": ws.title,
                "source_row_number": row_no,
            }

    return list(output.values())


def table_exists(conn, table):
    return conn.execute(text("""
        SELECT EXISTS (
            SELECT 1 FROM information_schema.tables
            WHERE table_schema='public' AND table_name=:t
        )
    """), {"t": table}).scalar()


def cols(conn, table):
    return [r[0] for r in conn.execute(text("""
        SELECT column_name
        FROM information_schema.columns
        WHERE table_schema='public' AND table_name=:t
        ORDER BY ordinal_position
    """), {"t": table}).fetchall()]


def insert_dynamic(conn, table, records):
    if not table_exists(conn, table):
        print(f"{table}: table not found")
        return 0

    table_cols = cols(conn, table)
    table_col_set = set(table_cols)

    mappings = {
        "sap_code": ["sap_code", "material_code", "item_code", "code"],
        "material_description": ["material_description", "item_description", "description", "tyre_description"],
        "key_code": ["key_code", "mold_key_code", "mould_key_code", "mold_key"],
        "casing_type": ["casing_type", "casing", "casing_code"],
        "line_name": ["line_name", "production_line", "machine_line_name", "line"],
        "day_plan": ["day_plan", "day_capacity"],
        "night_plan": ["night_plan", "night_capacity"],
        "total_plan": ["total_plan", "daily_capacity", "capacity", "total_capacity"],
        "planning_manager_approval_status": ["planning_manager_approval_status", "approval_status", "status"],
        "product_group": ["product_group"],
        "tyre_size": ["tyre_size", "size"],
        "weight": ["weight", "tyre_weight"],
        "curing_time": ["curing_time", "normal_curing_time"],
        "source_sheet_name": ["source_sheet_name", "source_sheet"],
        "source_row_number": ["source_row_number", "source_row"],
    }

    inserted = 0

    for rec in records:
        data = {}

        for src, targets in mappings.items():
            for t in targets:
                if t in table_col_set:
                    data[t] = rec.get(src)
                    break

        if "created_at" in table_col_set:
            data["created_at"] = datetime.now()
        if "updated_at" in table_col_set:
            data["updated_at"] = datetime.now()
        if "is_active" in table_col_set:
            data["is_active"] = True

        if not data:
            continue

        col_names = list(data.keys())
        sql = f'INSERT INTO {q(table)} ({", ".join(q(c) for c in col_names)}) VALUES ({", ".join(":" + c for c in col_names)})'

        try:
            conn.execute(text(sql), data)
            inserted += 1
        except Exception as e:
            print(f"SKIP {table} {rec.get('sap_code')}: {e}")

    return inserted


def seed_mold_master(conn, records):
    if not table_exists(conn, "mold_master"):
        return 0

    c = set(cols(conn, "mold_master"))
    keys = sorted({str(r.get("key_code")).strip() for r in records if r.get("key_code")})
    inserted = 0

    for key in keys:
        data = {}

        for field in ["mold_key_code", "mould_key_code", "key_code", "mold_code", "code"]:
            if field in c:
                data[field] = key
                break

        for field in ["description", "mold_description", "remarks"]:
            if field in c:
                data[field] = key
                break

        for field in ["mold_count", "total_mold_count", "total_molds", "total_count"]:
            if field in c:
                data[field] = 1

        for field in ["production_mold_count", "breakdown_mold_count"]:
            if field in c:
                data[field] = 0

        if "is_active" in c:
            data["is_active"] = True
        if "created_at" in c:
            data["created_at"] = datetime.now()
        if "updated_at" in c:
            data["updated_at"] = datetime.now()

        if data:
            names = list(data.keys())
            sql = f'INSERT INTO "mold_master" ({", ".join(q(x) for x in names)}) VALUES ({", ".join(":" + x for x in names)})'
            try:
                conn.execute(text(sql), data)
                inserted += 1
            except Exception as e:
                print("SKIP mold", key, e)

    return inserted


def seed_casing_master(conn, records):
    if not table_exists(conn, "casing_master"):
        return 0

    c = set(cols(conn, "casing_master"))
    casings = sorted({str(r.get("casing_type")).strip() for r in records if r.get("casing_type")})
    inserted = 0

    for casing in casings:
        data = {}

        for field in ["casing_type", "casing_code", "casing_name", "code"]:
            if field in c:
                data[field] = casing
                break

        for field in ["description", "remarks"]:
            if field in c:
                data[field] = casing
                break

        for field in ["available_casing_count", "casing_count", "total_casing_count", "total_count"]:
            if field in c:
                data[field] = 1

        for field in ["production_casing_count", "breakdown_casing_count"]:
            if field in c:
                data[field] = 0

        if "is_active" in c:
            data["is_active"] = True
        if "created_at" in c:
            data["created_at"] = datetime.now()
        if "updated_at" in c:
            data["updated_at"] = datetime.now()

        if data:
            names = list(data.keys())
            sql = f'INSERT INTO "casing_master" ({", ".join(q(x) for x in names)}) VALUES ({", ".join(":" + x for x in names)})'
            try:
                conn.execute(text(sql), data)
                inserted += 1
            except Exception as e:
                print("SKIP casing", casing, e)

    return inserted


def seed_line_cavities(conn):
    if not table_exists(conn, "production_lines") or not table_exists(conn, "production_line_cavities"):
        return 0

    line_cols = set(cols(conn, "production_lines"))
    cav_cols = set(cols(conn, "production_line_cavities"))

    line_name_col = None
    for x in ["line_name", "name", "production_line", "machine_line_name"]:
        if x in line_cols:
            line_name_col = x
            break

    if not line_name_col:
        return 0

    lines = conn.execute(text(f'SELECT id, {q(line_name_col)} AS line_name FROM production_lines ORDER BY id')).mappings().fetchall()
    inserted = 0

    for line in lines:
        data = {}

        if "production_line_id" in cav_cols:
            data["production_line_id"] = line["id"]
        if "line_id" in cav_cols:
            data["line_id"] = line["id"]
        if "line_name" in cav_cols:
            data["line_name"] = line["line_name"]
        if "production_line" in cav_cols:
            data["production_line"] = line["line_name"]

        for x in ["cavity_number", "cavity_no", "position_no", "position_number"]:
            if x in cav_cols:
                data[x] = 1
                break

        for x in ["cavity_name", "position_name", "name"]:
            if x in cav_cols:
                data[x] = f"{line['line_name']}-C1"
                break

        if "status" in cav_cols:
            data["status"] = "FREE"
        if "is_active" in cav_cols:
            data["is_active"] = True
        if "created_at" in cav_cols:
            data["created_at"] = datetime.now()
        if "updated_at" in cav_cols:
            data["updated_at"] = datetime.now()

        if data:
            names = list(data.keys())
            sql = f'INSERT INTO "production_line_cavities" ({", ".join(q(x) for x in names)}) VALUES ({", ".join(":" + x for x in names)})'
            try:
                conn.execute(text(sql), data)
                inserted += 1
            except Exception as e:
                print("SKIP cavity", line["line_name"], e)

    return inserted


def main():
    excel = DEFAULT_EXCEL
    replace = "--replace" in sys.argv
    seed_cavities = "--seed-line-cavities" in sys.argv

    for arg in sys.argv[1:]:
        if arg.lower().endswith((".xlsx", ".xlsm")):
            excel = arg

    print("Excel:", excel)
    records = read_excel(excel)
    print("Rows detected:", len(records))

    if not records:
        raise SystemExit("No rows detected. Check Excel file/header.")

    engine = create_engine(load_db_url())

    with engine.begin() as conn:
        if replace:
            for t in ["smds", "mold_master", "casing_master", "casing_units", "production_line_cavities"]:
                if table_exists(conn, t):
                    conn.execute(text(f'TRUNCATE TABLE {q(t)} RESTART IDENTITY CASCADE'))
                    print("Cleared:", t)

        print("smds inserted:", insert_dynamic(conn, "smds", records))
        print("mold_master inserted:", seed_mold_master(conn, records))
        print("casing_master inserted:", seed_casing_master(conn, records))

        if seed_cavities:
            print("production_line_cavities inserted:", seed_line_cavities(conn))

    print("\nFINAL COUNTS")
    with engine.connect() as conn:
        for t in ["smds", "mold_master", "casing_master", "production_lines", "production_line_cavities", "mpps_sap_stock_items"]:
            try:
                print(t, conn.execute(text(f'SELECT COUNT(*) FROM {q(t)}')).scalar())
            except Exception as e:
                print(t, "ERROR", e)


if __name__ == "__main__":
    main()
