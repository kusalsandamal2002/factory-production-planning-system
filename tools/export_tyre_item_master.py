from pathlib import Path
import csv

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from sqlalchemy import text

from app.database import engine


export_dir = Path("exports")
export_dir.mkdir(exist_ok=True)

csv_path = export_dir / "tyre_item_master_sap_description.csv"
xlsx_path = export_dir / "tyre_item_master_sap_description.xlsx"

with engine.connect() as conn:
    rows = conn.execute(text("""
        SELECT sap_code, description
        FROM tyre_item_master
        ORDER BY sap_code
    """)).all()

print("Tyre items found:", len(rows))

with csv_path.open("w", newline="", encoding="utf-8-sig") as file:
    writer = csv.writer(file)
    writer.writerow(["SAP Code", "Description"])

    for row in rows:
        writer.writerow([row.sap_code, row.description])

wb = Workbook()
ws = wb.active
ws.title = "Tyre Item Master"

ws.append(["SAP Code", "Description"])

for row in rows:
    ws.append([row.sap_code, row.description])

header_fill = PatternFill("solid", fgColor="EAF1FF")
header_font = Font(bold=True, color="0F172A")
thin = Side(style="thin", color="D9E2EF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border
    cell.alignment = Alignment(horizontal="center", vertical="center")

for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(vertical="center")

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions
ws.column_dimensions["A"].width = 18
ws.column_dimensions["B"].width = 65

wb.save(xlsx_path)

print("CSV exported:", csv_path.resolve())
print("Excel exported:", xlsx_path.resolve())
print("Done.")
