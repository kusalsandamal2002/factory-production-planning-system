from pathlib import Path
from openpyxl import load_workbook
from sqlalchemy import text
import re

from app.database import engine


SOURCE_FILE = Path("data/local_master/Tire production time with curing cycle.xlsx")


def clean(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def to_number(value) -> float:
    if value is None:
        return 0.0

    if isinstance(value, (int, float)):
        return float(value)

    value_text = clean(value)
    if not value_text:
        return 0.0

    try:
        return float(value_text)
    except ValueError:
        return 0.0


def total_minutes(hours_value, minutes_value) -> float:
    hours = to_number(hours_value)
    minutes = to_number(minutes_value)
    return round((hours * 60) + minutes, 2)


if not SOURCE_FILE.exists():
    raise SystemExit(f"Source file not found: {SOURCE_FILE}")

wb = load_workbook(SOURCE_FILE, read_only=True, data_only=True)
ws = wb["Sheet2"]

updates = []
skipped = 0

# Sheet2 source:
# B = SAP Code
# C = Material Description
# E = Curing Cycle Hours
# F = Curing Cycle Minutes
# G = Handling time
for row in ws.iter_rows(min_row=3, values_only=True):
    sap_code = clean(row[1] if len(row) > 1 else "")

    if not re.fullmatch(r"\d{6,14}", sap_code):
        skipped += 1
        continue

    normal_curing_minutes = total_minutes(
        row[4] if len(row) > 4 else None,
        row[5] if len(row) > 5 else None,
    )

    handling_minutes = to_number(row[6] if len(row) > 6 else None)

    if normal_curing_minutes <= 0 and handling_minutes <= 0:
        skipped += 1
        continue

    updates.append({
        "sap_code": sap_code,
        "normal_curing_minutes": normal_curing_minutes,
        "short_cycle_curing_minutes": 0,
        "handling_minutes": handling_minutes,
    })

wb.close()

with engine.begin() as conn:
    conn.execute(text("""
        ALTER TABLE tyre_item_master
        ADD COLUMN IF NOT EXISTS normal_curing_minutes NUMERIC(12, 2) NOT NULL DEFAULT 0
    """))

    conn.execute(text("""
        ALTER TABLE tyre_item_master
        ADD COLUMN IF NOT EXISTS short_cycle_curing_minutes NUMERIC(12, 2) NOT NULL DEFAULT 0
    """))

    conn.execute(text("""
        ALTER TABLE tyre_item_master
        ADD COLUMN IF NOT EXISTS handling_minutes NUMERIC(12, 2) NOT NULL DEFAULT 0
    """))

    # Source-only refresh:
    # Source file eke nathi ewa DB eke 0 thiyenawa. UI eke 0 display wenne "-" widiyata.
    conn.execute(text("""
        UPDATE tyre_item_master
        SET normal_curing_minutes = 0,
            short_cycle_curing_minutes = 0,
            handling_minutes = 0
    """))

    matched = 0

    for item in updates:
        result = conn.execute(
            text("""
                UPDATE tyre_item_master
                SET normal_curing_minutes = :normal_curing_minutes,
                    short_cycle_curing_minutes = 0,
                    handling_minutes = :handling_minutes,
                    updated_at = CURRENT_TIMESTAMP
                WHERE sap_code = :sap_code
            """),
            item,
        )
        matched += result.rowcount or 0

with engine.connect() as conn:
    total = conn.execute(text("SELECT COUNT(*) FROM tyre_item_master")).scalar_one()

    normal_count = conn.execute(text("""
        SELECT COUNT(*) FROM tyre_item_master WHERE normal_curing_minutes > 0
    """)).scalar_one()

    short_count = conn.execute(text("""
        SELECT COUNT(*) FROM tyre_item_master WHERE short_cycle_curing_minutes > 0
    """)).scalar_one()

    handling_count = conn.execute(text("""
        SELECT COUNT(*) FROM tyre_item_master WHERE handling_minutes > 0
    """)).scalar_one()

print("Source file:", SOURCE_FILE)
print("Source rows read:", len(updates))
print("Skipped rows:", skipped)
print("DB matched / updated:", matched)
print("Total tyre items:", total)
print("Normal curing from source:", normal_count)
print("Short cycle from source:", short_count)
print("Handling from source:", handling_count)
print("Missing values will display as '-' in UI.")
