from pathlib import Path
from datetime import datetime
import json
import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy import text

from app.database import engine


EXPORT_FILE = Path("exports/tyre_group_key_audit_with_aperture_black_default.xlsx")
TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")


def clean(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def normalize_desc(description: str) -> str:
    desc = clean(description).upper()
    desc = desc.replace("“", '"').replace("”", '"').replace("″", '"')
    desc = desc.replace("–", "-").replace("—", "-")
    desc = desc.replace("STD -", "STD-")
    desc = desc.replace("GUARD ON", "GUARD-ON")
    desc = desc.replace("GURD-ON", "GUARD-ON")
    desc = desc.replace("GURD ON", "GUARD-ON")
    desc = desc.replace("CURED ON", "CURED-ON")
    desc = re.sub(r"\s+", " ", desc)
    return desc


def normalize_size(raw: str) -> str:
    value = clean(raw).upper()
    value = re.sub(r"\s*-\s*", "-", value)
    value = value.replace(" ", "")

    if re.fullmatch(r"\d{1,2}\.\d{2}\d{2}", value):
        value = value[:-2] + "-" + value[-2:]

    return value


def has_token(desc: str, token: str) -> bool:
    escaped = re.escape(token)
    return re.search(rf"(?<![A-Z0-9]){escaped}(?![A-Z0-9])", desc) is not None


def extract_tyre_size(description: str) -> str:
    desc = normalize_desc(description)

    patterns = [
        r"\b\d{1,2}\.\d{2}\s*-\s*\d{1,2}\b",
        r"\b\d{1,2}\.\d{2}\d{2}\b",
        r"\b\d{2,3}/\d{2,3}\s*-\s*\d{1,2}\b",
        r"\b\d{1,2}X\d{1,2}\s+\d/\d\s*-\s*\d{1,2}(?:\s+\d/\d)?\b",
        r"\b\d{1,2}X\d{1,2}\d/\d\s*-\s*\d{1,2}(?:\s*\d/\d)?\b",
        r"\b\d{1,2}X\d{1,2}(?:\s+\d/\d)?\s*-\s*\d{1,2}(?:\s+\d/\d)?\b",
        r"\b\d{1,2}X\d{1,2}(?:\s+\d/\d)?\b",
    ]

    for pattern in patterns:
        match = re.search(pattern, desc)
        if match:
            return normalize_size(match.group(0))

    parts = desc.split()
    return normalize_size(parts[0]) if parts else "NO_SIZE"


def extract_rim_width(description: str) -> str:
    desc = normalize_desc(description)
    match = re.search(r'\b(\d{1,2}\.\d{2})\s*"', desc)
    if match:
        return match.group(1)
    return "NO_RIM"


def extract_aperture(description: str) -> str:
    desc = normalize_desc(description)

    if has_token(desc, "PLAIN") or has_token(desc, "PLN"):
        return "PLAIN"

    if has_token(desc, "SINGLE") or has_token(desc, "SINGAL") or has_token(desc, "SGL"):
        return "SINGLE"

    if has_token(desc, "TWO") or has_token(desc, "DOUBLE") or has_token(desc, "DBL"):
        return "TWO"

    return "UNKNOWN"


def extract_tread(description: str) -> str:
    desc = normalize_desc(description)

    if "GUARD-ON" in desc:
        return "GUARD-ON"

    if "CURED-ON" in desc:
        return "CURED-ON"

    for token in ["SKS", "TRX", "SM", "TR", "OPTIMA", "OPT", "ULTIMA", "ULT"]:
        if has_token(desc, token):
            if token == "OPTIMA":
                return "OPT"
            if token == "ULTIMA":
                return "ULT"
            return token

    return "NO_TREAD"


def extract_layer(description: str) -> str:
    desc = normalize_desc(description)

    if re.search(r"(?<![A-Z0-9])STD-2L(?![A-Z0-9])", desc) or has_token(desc, "2L"):
        return "2L"

    if re.search(r"(?<![A-Z0-9])STD-3L(?![A-Z0-9])", desc) or has_token(desc, "3L"):
        return "3L"

    return "NO_LAYER"


def extract_color(description: str) -> str:
    desc = normalize_desc(description)

    if has_token(desc, "GREY") or has_token(desc, "GRAY"):
        return "GREY"

    if has_token(desc, "BLACK"):
        return "BLACK"

    if has_token(desc, "NM") or desc.endswith(" NM"):
        return "NM"

    return "BLACK"


def make_group_key(description: str) -> dict:
    parts = {
        "tyre_size": extract_tyre_size(description),
        "rim_width": extract_rim_width(description),
        "aperture_type": extract_aperture(description),
        "tread_pattern": extract_tread(description),
        "layer": extract_layer(description),
        "color": extract_color(description),
    }

    parts["group_key"] = "|".join([
        parts["tyre_size"],
        parts["rim_width"],
        parts["aperture_type"],
        parts["tread_pattern"],
        parts["layer"],
        parts["color"],
    ])

    return parts


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 55)


with engine.begin() as conn:
    conn.execute(text(f"""
        CREATE TABLE IF NOT EXISTS tyre_process_groups_backup_aperture_{TIMESTAMP}
        AS SELECT * FROM tyre_process_groups
    """))

    conn.execute(text(f"""
        CREATE TABLE IF NOT EXISTS tyre_process_group_items_backup_aperture_{TIMESTAMP}
        AS SELECT * FROM tyre_process_group_items
    """))

    conn.execute(text("""
        ALTER TABLE tyre_process_groups
        ADD COLUMN IF NOT EXISTS rim_width VARCHAR(64) NOT NULL DEFAULT ''
    """))

    conn.execute(text("""
        ALTER TABLE tyre_process_groups
        ADD COLUMN IF NOT EXISTS aperture_type VARCHAR(64) NOT NULL DEFAULT ''
    """))

    conn.execute(text("""
        ALTER TABLE tyre_process_groups
        ADD COLUMN IF NOT EXISTS key_method VARCHAR(64) NOT NULL DEFAULT ''
    """))

    conn.execute(text("""
        ALTER TABLE tyre_process_groups
        ADD COLUMN IF NOT EXISTS key_parts_json TEXT NOT NULL DEFAULT ''
    """))

    conn.execute(text("""
        ALTER TABLE tyre_process_group_items
        ADD COLUMN IF NOT EXISTS key_parts_json TEXT NOT NULL DEFAULT ''
    """))

    conn.execute(text("""
        ALTER TABLE tyre_item_master
        ADD COLUMN IF NOT EXISTS process_group_key VARCHAR(255) NOT NULL DEFAULT ''
    """))

    rows = conn.execute(text("""
        SELECT sap_code, description
        FROM tyre_item_master
        WHERE sap_code IS NOT NULL
          AND sap_code <> ''
        ORDER BY sap_code
    """)).mappings().all()

    conn.execute(text("DELETE FROM tyre_process_group_items"))
    conn.execute(text("DELETE FROM tyre_process_groups"))

    group_cache = {}
    item_rows = []
    audit_rows = []

    for row in rows:
        sap_code = clean(row["sap_code"])
        description = clean(row["description"])

        if not sap_code or not description:
            continue

        parts = make_group_key(description)
        group_key = parts["group_key"]
        key_parts_json = json.dumps(parts, ensure_ascii=False)

        if group_key not in group_cache:
            result = conn.execute(
                text("""
                    INSERT INTO tyre_process_groups (
                        group_key,
                        tyre_size,
                        rim_width,
                        aperture_type,
                        pattern,
                        layer,
                        color,
                        key_method,
                        key_parts_json,
                        remarks,
                        updated_at
                    )
                    VALUES (
                        :group_key,
                        :tyre_size,
                        :rim_width,
                        :aperture_type,
                        :pattern,
                        :layer,
                        :color,
                        'SIZE_RIM_APERTURE_TREAD_LAYER_COLOR_V5_BLACK_DEFAULT',
                        :key_parts_json,
                        '',
                        CURRENT_TIMESTAMP
                    )
                    RETURNING id
                """),
                {
                    "group_key": group_key,
                    "tyre_size": parts["tyre_size"],
                    "rim_width": parts["rim_width"],
                    "aperture_type": parts["aperture_type"],
                    "pattern": parts["tread_pattern"],
                    "layer": parts["layer"],
                    "color": parts["color"],
                    "key_parts_json": key_parts_json,
                },
            )
            group_cache[group_key] = result.scalar_one()

        group_id = group_cache[group_key]

        conn.execute(
            text("""
                INSERT INTO tyre_process_group_items (
                    group_id,
                    sap_code,
                    description,
                    key_parts_json
                )
                VALUES (
                    :group_id,
                    :sap_code,
                    :description,
                    :key_parts_json
                )
            """),
            {
                "group_id": group_id,
                "sap_code": sap_code,
                "description": description,
                "key_parts_json": key_parts_json,
            },
        )

        conn.execute(
            text("""
                UPDATE tyre_item_master
                SET process_group_key = :group_key,
                    updated_at = CURRENT_TIMESTAMP
                WHERE sap_code = :sap_code
            """),
            {
                "group_key": group_key,
                "sap_code": sap_code,
            },
        )

        item_rows.append({
            "sap_code": sap_code,
            "description": description,
            **parts,
        })

        for field_name in ["tyre_size", "rim_width", "aperture_type", "tread_pattern", "layer", "color"]:
            value = parts[field_name]
            if value.startswith("NO_") or value == "UNKNOWN":
                audit_rows.append({
                    "sap_code": sap_code,
                    "group_key": group_key,
                    "issue": f"{field_name} not clearly identified: {value}",
                    "description": description,
                })

    groups = conn.execute(text("""
        SELECT
            g.group_key,
            g.tyre_size,
            g.rim_width,
            g.aperture_type,
            g.pattern AS tread_pattern,
            g.layer,
            g.color,
            COUNT(i.sap_code) AS sap_count
        FROM tyre_process_groups g
        LEFT JOIN tyre_process_group_items i ON i.group_id = g.id
        GROUP BY
            g.id,
            g.group_key,
            g.tyre_size,
            g.rim_width,
            g.aperture_type,
            g.pattern,
            g.layer,
            g.color
        ORDER BY COUNT(i.sap_code) DESC, g.group_key
    """)).mappings().all()

    group_items = conn.execute(text("""
        SELECT
            g.group_key,
            g.tyre_size,
            g.rim_width,
            g.aperture_type,
            g.pattern AS tread_pattern,
            g.layer,
            g.color,
            i.sap_code,
            i.description
        FROM tyre_process_group_items i
        JOIN tyre_process_groups g ON g.id = i.group_id
        ORDER BY g.group_key, i.sap_code
    """)).mappings().all()


wb = Workbook()
ws_summary = wb.active
ws_summary.title = "Summary"

ws_groups = wb.create_sheet("Group Keys")
ws_items = wb.create_sheet("SAP Codes By Group")
ws_audit = wb.create_sheet("Audit Issues")

ws_summary.append(["Metric", "Value"])
ws_summary.append(["Key Method", "SIZE_RIM_APERTURE_TREAD_LAYER_COLOR_V5_BLACK_DEFAULT"])
ws_summary.append(["SAP Codes", len(item_rows)])
ws_summary.append(["Group Keys", len(groups)])
ws_summary.append(["Group Key Parts", "Tyre Size | Rim/Width | Aperture | Tread | Layer | Color"])
ws_summary.append(["Curing/Handling Used In Key", "NO"])
ws_summary.append(["Backup Timestamp", TIMESTAMP])
style_sheet(ws_summary)

ws_groups.append([
    "Group Key",
    "Tyre Size",
    "Rim/Width",
    "Aperture",
    "Tread",
    "Layer",
    "Color",
    "SAP Count",
])

for row in groups:
    ws_groups.append([
        row["group_key"],
        row["tyre_size"],
        row["rim_width"],
        row["aperture_type"],
        row["tread_pattern"],
        row["layer"],
        row["color"],
        row["sap_count"],
    ])

style_sheet(ws_groups)

ws_items.append([
    "Group Key",
    "Tyre Size",
    "Rim/Width",
    "Aperture",
    "Tread",
    "Layer",
    "Color",
    "SAP Code",
    "Description",
])

for row in group_items:
    ws_items.append([
        row["group_key"],
        row["tyre_size"],
        row["rim_width"],
        row["aperture_type"],
        row["tread_pattern"],
        row["layer"],
        row["color"],
        row["sap_code"],
        row["description"],
    ])

style_sheet(ws_items)

ws_audit.append(["SAP Code", "Group Key", "Issue", "Description"])

for row in audit_rows:
    ws_audit.append([
        row["sap_code"],
        row["group_key"],
        row["issue"],
        row["description"],
    ])

style_sheet(ws_audit)

ws_groups.column_dimensions["A"].width = 50
ws_items.column_dimensions["A"].width = 50
ws_items.column_dimensions["I"].width = 55
ws_audit.column_dimensions["B"].width = 50
ws_audit.column_dimensions["D"].width = 55

wb.save(EXPORT_FILE)

print("Aperture tyre group key rebuild completed.")
print("SAP Codes:", len(item_rows))
print("Group Keys:", len(groups))
print("Audit Rows:", len(audit_rows))
print("Excel:", EXPORT_FILE.resolve())
