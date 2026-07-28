from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import bindparam, text

from app.database import get_session


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def import_core_values(
    workbook_path: str | Path,
    *,
    sheet_name: str = "OVEN",
) -> tuple[int, int]:
    path = Path(workbook_path).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(path)

    workbook = load_workbook(
        path,
        read_only=True,
        data_only=True,
    )
    if sheet_name not in workbook.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' was not found."
        )

    sheet = workbook[sheet_name]
    mapping: dict[str, str] = {}

    # OVEN workbook: Tyre Code = column D, CORE = column N.
    for row in sheet.iter_rows(
        min_row=3,
        values_only=True,
    ):
        tyre_code = _clean(
            row[3] if len(row) > 3 else None
        )
        core = _clean(
            row[13] if len(row) > 13 else None
        )

        if not tyre_code or not core:
            continue

        previous = mapping.get(tyre_code)
        if previous and previous != core:
            raise ValueError(
                "Conflicting CORE values found for "
                f"{tyre_code}: {previous} and {core}"
            )
        mapping[tyre_code] = core

    if not mapping:
        return 0, 0

    with get_session() as session:
        session.execute(
            text(
                """
                ALTER TABLE smds
                ADD COLUMN IF NOT EXISTS core TEXT
                """
            )
        )

        statement = text(
            """
            UPDATE smds
            SET
                core = :core,
                updated_at = CURRENT_TIMESTAMP
            WHERE sap_code = :sap_code
            """
        )

        updated = 0
        for sap_code, core in mapping.items():
            result = session.execute(
                statement,
                {
                    "sap_code": sap_code,
                    "core": core,
                },
            )
            updated += int(result.rowcount or 0)

    return len(mapping), updated


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Import CORE reference values from the "
            "legacy OVEN Excel worksheet into SMDS."
        )
    )
    parser.add_argument(
        "workbook",
        help="Path to the OVEN planning workbook.",
    )
    parser.add_argument(
        "--sheet",
        default="OVEN",
        help="Worksheet name. Default: OVEN",
    )
    args = parser.parse_args()

    found, updated = import_core_values(
        args.workbook,
        sheet_name=args.sheet,
    )

    print("CORE values found:", found)
    print("SMDS rows updated:", updated)


if __name__ == "__main__":
    main()
