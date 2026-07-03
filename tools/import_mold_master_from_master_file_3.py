from pathlib import Path
import re
from collections import defaultdict
from openpyxl import load_workbook
from sqlalchemy import text
from app.database import engine

SOURCE_FILE = Path("data/local_master/Master File (3).xlsx")

def clean(v):
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v).strip())

def norm_header(v):
    return re.sub(r"[^a-z0-9]+", "", clean(v).lower())

def to_int(v):
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(round(float(v)))
    t = clean(v)
    if not t:
        return 0
    try:
        return int(round(float(t)))
    except ValueError:
        return 0

def norm_casing(v):
    t = clean(v)
    if not t:
        return "No Casing"

    low = t.lower().replace(" ", "")
    if low in {"nocasing", "no", "none", "-"}:
        return "No Casing"

    fixes = {
        "B5 Specal 02": "B5 Special 02",
        "B5 Specal 01": "B5 Special 01",
        "B5 Specal 03": "B5 Special 03",
    }
    return fixes.get(t, t)

if not SOURCE_FILE.exists():
    raise SystemExit(f"File not found: {SOURCE_FILE}. Copy Master File (3).xlsx into data/local_master first.")

wb = load_workbook(SOURCE_FILE, read_only=True, data_only=True)

target_ws = None
for ws in wb.worksheets:
    if "mold" in ws.title.lower() and "casing" in ws.title.lower():
        target_ws = ws
        break

if target_ws is None:
    raise SystemExit("Could not find sheet containing 'Mold' and 'Casing'.")

ws = target_ws

expected = {
    "keycode": "mold_key_code",
    "numberofmolds": "mold_count",
    "casingtype": "casing_type",
    "numofcasing": "casing_count",
    "numberofcasing": "casing_count",
    "noofcasing": "casing_count",
}

header_row = None
cols = {}

for row_no, row in enumerate(ws.iter_rows(min_row=1, max_row=50, values_only=True), start=1):
    found = {}
    for col_no, value in enumerate(row, start=1):
        key = norm_header(value)
        if key in expected:
            found[expected[key]] = col_no

    if {"mold_key_code", "mold_count", "casing_type", "casing_count"}.issubset(found):
        header_row = row_no
        cols = found
        break

if not header_row:
    raise SystemExit("Header not found. Need columns: Key Code, Number of Molds, Casing Type, Num of casing")

raw = []
agg = {}

for row_no, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
    key = clean(row[cols["mold_key_code"] - 1] if len(row) >= cols["mold_key_code"] else "")
    if not key:
        continue

    mold_count = to_int(row[cols["mold_count"] - 1] if len(row) >= cols["mold_count"] else 0)
    casing_type = norm_casing(row[cols["casing_type"] - 1] if len(row) >= cols["casing_type"] else "")
    casing_count = to_int(row[cols["casing_count"] - 1] if len(row) >= cols["casing_count"] else 0)

    raw.append((row_no, key, mold_count, casing_type, casing_count))

    if key not in agg:
        agg[key] = {
            "mold_key_code": key,
            "mold_count": 0,
            "casing_types": set(),
            "casing_count": 0,
            "rows": [],
        }

    agg[key]["mold_count"] += mold_count
    agg[key]["casing_types"].add(casing_type)
    agg[key]["casing_count"] += casing_count
    agg[key]["rows"].append(row_no)

wb.close()

master = []
for key, item in agg.items():
    casing_types = sorted(item["casing_types"])
    casing_type = casing_types[0] if len(casing_types) == 1 else "Mixed: " + ", ".join(casing_types)

    master.append({
        "mold_key_code": key,
        "mold_count": item["mold_count"],
        "casing_type": casing_type,
        "casing_count": item["casing_count"],
        "source_rows": ", ".join(str(x) for x in item["rows"]),
        "source_row_count": len(item["rows"]),
    })

with engine.begin() as conn:
    conn.execute(text("""
        CREATE TABLE IF NOT EXISTS mold_master (
            id BIGSERIAL PRIMARY KEY,
            mold_key_code VARCHAR(255) NOT NULL UNIQUE,
            mold_count INTEGER NOT NULL DEFAULT 0,
            casing_type VARCHAR(255) NOT NULL DEFAULT '',
            casing_count INTEGER NOT NULL DEFAULT 0,
            status VARCHAR(32) NOT NULL DEFAULT 'Active',
            remarks TEXT NOT NULL DEFAULT '',
            source_file VARCHAR(255) NOT NULL DEFAULT '',
            source_sheet VARCHAR(255) NOT NULL DEFAULT '',
            source_rows TEXT NOT NULL DEFAULT '',
            created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
    """))

    conn.execute(text("DELETE FROM mold_master"))

    for item in master:
        conn.execute(text("""
            INSERT INTO mold_master (
                mold_key_code,
                mold_count,
                casing_type,
                casing_count,
                status,
                remarks,
                source_file,
                source_sheet,
                source_rows,
                updated_at
            )
            VALUES (
                :mold_key_code,
                :mold_count,
                :casing_type,
                :casing_count,
                'Active',
                '',
                :source_file,
                :source_sheet,
                :source_rows,
                CURRENT_TIMESTAMP
            )
        """), {
            **item,
            "source_file": SOURCE_FILE.name,
            "source_sheet": ws.title,
        })

    summary = conn.execute(text("""
        SELECT
            COUNT(*) AS keys,
            COALESCE(SUM(mold_count), 0) AS molds,
            COALESCE(SUM(casing_count), 0) AS casings
        FROM mold_master
    """)).mappings().one()

    casing_summary = conn.execute(text("""
        SELECT casing_type, COUNT(*) AS keys, SUM(mold_count) AS molds, SUM(casing_count) AS casings
        FROM mold_master
        GROUP BY casing_type
        ORDER BY casing_type
    """)).mappings().all()

duplicates = [x for x in master if x["source_row_count"] > 1]

print("Mold Master import completed from correct source.")
print("Source:", SOURCE_FILE)
print("Sheet:", ws.title)
print("Raw mold rows:", len(raw))
print("Unique mold keys:", summary["keys"])
print("Total mold count:", summary["molds"])
print("Total casing count:", summary["casings"])
print("Duplicate key count:", len(duplicates))
print("")
print("Casing Summary:")
for row in casing_summary:
    print(row["casing_type"], "| Keys:", row["keys"], "| Molds:", row["molds"], "| Casings:", row["casings"])
