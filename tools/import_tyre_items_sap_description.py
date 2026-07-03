from __future__ import annotations

import re
from pathlib import Path
from collections import defaultdict
from datetime import date, datetime

from openpyxl import load_workbook
from sqlalchemy import text

from app.database import engine


SOURCE_DIR = Path("data/local_master")


CODE_HEADERS = [
    "sap code",
    "sap",
    "material",
    "material code",
    "product code",
    "product codes",
    "tyre code",
    "tire code",
    "item code",
]

DESC_HEADERS = [
    "description",
    "material description",
    "full tyre description",
    "full  tyre  description",
    "full tire description",
    "tyre description",
    "tire description",
    "item description",
]


def clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def normalize_code(value) -> str | None:
    if value is None:
        return None

    if isinstance(value, (datetime, date, bool)):
        return None

    if isinstance(value, int):
        code = str(value)
    elif isinstance(value, float):
        if not value.is_integer():
            return None
        code = str(int(value))
    else:
        code = str(value).strip()
        code = code.replace(" ", "")
        code = code.replace(",", "")
        if code.endswith(".0"):
            code = code[:-2]

    if not code:
        return None

    # Numeric SAP code preferred.
    if re.fullmatch(r"\d{6,14}", code):
        return code

    # Allow alphanumeric item codes only if they contain numbers.
    if re.fullmatch(r"[A-Za-z0-9\-/_.]{5,30}", code) and re.search(r"\d", code):
        return code.upper()

    return None


def normalize_desc(value) -> str | None:
    desc = clean(value)

    if not desc:
        return None

    bad = {"0", "N/A", "NA", "NONE", "NULL", "#N/A", "#VALUE!", "-"}
    if desc.upper() in bad:
        return None

    if len(desc) < 3:
        return None

    if not re.search(r"[A-Za-z]", desc):
        return None

    # Avoid header-like values.
    headerish = desc.lower()
    if headerish in CODE_HEADERS or headerish in DESC_HEADERS:
        return None

    return desc


def find_columns(row) -> tuple[list[int], list[int]]:
    code_cols = []
    desc_cols = []

    for idx, value in enumerate(row):
        header = clean(value).lower()

        if not header:
            continue

        if any(word == header or word in header for word in DESC_HEADERS):
            desc_cols.append(idx)
        elif any(word == header or word in header for word in CODE_HEADERS):
            code_cols.append(idx)

    return code_cols, desc_cols


def add_record(records, code: str | None, desc: str | None, source: str, priority: int) -> None:
    if not code or not desc:
        return

    records[code].append({
        "description": desc,
        "source": source,
        "priority": priority,
    })


def scan_sheet(ws, workbook_name: str, records) -> None:
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return

    # Header-based scan.
    for header_index, row in enumerate(rows[:60]):
        code_cols, desc_cols = find_columns(row)

        if not code_cols or not desc_cols:
            continue

        for code_col in code_cols:
            for desc_col in desc_cols:
                if code_col == desc_col:
                    continue

                for data_row in rows[header_index + 1:]:
                    if code_col >= len(data_row) or desc_col >= len(data_row):
                        continue

                    code = normalize_code(data_row[code_col])
                    desc = normalize_desc(data_row[desc_col])

                    add_record(
                        records,
                        code,
                        desc,
                        f"{workbook_name} / {ws.title}",
                        priority=100,
                    )

    # Fallback scan for messy master files.
    if "master" in workbook_name.lower() or "mold" in workbook_name.lower():
        for row in rows:
            values = list(row)

            for idx, value in enumerate(values):
                code = normalize_code(value)

                if not code:
                    continue

                for offset in range(1, 5):
                    desc_idx = idx + offset

                    if desc_idx >= len(values):
                        break

                    desc = normalize_desc(values[desc_idx])

                    if desc:
                        add_record(
                            records,
                            code,
                            desc,
                            f"{workbook_name} / {ws.title}",
                            priority=40,
                        )
                        break


def choose_best(candidates: list[dict]) -> dict:
    return sorted(
        candidates,
        key=lambda item: (item["priority"], len(item["description"])),
        reverse=True,
    )[0]


def ensure_table() -> None:
    with engine.begin() as conn:
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS tyre_item_master (
                id BIGSERIAL PRIMARY KEY,
                sap_code VARCHAR(128) NOT NULL UNIQUE,
                description TEXT NOT NULL DEFAULT '',
                status VARCHAR(32) NOT NULL DEFAULT 'Active',
                remarks TEXT NOT NULL DEFAULT '',
                created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
        """))

        conn.execute(text("""
            ALTER TABLE tyre_item_master
            ADD COLUMN IF NOT EXISTS description TEXT NOT NULL DEFAULT ''
        """))

        conn.execute(text("""
            ALTER TABLE tyre_item_master
            ADD COLUMN IF NOT EXISTS status VARCHAR(32) NOT NULL DEFAULT 'Active'
        """))

        conn.execute(text("""
            ALTER TABLE tyre_item_master
            ADD COLUMN IF NOT EXISTS remarks TEXT NOT NULL DEFAULT ''
        """))

        conn.execute(text("""
            ALTER TABLE tyre_item_master
            ADD COLUMN IF NOT EXISTS updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
        """))


def import_items(items: list[dict]) -> None:
    ensure_table()

    with engine.begin() as conn:
        for item in items:
            conn.execute(
                text("""
                    INSERT INTO tyre_item_master (
                        sap_code,
                        description,
                        status,
                        remarks
                    )
                    VALUES (
                        :sap_code,
                        :description,
                        'Active',
                        :remarks
                    )
                    ON CONFLICT (sap_code)
                    DO UPDATE SET
                        description = EXCLUDED.description,
                        remarks = EXCLUDED.remarks,
                        updated_at = CURRENT_TIMESTAMP
                """),
                {
                    "sap_code": item["sap_code"],
                    "description": item["description"],
                    "remarks": f"Imported from {item['source']}",
                },
            )


def main() -> None:
    files = sorted(SOURCE_DIR.glob("*.xlsx"))

    if not files:
        raise SystemExit("No Excel files found in data/local_master")

    records = defaultdict(list)

    for file_path in files:
        print("Scanning:", file_path.name)

        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
        except Exception as exc:
            print("  skipped:", exc)
            continue

        for ws in wb.worksheets:
            try:
                scan_sheet(ws, file_path.name, records)
            except Exception as exc:
                print(f"  sheet skipped {ws.title}: {exc}")

        wb.close()

    items = []

    for code, candidates in records.items():
        best = choose_best(candidates)
        items.append({
            "sap_code": code,
            "description": best["description"],
            "source": best["source"],
        })

    items.sort(key=lambda row: row["sap_code"])

    print("")
    print("Unique SAP items found:", len(items))
    print("")

    for item in items[:30]:
        print(item["sap_code"], "|", item["description"])

    if len(items) > 30:
        print("... more:", len(items) - 30)

    import_items(items)

    with engine.connect() as conn:
        total = conn.execute(text("SELECT COUNT(*) FROM tyre_item_master")).scalar_one()

    print("")
    print("Imported / updated:", len(items))
    print("Total rows in tyre_item_master:", total)
    print("Done.")


if __name__ == "__main__":
    main()
