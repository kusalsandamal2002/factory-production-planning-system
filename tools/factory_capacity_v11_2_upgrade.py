from __future__ import annotations

import argparse
from datetime import date, timedelta
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook
from sqlalchemy import text

from app.database import get_session
from app.services.factory_resource_intelligence_service import FactoryResourceIntelligenceService


def _txt(value) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\u200d", "").split()).strip()


def _int(value) -> int:
    try:
        return max(0, int(round(float(value or 0))))
    except Exception:
        return 0


def _float(value) -> float:
    try:
        return max(0.0, float(value or 0))
    except Exception:
        return 0.0


def _sheet(wb, wanted: str):
    wanted = wanted.strip().upper()
    for name in wb.sheetnames:
        if name.strip().upper() == wanted:
            return wb[name]
    return None


def _fast_analysis(path: Path, plan_date: date, workbook_name: str):
    wb = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        oven = _sheet(wb, "OVEN")
        band = _sheet(wb, "BAND")
        if oven is None:
            return None

        band_rows = []
        seen_band = set()
        if band is not None:
            for row_no, row in enumerate(band.iter_rows(min_row=4, max_col=8, values_only=True), start=4):
                mold = _txt(row[0] if row else None)
                key = mold.upper()
                if not mold or key in seen_band:
                    continue
                seen_band.add(key)
                band_rows.append({"mold_code": mold, "source_sheet": band.title, "source_row": row_no})

        oven_rows = []
        structure = {}
        for row_no, row in enumerate(oven.iter_rows(min_row=3, min_col=2, max_col=27, values_only=True), start=3):
            # row indexes are relative to Excel column B.
            line = _txt(row[0] if len(row) > 0 else None)
            cavity = _txt(row[1] if len(row) > 1 else None)
            sap = _txt(row[2] if len(row) > 2 else None)
            desc = _txt(row[3] if len(row) > 3 else None)
            if line and cavity:
                key = (line, cavity)
                rec = structure.setdefault(key, {
                    "line_name": line,
                    "cavity_code": cavity,
                    "first_source_row": row_no,
                    "last_source_row": row_no,
                    "allocation_slot_capacity": 0,
                    "source_sheet": oven.title,
                })
                rec["last_source_row"] = row_no
                rec["allocation_slot_capacity"] += 1
            if not sap or not cavity:
                continue
            total_to_produce = _int(row[8] if len(row) > 8 else None)   # J
            today = _int(row[9] if len(row) > 9 else None)             # K
            day_qty = _int(row[10] if len(row) > 10 else None)         # L
            night_qty = _int(row[11] if len(row) > 11 else None)       # M
            next_qty = _int(row[13] if len(row) > 13 else None)        # O
            weight = _float(row[15] if len(row) > 15 else None)        # Q
            balance = _int(row[19] if len(row) > 19 else None)         # U
            casing = _txt(row[20] if len(row) > 20 else None)          # V
            mold = _txt(row[25] if len(row) > 25 else None)            # AA
            if mold.startswith("#"):
                mold = ""
            for shift, qty, offset in (("DAY", day_qty, 0), ("NIGHT", night_qty, 0), ("NEXT DAY", next_qty, 1)):
                if qty <= 0:
                    continue
                d = plan_date + timedelta(days=offset)
                oven_rows.append({
                    "plan_date": d.isoformat(),
                    "line_name": line,
                    "oven_code": cavity,
                    "shift_name": shift,
                    "sap_code": sap,
                    "description": desc,
                    "planned_qty": qty,
                    "today_qty": today,
                    "total_to_produce_qty": total_to_produce,
                    "next_day_qty": next_qty,
                    "balance_qty": balance,
                    "planned_weight_kg": qty * weight,
                    "unit_weight_kg": weight,
                    "casing_evidence": casing,
                    "mold_code": mold,
                    "source_sheet": oven.title,
                    "source_row": row_no,
                })
        return SimpleNamespace(
            oven_rows=oven_rows,
            oven_resource_rows=list(structure.values()),
            band_rows=band_rows,
            bead_rows=[],
            workbook_name=workbook_name,
            plan_date=plan_date.isoformat(),
        )
    finally:
        wb.close()


def run(max_files: int = 0) -> dict[str, int]:
    """Resumable one-time V11.2 history upgrade.

    Each workbook is committed independently so an interruption never throws away
    hours of already-normalized history. All writes are keyed/upserted by the
    original import run, making a rerun safe and idempotent.
    """
    service = FactoryResourceIntelligenceService()
    processed = skipped = failed = allocations = mold_shift = 0

    # Read the catalogue in a short transaction first.
    with get_session() as session:
        service.ensure_schema(session)
        runs = service._safe_rows(
            session,
            """
            SELECT DISTINCT ON (COALESCE(NULLIF(workbook_hash,''), 'RUN:' || id::text))
                   id, workbook_name, workbook_path, archive_path, plan_date, workbook_hash
            FROM excel_import_runs
            WHERE status IN ('COMMITTED','COMMITTED WITH WARNINGS')
              AND rollback_at IS NULL AND plan_date IS NOT NULL
            ORDER BY COALESCE(NULLIF(workbook_hash,''), 'RUN:' || id::text), plan_date DESC, id DESC
            """,
        )

    runs.sort(key=lambda r: (r.get("plan_date") or date.min, int(r.get("id") or 0)))
    if max_files > 0:
        runs = runs[-max_files:]
    total = len(runs)
    latest_date = max((r.get("plan_date") for r in runs if r.get("plan_date")), default=None)

    # If the project was moved since older imports, locate archived workbooks by
    # filename once instead of failing every stale absolute path.
    project_root = Path(__file__).resolve().parents[1]
    fallback_by_name: dict[str, Path] = {}
    for root in (project_root / "data_sources" / "import_archive", project_root / "data_sources"):
        if not root.exists():
            continue
        try:
            for candidate in root.rglob("*.xlsx"):
                fallback_by_name.setdefault(candidate.name.upper(), candidate)
        except Exception:
            pass

    for idx, run in enumerate(runs, start=1):
        candidates = [run.get("archive_path"), run.get("workbook_path")]
        source = next((Path(str(p)) for p in candidates if p and Path(str(p)).exists()), None)
        if source is None:
            source = fallback_by_name.get(str(run.get("workbook_name") or "").upper())
        pct = int(idx / max(1, total) * 100)
        if source is None:
            skipped += 1
            print(f"[{pct:3d}%] SKIP #{run.get('id')}: source workbook not found")
            continue
        try:
            analysis = _fast_analysis(source, run["plan_date"], str(run.get("workbook_name") or source.name))
            if analysis is None:
                skipped += 1
                print(f"[{pct:3d}%] SKIP {source.name}: OVEN sheet not found")
                continue
            # One transaction per workbook = resumable and bounded DB work.
            with get_session() as session:
                service.ensure_schema(session)
                result = service.capture_workbook_resources(
                    session,
                    import_run_id=int(run["id"]),
                    analysis=analysis,
                    import_mode="LIVE" if run.get("plan_date") == latest_date else "HISTORICAL",
                )
                service.sync_operational_oven_columns(session, import_run_id=int(run["id"]))
            processed += 1
            allocations += int(result.get("fi_plan_allocations") or 0)
            mold_shift += int(result.get("fi_mold_shift_usage") or 0)
            print(
                f"[{pct:3d}%] OK   {source.name}: "
                f"{len(analysis.band_rows)} BAND mold codes, "
                f"{result.get('fi_mold_shift_usage', 0)} mold-shift observations"
            )
        except Exception as exc:
            failed += 1
            print(f"[{pct:3d}%] FAIL {source.name}: {exc}")

    print("[100%] Rebuilding unique LINE/CAVITY/MOLD memory and mold profiles...")
    with get_session() as session:
        service.ensure_schema(session)
        service.rebuild_execution_observations(session)
        service.rebuild_mold_profiles(session)
        service.rebuild_resource_memory(session)
        state = service.refresh_state(session)
        session.execute(
            text("UPDATE mpps_fi_state SET model_version=:v, updated_at=CURRENT_TIMESTAMP WHERE id=1"),
            {"v": service.MODEL_VERSION},
        )
        print(
            f"State: latest={state.get('latest_plan_date')} "
            f"execution={state.get('execution_observations')} "
            f"profiles={state.get('capacity_profiles')}"
        )
    return {
        "processed": processed,
        "skipped": skipped,
        "failed": failed,
        "allocations": allocations,
        "mold_shift_usage": mold_shift,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="MPPS Factory Capacity Intelligence V11.2 one-time migration/backfill")
    parser.add_argument("--max-files", type=int, default=0, help="0 = all committed unique workbooks")
    args = parser.parse_args()
    result = run(max_files=max(0, int(args.max_files)))
    print("V11.2 migration complete:", result)
    return 0 if result["failed"] == 0 else 2


if __name__ == "__main__":
    raise SystemExit(main())
