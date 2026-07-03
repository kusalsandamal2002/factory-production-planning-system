from pathlib import Path
from openpyxl import load_workbook
import re

source_dir = Path("data/local_master")

files = []
for pattern in ["*curing*.xlsx", "*production time*.xlsx", "*Tire production time*.xlsx", "*Tyre production time*.xlsx"]:
    files.extend(source_dir.glob(pattern))

files = sorted(set(files))

if not files:
    print("No curing/production time Excel file found in data/local_master")
    print("Available Excel files:")
    for f in source_dir.glob("*.xlsx"):
        print("-", f.name)
    raise SystemExit

keywords = [
    "sap", "material", "code",
    "description", "desc",
    "normal", "short", "cycle", "curing", "cure",
    "handling", "handle", "time", "min", "minutes"
]

def clean(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())

def row_score(row):
    score = 0
    joined = " ".join(clean(v).lower() for v in row)
    for kw in keywords:
        if kw in joined:
            score += 1
    return score

for file_path in files:
    print("")
    print("=" * 100)
    print("FILE:", file_path.name)
    print("=" * 100)

    wb = load_workbook(file_path, read_only=True, data_only=True)

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True, max_row=60))
        if not rows:
            continue

        scored = []
        for idx, row in enumerate(rows, start=1):
            score = row_score(row)
            if score >= 2:
                scored.append((score, idx, row))

        if not scored:
            continue

        scored.sort(reverse=True, key=lambda x: x[0])

        print("")
        print("SHEET:", ws.title)

        for score, header_row_no, header in scored[:3]:
            print("")
            print("Possible header row:", header_row_no, "| score:", score)
            headers = [clean(v) for v in header]
            for col_no, header_name in enumerate(headers, start=1):
                if header_name:
                    print(f"  Col {col_no}: {header_name}")

            print("")
            print("Sample rows:")
            sample_start = header_row_no + 1
            for r in ws.iter_rows(min_row=sample_start, max_row=sample_start + 7, values_only=True):
                values = [clean(v) for v in r]
                compact = [v for v in values if v][:12]
                if compact:
                    print("  ", " | ".join(compact))

    wb.close()
