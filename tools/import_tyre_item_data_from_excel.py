from __future__ import annotations

import argparse
import re
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import text

from app.database import engine


CODE_HEADER_WORDS = (
    "sap code",
    "sap",
    "product codes",
    "product code",
    "tyre code",
    "tire code",
    "material",
)

DESC_HEADER_WORDS = (
    "description",
    "material description",
    "full tyre description",
    "full  tyre  description",
    "full tire description",
)


def clean_text(value) -> str:
    if value is None:
        return ""

    if isinstance(value, (datetime, date)):
        return ""

    text_value = str(value).strip()
    text_value = re.sub(r"\s+", " ", text_value)
    return text_value


def normalize_sap_code(value) -> str | None:
    if value is None:
        return None

    if isinstance(value, (datetime, date)):
        return None

    if isinstance(value, bool):
        return None

    if isinstance(value, int):
        raw = str(value)
    elif isinstance(value, float):
        if not value.is_integer():
            return None
        raw = str(int(value))
    else:
        raw = str(value).strip()
        raw = raw.replace(" ", "")
        raw = raw.replace(",", "")
        if raw.endswith(".0"):
            raw = raw[:-2]

    if not re.fullmatch(r"\d{6,12}", raw):
        return None

    if int(raw) == 0:
        return None

    return raw


def normalize_description(value) -> str | None:
    desc = clean_text(value)

    if not desc:
        return None

    bad_values = {
        "0",
        "#N/A",
        "#VALUE!",
        "N/A",
        "NA",
        "NONE",
        "NULL",
    }

    if desc.upper() in bad_values:
        return None

    if re.fullmatch(r"[-+]?\d+(\.\d+)?", desc):
        return None

    if len(desc) < 3:
        return None

    if not re.search(r"[A-Za-z]", desc):
        return None

    return desc


def header_indexes(row) -> tuple[list[int], list[int]]:
    code_cols: list[int] = []
    desc_cols: list[int] = []

    for idx, value in enumerate(row):
        header = clean_text(value).lower()

        if not header:
            continue

        is_desc = any(word in header for word in DESC_HEADER_WORDS)
        is_code = any(word in header for word in CODE_HEADER_WORDS)

        if is_desc:
            desc_cols.append(idx)
        elif is_code:
            code_cols.append(idx)

    return code_cols, desc_cols


def add_record(records, sap_code: str | None, description: str | None, priority: int, source: str) -> None:
    if not sap_code or not description:
        return

    records[sap_code].append(
        {
            "description": description,
            "priority": priority,
            "source": source,
        }
    )


def extract_from_sheet(ws, workbook_name: str, records) -> None:
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return

    # 1) Header-based extraction.
    for header_row_index, row in enumerate(rows[:40]):
        code_cols, desc_cols = header_indexes(row)

        if not code_cols or not desc_cols:
            continue

        for code_col in code_cols:
            for desc_col in desc_cols:
                if code_col == desc_col:
                    continue

                for data_row in rows[header_row_index + 1:]:
                    if code_col >= len(data_row) or desc_col >= len(data_row):
                        continue

                    sap_code = normalize_sap_code(data_row[code_col])
                    description = normalize_description(data_row[desc_col])

                    add_record(
                        records,
                        sap_code,
                        description,
                        priority=20,
                        source=f"{workbook_name} / {ws.title}",
                    )

    # 2) Generic nearby scan for messy Excel sheets.
    for row in rows:
        values = list(row)

        for idx, value in enumerate(values):
            sap_code = normalize_sap_code(value)

            if not sap_code:
                continue

            # Look for a description within the next 5 cells.
            # This handles Material + Description, Tyre Code + Description,
            # and messy shifted source files.
            for offset in range(1, 6):
                desc_idx = idx + offset

                if desc_idx >= len(values):
                    break

                description = normalize_description(values[desc_idx])

                if description:
                    add_record(
                        records,
                        sap_code,
                        description,
                        priority=10 if offset <= 2 else 5,
                        source=f"{workbook_name} / {ws.title}",
                    )


def choose_best_description(candidates: list[dict]) -> tuple[str, str]:
    # Highest priority first, then longest useful description.
    candidates = sorted(
        candidates,
        key=lambda item: (item["priority"], len(item["description"])),
        reverse=True,
    )

    best = candidates[0]
    return best["description"], best["source"]


def ensure_table() -> None:
    with engine.begin() as conn:
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS tyre_item_master (
                id BIGSERIAL PRIMARY KEY,
                sap_code VARCHAR(128) NOT NULL UNIQUE,
                description TEXT NOT NULL DEFAULT '',
                status VARCHAR(32) NOT NULL DEFAULT 'Active',
                remarks TEXT NOT NULL DEFAULT '',
                created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
        """))

        conn.execute(text("""
            ALTER TABLE tyre_item_master
            ADD COLUMN IF NOT EXISTS description TEXT NOT NULL DEFAULT ''
        """))

        conn.execute(text("""
            ALTER TABLE tyre_item_master
            ADD COLUMN IF NOT EXISTS status VARCHAR(32) NOT NULL DEFAULT 'Active'
        """))

        conn.execute(text("""
            ALTER TABLE tyre_item_master
            ADD COLUMN IF NOT EXISTS remarks TEXT NOT NULL DEFAULT ''
        """))

        conn.execute(text("""
            ALTER TABLE tyre_item_master
            ADD COLUMN IF NOT EXISTS updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
        """))


def import_to_database(items: list[dict], clear_first: bool = False) -> int:
    ensure_table()

    with engine.begin() as conn:
        if clear_first:
            conn.execute(text("DELETE FROM tyre_item_master"))

        for item in items:
            conn.execute(
                text("""
                    INSERT INTO tyre_item_master (
                        sap_code,
                        description,
                        status,
                        remarks
                    )
                    VALUES (
                        :sap_code,
                        :description,
                        'Active',
                        :remarks
                    )
                    ON CONFLICT (sap_code)
                    DO UPDATE SET
                        description = EXCLUDED.description,
                        remarks = EXCLUDED.remarks,
                        updated_at = CURRENT_TIMESTAMP
                """),
                {
                    "sap_code": item["sap_code"],
                    "description": item["description"],
                    "remarks": f"Imported from {item['source']}",
                },
            )

    return len(items)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", default="data/local_master", help="Folder containing source Excel files")
    parser.add_argument("--dry-run", action="store_true", help="Preview only; do not write database")
    parser.add_argument("--clear-first", action="store_true", help="Delete existing tyre items before import")
    args = parser.parse_args()

    source_dir = Path(args.source)

    if not source_dir.exists():
        raise SystemExit(f"Source folder not found: {source_dir}")

    files = sorted(source_dir.glob("*.xlsx"))

    if not files:
        raise SystemExit(f"No .xlsx files found in: {source_dir}")

    records = defaultdict(list)

    for file_path in files:
        print(f"Scanning: {file_path.name}")

        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
        except Exception as exc:
            print(f"  SKIP: could not open workbook: {exc}")
            continue

        for ws in wb.worksheets:
            try:
                extract_from_sheet(ws, file_path.name, records)
            except Exception as exc:
                print(f"  SKIP sheet {ws.title}: {exc}")

        wb.close()

    items: list[dict] = []

    for sap_code, candidates in records.items():
        description, source = choose_best_description(candidates)

        items.append(
            {
                "sap_code": sap_code,
                "description": description,
                "source": source,
            }
        )

    items.sort(key=lambda item: item["sap_code"])

    print()
    print(f"Unique tyre items found: {len(items)}")
    print()

    for item in items[:20]:
        print(f"{item['sap_code']} | {item['description']}")

    if len(items) > 20:
        print(f"... {len(items) - 20} more")

    if args.dry_run:
        print()
        print("Dry run only. Database not changed.")
        return

    imported_count = import_to_database(items, clear_first=args.clear_first)

    with engine.connect() as conn:
        total_rows = conn.execute(text("SELECT COUNT(*) FROM tyre_item_master")).scalar_one()

    print()
    print(f"Imported / updated rows: {imported_count}")
    print(f"Total rows in tyre_item_master: {total_rows}")
    print("Done.")


if __name__ == "__main__":
    main()
