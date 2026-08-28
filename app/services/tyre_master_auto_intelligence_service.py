from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from hashlib import sha256
import json
import math
from pathlib import Path
import re
import threading
from typing import Any

from sqlalchemy import text

from app.database import engine, get_session

try:
    from app.services.operational_source_service import OperationalSourceService
except Exception:
    OperationalSourceService = None


_SCHEMA_LOCK = threading.Lock()
_SCHEMA_READY = False


def _clean(value: Any) -> str:
    if value is None:
        return ""
    text_value = str(value).strip()
    if text_value.lower() in {"none", "nan", "null", "#n/a", "#value!"}:
        return ""
    return text_value


def _number(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        try:
            if math.isnan(float(value)):
                return 0.0
        except Exception:
            pass
        return float(value)
    text_value = _clean(value).replace(",", "")
    if not text_value:
        return 0.0
    try:
        return float(text_value)
    except Exception:
        match = re.search(r"-?\d+(?:\.\d+)?", text_value)
        return float(match.group(0)) if match else 0.0


def _sap(value: Any) -> str:
    text_value = _clean(value)
    if text_value.endswith(".0"):
        text_value = text_value[:-2]
    digits = re.sub(r"\D", "", text_value)
    return digits if len(digits) >= 6 else ""


def _guess_size(description: str) -> str:
    desc = re.sub(r"\s+", " ", _clean(description))
    if not desc:
        return ""
    parts = desc.split()
    if len(parts) >= 2:
        if (
            re.match(r"^\d+(?:\.\d+)?X\d+(?:\.\d+)?$", parts[0], re.I)
            and re.match(r"^\d+/\d+-\d+", parts[1])
        ):
            return f"{parts[0]} {parts[1]}"
        if (
            re.match(r".*-\d+$", parts[0])
            and re.fullmatch(r"\d+/\d+", parts[1])
        ):
            return f"{parts[0]} {parts[1]}"
    return parts[0] if parts else ""


def _file_fingerprint(path: Path) -> str:
    stat = path.stat()
    return f"{path.resolve()}|{stat.st_size}|{stat.st_mtime_ns}"


def _sha256_file(path: Path) -> str:
    digest = sha256()
    with path.open("rb") as handle:
        while True:
            block = handle.read(1024 * 1024)
            if not block:
                break
            digest.update(block)
    return digest.hexdigest()


def _excel_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            from openpyxl.utils.datetime import from_excel
            converted = from_excel(value)
            return converted.date() if isinstance(converted, datetime) else converted
        except Exception:
            return None
    text_value = _clean(value)
    for pattern in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text_value, pattern).date()
        except Exception:
            pass
    return None


@dataclass
class WorkbookLocation:
    path: Path
    workbook_name: str
    plan_date: date | None
    source_label: str = ""


class TyreMasterAutoIntelligenceService:
    MODULES = (
        (
            "MASTER_HEALTH",
            "Master Data Health",
            "Missing / inconsistent master fields and auto-sync quality.",
            1,
        ),
        (
            "SIMILAR_TYRE",
            "Similar Tyre Intelligence",
            "Description / size based nearest-tyre learning for new SAP items.",
            100,
        ),
        (
            "CURING_TIME",
            "Curing Time Intelligence",
            "Learn curing-time baselines and abnormal curing values.",
            50,
        ),
        (
            "LINE_OVEN_COMPATIBILITY",
            "Line & Oven Compatibility",
            "Learn successful SAP → line → oven/cavity relationships.",
            100,
        ),
        (
            "WEIGHT_BASELINE",
            "Tyre Weight Intelligence",
            "Learn stable tyre-weight baselines and weight anomalies.",
            100,
        ),
        (
            "SHIFT_PRODUCTIVITY",
            "Day / Night Productivity",
            "Learn shift-specific plan and actual production behaviour.",
            100,
        ),
        (
            "PLAN_ACHIEVEMENT",
            "Plan Achievement Intelligence",
            "Learn plan-vs-actual achievement by SAP / line / shift.",
            100,
        ),
        (
            "STOCK_PRODUCTION_RISK",
            "Stock & Production Risk",
            "Learn stock coverage, production gap and replenishment pressure.",
            50,
        ),
    )

    @classmethod
    def ensure_schema(cls) -> None:
        global _SCHEMA_READY
        if _SCHEMA_READY:
            return

        with _SCHEMA_LOCK:
            if _SCHEMA_READY:
                return

            statements = [
                """
                CREATE TABLE IF NOT EXISTS mpps_tyre_master_sync_state (
                    id INTEGER PRIMARY KEY,
                    source_fingerprint TEXT NOT NULL DEFAULT '',
                    source_hash TEXT NOT NULL DEFAULT '',
                    workbook_name TEXT NOT NULL DEFAULT '',
                    workbook_path TEXT NOT NULL DEFAULT '',
                    plan_date DATE,
                    status VARCHAR(40) NOT NULL DEFAULT 'NEVER_SYNCED',
                    item_count INTEGER NOT NULL DEFAULT 0,
                    mapping_count INTEGER NOT NULL DEFAULT 0,
                    observation_count INTEGER NOT NULL DEFAULT 0,
                    last_message TEXT NOT NULL DEFAULT '',
                    last_synced_at TIMESTAMP
                )
                """,
                """
                CREATE TABLE IF NOT EXISTS mpps_tyre_factory_mapping (
                    id BIGSERIAL PRIMARY KEY,
                    sap_code TEXT NOT NULL,
                    line TEXT NOT NULL DEFAULT '',
                    oven_no TEXT NOT NULL DEFAULT '',
                    heel TEXT NOT NULL DEFAULT '',
                    soft TEXT NOT NULL DEFAULT '',
                    tread TEXT NOT NULL DEFAULT '',
                    remark TEXT NOT NULL DEFAULT '',
                    weight_kg NUMERIC(14,3) NOT NULL DEFAULT 0,
                    day_plan NUMERIC(14,3) NOT NULL DEFAULT 0,
                    night_plan NUMERIC(14,3) NOT NULL DEFAULT 0,
                    day_produced NUMERIC(14,3) NOT NULL DEFAULT 0,
                    night_produced NUMERIC(14,3) NOT NULL DEFAULT 0,
                    last_plan_date DATE,
                    last_workbook TEXT NOT NULL DEFAULT '',
                    updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(sap_code, line, oven_no)
                )
                """,
                """
                CREATE TABLE IF NOT EXISTS mpps_tyre_workbook_observation (
                    id BIGSERIAL PRIMARY KEY,
                    workbook_hash TEXT NOT NULL,
                    workbook_name TEXT NOT NULL,
                    plan_date DATE,
                    sap_code TEXT NOT NULL,
                    description TEXT NOT NULL DEFAULT '',
                    line TEXT NOT NULL DEFAULT '',
                    oven_no TEXT NOT NULL DEFAULT '',
                    heel TEXT NOT NULL DEFAULT '',
                    soft TEXT NOT NULL DEFAULT '',
                    tread TEXT NOT NULL DEFAULT '',
                    remark TEXT NOT NULL DEFAULT '',
                    weight_kg NUMERIC(14,3) NOT NULL DEFAULT 0,
                    day_plan NUMERIC(14,3) NOT NULL DEFAULT 0,
                    night_plan NUMERIC(14,3) NOT NULL DEFAULT 0,
                    day_produced NUMERIC(14,3) NOT NULL DEFAULT 0,
                    night_produced NUMERIC(14,3) NOT NULL DEFAULT 0,
                    next_day_plan NUMERIC(14,3) NOT NULL DEFAULT 0,
                    total_to_produce NUMERIC(14,3) NOT NULL DEFAULT 0,
                    today_qty NUMERIC(14,3) NOT NULL DEFAULT 0,
                    total_stock NUMERIC(14,3) NOT NULL DEFAULT 0,
                    current_stock NUMERIC(14,3) NOT NULL DEFAULT 0,
                    scrap NUMERIC(14,3) NOT NULL DEFAULT 0,
                    blocked NUMERIC(14,3) NOT NULL DEFAULT 0,
                    source_sheet TEXT NOT NULL DEFAULT '',
                    created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(workbook_hash, sap_code, line, oven_no, source_sheet)
                )
                """,
                """
                CREATE TABLE IF NOT EXISTS mpps_tyre_ml_registry (
                    module_key VARCHAR(64) PRIMARY KEY,
                    module_name TEXT NOT NULL,
                    purpose TEXT NOT NULL,
                    status VARCHAR(40) NOT NULL DEFAULT 'LEARNING',
                    training_rows INTEGER NOT NULL DEFAULT 0,
                    history_days INTEGER NOT NULL DEFAULT 0,
                    readiness_score NUMERIC(10,4) NOT NULL DEFAULT 0,
                    model_version VARCHAR(64) NOT NULL DEFAULT 'V33',
                    last_trained_at TIMESTAMP,
                    updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
                )
                """,
                """
                CREATE TABLE IF NOT EXISTS mpps_tyre_ml_features (
                    module_key VARCHAR(64) NOT NULL,
                    entity_key TEXT NOT NULL,
                    feature_json JSONB NOT NULL DEFAULT '{}'::jsonb,
                    sample_count INTEGER NOT NULL DEFAULT 0,
                    updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    PRIMARY KEY(module_key, entity_key)
                )
                """,
                """
                CREATE INDEX IF NOT EXISTS ix_tyre_obs_sap
                ON mpps_tyre_workbook_observation (sap_code)
                """,
                """
                CREATE INDEX IF NOT EXISTS ix_tyre_obs_plan_date
                ON mpps_tyre_workbook_observation (plan_date)
                """,
                """
                CREATE INDEX IF NOT EXISTS ix_tyre_obs_line_oven
                ON mpps_tyre_workbook_observation (line, oven_no)
                """,
                """
                CREATE INDEX IF NOT EXISTS ix_tyre_mapping_sap
                ON mpps_tyre_factory_mapping (sap_code)
                """,
            ]

            with engine.begin() as conn:
                for statement in statements:
                    conn.execute(text(statement))

                conn.execute(
                    text(
                        """
                        INSERT INTO mpps_tyre_master_sync_state (id)
                        VALUES (1)
                        ON CONFLICT (id) DO NOTHING
                        """
                    )
                )

                for key, name, purpose, _minimum in cls.MODULES:
                    conn.execute(
                        text(
                            """
                            INSERT INTO mpps_tyre_ml_registry (
                                module_key, module_name, purpose,
                                status, model_version
                            )
                            VALUES (
                                :key, :name, :purpose,
                                'LEARNING', 'V33'
                            )
                            ON CONFLICT (module_key) DO UPDATE
                            SET module_name=EXCLUDED.module_name,
                                purpose=EXCLUDED.purpose,
                                model_version='V33',
                                updated_at=CURRENT_TIMESTAMP
                            """
                        ),
                        {
                            "key": key,
                            "name": name,
                            "purpose": purpose,
                        },
                    )

            _SCHEMA_READY = True

    @classmethod
    def _latest_db_workbook(cls) -> tuple[str, str] | None:
        try:
            with engine.connect() as conn:
                exists = bool(
                    conn.execute(
                        text(
                            """
                            SELECT EXISTS (
                                SELECT 1
                                FROM information_schema.tables
                                WHERE table_schema='public'
                                  AND table_name='excel_workbooks'
                            )
                            """
                        )
                    ).scalar()
                )
                if not exists:
                    return None

                row = conn.execute(
                    text(
                        """
                        SELECT
                            COALESCE(file_path, '') AS file_path,
                            COALESCE(original_file_name, '') AS original_file_name
                        FROM excel_workbooks
                        ORDER BY imported_at DESC, id DESC
                        LIMIT 1
                        """
                    )
                ).mappings().first()

                if not row:
                    return None

                return (
                    _clean(row.get("file_path")),
                    _clean(row.get("original_file_name")),
                )
        except Exception:
            return None

    @classmethod
    def locate_latest_workbook(cls) -> WorkbookLocation | None:
        project_root = Path(__file__).resolve().parents[2]
        candidates: list[tuple[Path, str, date | None, str]] = []

        db_workbook = cls._latest_db_workbook()
        if db_workbook:
            path_text, file_name = db_workbook
            if path_text:
                candidate = Path(path_text).expanduser()
                if candidate.exists():
                    candidates.append((candidate, file_name or candidate.name, None, "excel_workbooks"))

            if file_name:
                search_roots = [
                    project_root,
                    project_root / "data",
                    project_root / "uploads",
                    project_root / "storage",
                    project_root / "runtime",
                    Path.home() / "Downloads",
                ]
                for root in search_roots:
                    try:
                        exact = root / file_name
                        if exact.exists():
                            candidates.append((exact, file_name, None, "excel_workbooks"))
                    except Exception:
                        pass

        if OperationalSourceService is not None:
            try:
                with get_session() as session:
                    source = OperationalSourceService.latest(session)

                if source is not None:
                    plan_date = getattr(source, "plan_date", None)
                    workbook_name = _clean(
                        getattr(source, "workbook_name", "")
                        or getattr(source, "file_name", "")
                        or getattr(source, "source_file", "")
                    )
                    for attribute in (
                        "file_path",
                        "workbook_path",
                        "stored_path",
                        "source_path",
                        "upload_path",
                        "path",
                    ):
                        value = _clean(getattr(source, attribute, ""))
                        if value:
                            candidate = Path(value).expanduser()
                            if candidate.exists():
                                candidates.append(
                                    (
                                        candidate,
                                        workbook_name or candidate.name,
                                        plan_date if isinstance(plan_date, date) else None,
                                        "OperationalSourceService",
                                    )
                                )

                    if workbook_name:
                        for root in (
                            project_root,
                            project_root / "data",
                            project_root / "uploads",
                            project_root / "storage",
                            Path.home() / "Downloads",
                        ):
                            try:
                                exact = root / workbook_name
                                if exact.exists():
                                    candidates.append(
                                        (
                                            exact,
                                            workbook_name,
                                            plan_date if isinstance(plan_date, date) else None,
                                            "OperationalSourceService",
                                        )
                                    )
                            except Exception:
                                pass
            except Exception:
                pass

        if not candidates:
            search_roots = [
                project_root / "uploads",
                project_root / "data",
                project_root / "storage",
                Path.home() / "Downloads",
            ]
            for root in search_roots:
                if not root.exists():
                    continue
                try:
                    for candidate in root.glob("*.xlsx"):
                        name_upper = candidate.name.upper()
                        if "OVEN" in name_upper and "SHEET" in name_upper:
                            candidates.append((candidate, candidate.name, None, "fallback"))
                except Exception:
                    pass

        if not candidates:
            return None

        candidates.sort(
            key=lambda item: item[0].stat().st_mtime if item[0].exists() else 0,
            reverse=True,
        )
        path, name, plan_date, label = candidates[0]
        return WorkbookLocation(
            path=path,
            workbook_name=name or path.name,
            plan_date=plan_date,
            source_label=label,
        )

    @classmethod
    def _parse_workbook(
        cls,
        path: Path,
        plan_date_hint: date | None,
    ) -> tuple[list[dict[str, Any]], dict[str, dict[str, Any]], date | None]:
        from openpyxl import load_workbook

        workbook = load_workbook(
            path,
            read_only=True,
            data_only=True,
        )

        observations: list[dict[str, Any]] = []
        stock_by_sap: dict[str, dict[str, Any]] = {}
        plan_date = plan_date_hint

        try:
            daily_name = next(
                (name for name in workbook.sheetnames if name.strip().lower() == "daily  plan".strip().lower()),
                None,
            )
            if daily_name:
                ws = workbook[daily_name]
                if plan_date is None:
                    plan_date = _excel_date(ws.cell(row=3, column=3).value)

                for row_index in range(7, ws.max_row + 1):
                    sap_code = _sap(ws.cell(row=row_index, column=4).value)
                    if not sap_code:
                        continue

                    line = _clean(ws.cell(row=row_index, column=2).value)
                    oven_no = _clean(ws.cell(row=row_index, column=3).value)
                    description = _clean(ws.cell(row=row_index, column=5).value)

                    observations.append(
                        {
                            "sap_code": sap_code,
                            "description": description,
                            "line": line,
                            "oven_no": oven_no,
                            "heel": "",
                            "soft": "",
                            "tread": "",
                            "remark": _clean(ws.cell(row=row_index, column=7).value),
                            "weight_kg": _number(ws.cell(row=row_index, column=9).value),
                            "day_plan": _number(ws.cell(row=row_index, column=11).value),
                            "day_produced": _number(ws.cell(row=row_index, column=12).value),
                            "night_plan": _number(ws.cell(row=row_index, column=15).value),
                            "night_produced": _number(ws.cell(row=row_index, column=16).value),
                            "next_day_plan": _number(ws.cell(row=row_index, column=19).value),
                            "total_to_produce": _number(ws.cell(row=row_index, column=8).value),
                            "today_qty": 0.0,
                            "source_sheet": "Daily Plan",
                        }
                    )

            oven_name = next(
                (name for name in workbook.sheetnames if name.strip().upper() == "OVEN"),
                None,
            )
            if oven_name:
                ws = workbook[oven_name]
                for row_index in range(3, ws.max_row + 1):
                    sap_code = _sap(ws.cell(row=row_index, column=4).value)
                    if not sap_code:
                        continue

                    observations.append(
                        {
                            "sap_code": sap_code,
                            "description": _clean(ws.cell(row=row_index, column=5).value),
                            "line": _clean(ws.cell(row=row_index, column=2).value),
                            "oven_no": _clean(ws.cell(row=row_index, column=3).value),
                            "heel": _clean(ws.cell(row=row_index, column=6).value),
                            "soft": _clean(ws.cell(row=row_index, column=7).value),
                            "tread": _clean(ws.cell(row=row_index, column=8).value),
                            "remark": _clean(ws.cell(row=row_index, column=9).value),
                            "weight_kg": _number(ws.cell(row=row_index, column=17).value),
                            "day_plan": _number(ws.cell(row=row_index, column=12).value),
                            "day_produced": 0.0,
                            "night_plan": _number(ws.cell(row=row_index, column=13).value),
                            "night_produced": 0.0,
                            "next_day_plan": _number(ws.cell(row=row_index, column=15).value),
                            "total_to_produce": _number(ws.cell(row=row_index, column=10).value),
                            "today_qty": _number(ws.cell(row=row_index, column=11).value),
                            "source_sheet": "OVEN",
                        }
                    )

            prod_name = next(
                (name for name in workbook.sheetnames if name.strip().upper() == "PROD"),
                None,
            )
            if prod_name:
                ws = workbook[prod_name]
                for row_index in range(4, ws.max_row + 1):
                    sap_code = _sap(ws.cell(row=row_index, column=2).value)
                    if not sap_code:
                        continue

                    stock_by_sap[sap_code] = {
                        "description": _clean(ws.cell(row=row_index, column=3).value),
                        "total_stock": _number(ws.cell(row=row_index, column=4).value),
                        "scrap": _number(ws.cell(row=row_index, column=5).value),
                        "blocked": _number(ws.cell(row=row_index, column=6).value),
                        "current_stock": _number(ws.cell(row=row_index, column=227).value),
                    }

        finally:
            workbook.close()

        return observations, stock_by_sap, plan_date

    @classmethod
    def _smds_columns(cls, conn) -> set[str]:
        rows = conn.execute(
            text(
                """
                SELECT column_name
                FROM information_schema.columns
                WHERE table_schema='public'
                  AND table_name='smds'
                """
            )
        ).scalars().all()
        return {str(value) for value in rows}

    @classmethod
    def _upsert_master(
        cls,
        conn,
        sap_code: str,
        item: dict[str, Any],
        workbook_name: str,
    ) -> None:
        columns = cls._smds_columns(conn)
        if "sap_code" not in columns:
            return

        available: dict[str, Any] = {
            "sap_code": sap_code,
            "material_description": _clean(item.get("description")),
            "line": _clean(item.get("line")),
            "heel": _clean(item.get("heel")),
            "soft": _clean(item.get("soft")),
            "tred": _clean(item.get("tread")),
            "weight_per_tyre_kg": _number(item.get("weight_kg")),
            "day_plan": _number(item.get("day_plan")),
            "night_plan": _number(item.get("night_plan")),
            "total_plan": _number(item.get("day_plan")) + _number(item.get("night_plan")),
            "source_file": workbook_name,
            "source_sheet": "AUTO OVEN V33",
        }

        fields = [
            key
            for key in available
            if key in columns
        ]

        if "material_description" not in fields:
            return

        insert_columns = ", ".join(fields)
        bind_names = ", ".join(f":{field}" for field in fields)

        update_fields = []
        for field in fields:
            if field == "sap_code":
                continue

            if field in {
                "weight_per_tyre_kg",
                "day_plan",
                "night_plan",
                "total_plan",
            }:
                update_fields.append(
                    f"{field}=CASE "
                    f"WHEN EXCLUDED.{field} IS NOT NULL AND EXCLUDED.{field} <> 0 "
                    f"THEN EXCLUDED.{field} ELSE smds.{field} END"
                )
            else:
                update_fields.append(
                    f"{field}=CASE "
                    f"WHEN NULLIF(BTRIM(CAST(EXCLUDED.{field} AS TEXT)), '') IS NOT NULL "
                    f"THEN EXCLUDED.{field} ELSE smds.{field} END"
                )

        if "updated_at" in columns:
            update_fields.append("updated_at=CURRENT_TIMESTAMP")

        sql = f"""
            INSERT INTO smds ({insert_columns})
            VALUES ({bind_names})
            ON CONFLICT (sap_code) DO UPDATE
            SET {", ".join(update_fields)}
        """

        conn.execute(
            text(sql),
            {field: available[field] for field in fields},
        )

    @classmethod
    def _merge_current_master(
        cls,
        observations: list[dict[str, Any]],
        stock_by_sap: dict[str, dict[str, Any]],
    ) -> dict[str, dict[str, Any]]:
        grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in observations:
            grouped[row["sap_code"]].append(row)

        result: dict[str, dict[str, Any]] = {}

        for sap_code, rows in grouped.items():
            def score(row: dict[str, Any]) -> float:
                return (
                    _number(row.get("day_plan"))
                    + _number(row.get("night_plan"))
                    + _number(row.get("today_qty"))
                    + (1 if _clean(row.get("line")) else 0)
                    + (1 if _clean(row.get("oven_no")) else 0)
                )

            selected = max(rows, key=score)
            merged = dict(selected)

            for field in (
                "description",
                "line",
                "oven_no",
                "heel",
                "soft",
                "tread",
                "remark",
            ):
                if not _clean(merged.get(field)):
                    for candidate in rows:
                        if _clean(candidate.get(field)):
                            merged[field] = candidate.get(field)
                            break

            for field in (
                "weight_kg",
                "day_plan",
                "night_plan",
                "day_produced",
                "night_produced",
                "next_day_plan",
                "total_to_produce",
                "today_qty",
            ):
                if _number(merged.get(field)) == 0:
                    values = [
                        _number(candidate.get(field))
                        for candidate in rows
                        if _number(candidate.get(field)) != 0
                    ]
                    if values:
                        merged[field] = max(values)

            stock = stock_by_sap.get(sap_code, {})
            if not _clean(merged.get("description")):
                merged["description"] = _clean(stock.get("description"))
            merged.update(
                {
                    "total_stock": _number(stock.get("total_stock")),
                    "current_stock": _number(stock.get("current_stock")),
                    "scrap": _number(stock.get("scrap")),
                    "blocked": _number(stock.get("blocked")),
                }
            )
            result[sap_code] = merged

        for sap_code, stock in stock_by_sap.items():
            if sap_code in result:
                continue
            result[sap_code] = {
                "sap_code": sap_code,
                "description": _clean(stock.get("description")),
                "line": "",
                "oven_no": "",
                "heel": "",
                "soft": "",
                "tread": "",
                "remark": "",
                "weight_kg": 0.0,
                "day_plan": 0.0,
                "night_plan": 0.0,
                "day_produced": 0.0,
                "night_produced": 0.0,
                "next_day_plan": 0.0,
                "total_to_produce": 0.0,
                "today_qty": 0.0,
                "total_stock": _number(stock.get("total_stock")),
                "current_stock": _number(stock.get("current_stock")),
                "scrap": _number(stock.get("scrap")),
                "blocked": _number(stock.get("blocked")),
            }

        return result

    @classmethod
    def sync_latest(cls, force: bool = False) -> dict[str, Any]:
        cls.ensure_schema()

        location = cls.locate_latest_workbook()
        if location is None:
            return {
                "changed": False,
                "status": "NO_WORKBOOK",
                "message": "No committed OVEN workbook file was found.",
            }

        try:
            quick_fingerprint = _file_fingerprint(location.path)
        except Exception as exc:
            return {
                "changed": False,
                "status": "FILE_ERROR",
                "message": str(exc),
            }

        with engine.connect() as conn:
            state = conn.execute(
                text(
                    """
                    SELECT source_fingerprint
                    FROM mpps_tyre_master_sync_state
                    WHERE id=1
                    """
                )
            ).scalar()

        if not force and _clean(state) == quick_fingerprint:
            return {
                "changed": False,
                "status": "UP_TO_DATE",
                "workbook_name": location.workbook_name,
                "message": "Tyre Master already matches the latest workbook.",
            }

        workbook_hash = _sha256_file(location.path)
        observations, stock_by_sap, plan_date = cls._parse_workbook(
            location.path,
            location.plan_date,
        )
        master = cls._merge_current_master(
            observations,
            stock_by_sap,
        )

        mapping_count = 0
        observation_count = 0

        with engine.begin() as conn:
            for sap_code, item in master.items():
                cls._upsert_master(
                    conn,
                    sap_code,
                    item,
                    location.workbook_name,
                )

            for row in observations:
                sap_code = row["sap_code"]
                stock = stock_by_sap.get(sap_code, {})

                conn.execute(
                    text(
                        """
                        INSERT INTO mpps_tyre_factory_mapping (
                            sap_code, line, oven_no,
                            heel, soft, tread, remark,
                            weight_kg, day_plan, night_plan,
                            day_produced, night_produced,
                            last_plan_date, last_workbook,
                            updated_at
                        )
                        VALUES (
                            :sap_code, :line, :oven_no,
                            :heel, :soft, :tread, :remark,
                            :weight_kg, :day_plan, :night_plan,
                            :day_produced, :night_produced,
                            :plan_date, :workbook_name,
                            CURRENT_TIMESTAMP
                        )
                        ON CONFLICT (sap_code, line, oven_no) DO UPDATE
                        SET heel=CASE WHEN EXCLUDED.heel<>'' THEN EXCLUDED.heel ELSE mpps_tyre_factory_mapping.heel END,
                            soft=CASE WHEN EXCLUDED.soft<>'' THEN EXCLUDED.soft ELSE mpps_tyre_factory_mapping.soft END,
                            tread=CASE WHEN EXCLUDED.tread<>'' THEN EXCLUDED.tread ELSE mpps_tyre_factory_mapping.tread END,
                            remark=CASE WHEN EXCLUDED.remark<>'' THEN EXCLUDED.remark ELSE mpps_tyre_factory_mapping.remark END,
                            weight_kg=CASE WHEN EXCLUDED.weight_kg<>0 THEN EXCLUDED.weight_kg ELSE mpps_tyre_factory_mapping.weight_kg END,
                            day_plan=EXCLUDED.day_plan,
                            night_plan=EXCLUDED.night_plan,
                            day_produced=EXCLUDED.day_produced,
                            night_produced=EXCLUDED.night_produced,
                            last_plan_date=EXCLUDED.last_plan_date,
                            last_workbook=EXCLUDED.last_workbook,
                            updated_at=CURRENT_TIMESTAMP
                        """
                    ),
                    {
                        **row,
                        "workbook_name": location.workbook_name,
                        "plan_date": plan_date,
                    },
                )
                mapping_count += 1

                conn.execute(
                    text(
                        """
                        INSERT INTO mpps_tyre_workbook_observation (
                            workbook_hash, workbook_name, plan_date,
                            sap_code, description, line, oven_no,
                            heel, soft, tread, remark,
                            weight_kg, day_plan, night_plan,
                            day_produced, night_produced,
                            next_day_plan, total_to_produce, today_qty,
                            total_stock, current_stock, scrap, blocked,
                            source_sheet
                        )
                        VALUES (
                            :workbook_hash, :workbook_name, :plan_date,
                            :sap_code, :description, :line, :oven_no,
                            :heel, :soft, :tread, :remark,
                            :weight_kg, :day_plan, :night_plan,
                            :day_produced, :night_produced,
                            :next_day_plan, :total_to_produce, :today_qty,
                            :total_stock, :current_stock, :scrap, :blocked,
                            :source_sheet
                        )
                        ON CONFLICT (
                            workbook_hash, sap_code, line, oven_no, source_sheet
                        ) DO NOTHING
                        """
                    ),
                    {
                        **row,
                        "workbook_hash": workbook_hash,
                        "workbook_name": location.workbook_name,
                        "plan_date": plan_date,
                        "total_stock": _number(stock.get("total_stock")),
                        "current_stock": _number(stock.get("current_stock")),
                        "scrap": _number(stock.get("scrap")),
                        "blocked": _number(stock.get("blocked")),
                    },
                )
                observation_count += 1

            conn.execute(
                text(
                    """
                    UPDATE mpps_tyre_master_sync_state
                    SET source_fingerprint=:fingerprint,
                        source_hash=:source_hash,
                        workbook_name=:workbook_name,
                        workbook_path=:workbook_path,
                        plan_date=:plan_date,
                        status='SYNCED',
                        item_count=:item_count,
                        mapping_count=:mapping_count,
                        observation_count=:observation_count,
                        last_message=:message,
                        last_synced_at=CURRENT_TIMESTAMP
                    WHERE id=1
                    """
                ),
                {
                    "fingerprint": quick_fingerprint,
                    "source_hash": workbook_hash,
                    "workbook_name": location.workbook_name,
                    "workbook_path": str(location.path),
                    "plan_date": plan_date,
                    "item_count": len(master),
                    "mapping_count": mapping_count,
                    "observation_count": observation_count,
                    "message": (
                        f"Auto-synced {len(master)} tyre items "
                        f"from {location.workbook_name}"
                    ),
                },
            )

        learning = cls.train_all(incremental=True)

        return {
            "changed": True,
            "status": "SYNCED",
            "workbook_name": location.workbook_name,
            "plan_date": plan_date.isoformat() if plan_date else "",
            "item_count": len(master),
            "mapping_count": mapping_count,
            "observation_count": observation_count,
            "learning": learning,
            "message": (
                f"Auto-synced {len(master):,} tyre items and "
                f"{observation_count:,} learning observations."
            ),
        }

    @classmethod
    def _history_counts(cls, conn) -> dict[str, int]:
        row = conn.execute(
            text(
                """
                SELECT
                    COUNT(*) AS observations,
                    COUNT(DISTINCT plan_date) FILTER (
                        WHERE plan_date IS NOT NULL
                    ) AS history_days,
                    COUNT(*) FILTER (
                        WHERE weight_kg<>0
                    ) AS weight_rows,
                    COUNT(*) FILTER (
                        WHERE line<>'' AND oven_no<>''
                    ) AS line_oven_rows,
                    COUNT(*) FILTER (
                        WHERE day_plan<>0 OR night_plan<>0
                    ) AS shift_rows,
                    COUNT(*) FILTER (
                        WHERE day_produced<>0 OR night_produced<>0
                    ) AS actual_rows,
                    COUNT(*) FILTER (
                        WHERE current_stock<>0 OR total_stock<>0
                    ) AS stock_rows
                FROM mpps_tyre_workbook_observation
                """
            )
        ).mappings().first()

        return {
            key: int((row or {}).get(key) or 0)
            for key in (
                "observations",
                "history_days",
                "weight_rows",
                "line_oven_rows",
                "shift_rows",
                "actual_rows",
                "stock_rows",
            )
        }

    @classmethod
    def _master_health(cls, conn) -> tuple[int, float]:
        columns = cls._smds_columns(conn)
        if not columns:
            return 0, 0.0

        total = int(
            conn.execute(text("SELECT COUNT(*) FROM smds")).scalar()
            or 0
        )
        if total <= 0:
            return 0, 0.0

        checks = []
        for column in (
            "sap_code",
            "material_description",
            "line",
            "key_code",
            "casing_type",
            "curing_cycle",
            "weight_per_tyre_kg",
        ):
            if column not in columns:
                continue
            if column == "weight_per_tyre_kg":
                checks.append(
                    f"CASE WHEN {column} IS NULL OR {column}=0 THEN 1 ELSE 0 END"
                )
            else:
                checks.append(
                    f"CASE WHEN {column} IS NULL OR BTRIM(CAST({column} AS TEXT))='' "
                    f"OR BTRIM(CAST({column} AS TEXT))='-' THEN 1 ELSE 0 END"
                )

        if not checks:
            return total, 100.0

        missing = int(
            conn.execute(
                text(
                    f"""
                    SELECT COALESCE(SUM(
                        {" + ".join(checks)}
                    ), 0)
                    FROM smds
                    """
                )
            ).scalar()
            or 0
        )
        denominator = max(1, total * len(checks))
        health = max(0.0, min(100.0, 100.0 - (missing / denominator) * 100.0))
        return total, health

    @classmethod
    def _rebuild_features(cls, conn) -> None:
        conn.execute(
            text(
                """
                DELETE FROM mpps_tyre_ml_features
                WHERE module_key IN (
                    'LINE_OVEN_COMPATIBILITY',
                    'WEIGHT_BASELINE',
                    'SHIFT_PRODUCTIVITY',
                    'PLAN_ACHIEVEMENT',
                    'STOCK_PRODUCTION_RISK'
                )
                """
            )
        )

        conn.execute(
            text(
                """
                INSERT INTO mpps_tyre_ml_features (
                    module_key, entity_key, feature_json, sample_count
                )
                SELECT
                    'LINE_OVEN_COMPATIBILITY',
                    sap_code || '|' || line || '|' || oven_no,
                    jsonb_build_object(
                        'sap_code', sap_code,
                        'line', line,
                        'oven_no', oven_no,
                        'avg_plan', AVG(day_plan + night_plan),
                        'avg_today', AVG(today_qty),
                        'last_seen', MAX(plan_date)
                    ),
                    COUNT(*)
                FROM mpps_tyre_workbook_observation
                WHERE line<>'' AND oven_no<>''
                GROUP BY sap_code, line, oven_no
                """
            )
        )

        conn.execute(
            text(
                """
                INSERT INTO mpps_tyre_ml_features (
                    module_key, entity_key, feature_json, sample_count
                )
                SELECT
                    'WEIGHT_BASELINE',
                    sap_code,
                    jsonb_build_object(
                        'avg_weight', AVG(weight_kg),
                        'min_weight', MIN(weight_kg),
                        'max_weight', MAX(weight_kg),
                        'last_seen', MAX(plan_date)
                    ),
                    COUNT(*)
                FROM mpps_tyre_workbook_observation
                WHERE weight_kg>0
                GROUP BY sap_code
                """
            )
        )

        conn.execute(
            text(
                """
                INSERT INTO mpps_tyre_ml_features (
                    module_key, entity_key, feature_json, sample_count
                )
                SELECT
                    'SHIFT_PRODUCTIVITY',
                    sap_code || '|' || line,
                    jsonb_build_object(
                        'avg_day_plan', AVG(day_plan),
                        'avg_night_plan', AVG(night_plan),
                        'avg_day_actual', AVG(day_produced),
                        'avg_night_actual', AVG(night_produced),
                        'last_seen', MAX(plan_date)
                    ),
                    COUNT(*)
                FROM mpps_tyre_workbook_observation
                WHERE line<>'' AND (day_plan<>0 OR night_plan<>0)
                GROUP BY sap_code, line
                """
            )
        )

        conn.execute(
            text(
                """
                INSERT INTO mpps_tyre_ml_features (
                    module_key, entity_key, feature_json, sample_count
                )
                SELECT
                    'PLAN_ACHIEVEMENT',
                    sap_code || '|' || line,
                    jsonb_build_object(
                        'avg_achievement',
                        AVG(
                            CASE
                                WHEN (day_plan + night_plan)>0
                                THEN (day_produced + night_produced)
                                     / NULLIF(day_plan + night_plan, 0)
                                ELSE NULL
                            END
                        ),
                        'last_seen', MAX(plan_date)
                    ),
                    COUNT(*)
                FROM mpps_tyre_workbook_observation
                WHERE line<>''
                  AND (day_plan + night_plan)>0
                  AND (day_produced + night_produced)>0
                GROUP BY sap_code, line
                """
            )
        )

        conn.execute(
            text(
                """
                INSERT INTO mpps_tyre_ml_features (
                    module_key, entity_key, feature_json, sample_count
                )
                SELECT DISTINCT ON (sap_code)
                    'STOCK_PRODUCTION_RISK',
                    sap_code,
                    jsonb_build_object(
                        'current_stock', current_stock,
                        'total_stock', total_stock,
                        'planned_qty', day_plan + night_plan,
                        'production_gap',
                        GREATEST(
                            (day_plan + night_plan) - current_stock,
                            0
                        ),
                        'last_seen', plan_date
                    ),
                    1
                FROM mpps_tyre_workbook_observation
                WHERE current_stock<>0
                   OR total_stock<>0
                   OR (day_plan + night_plan)<>0
                ORDER BY sap_code, plan_date DESC NULLS LAST, id DESC
                """
            )
        )

    @classmethod
    def train_all(cls, incremental: bool = False) -> dict[str, Any]:
        cls.ensure_schema()

        with engine.begin() as conn:
            counts = cls._history_counts(conn)
            item_count, health = cls._master_health(conn)
            cls._rebuild_features(conn)

            curing_columns = cls._smds_columns(conn)
            curing_rows = 0
            if "normal_curing_minutes" in curing_columns:
                curing_rows = int(
                    conn.execute(
                        text(
                            """
                            SELECT COUNT(*)
                            FROM smds
                            WHERE normal_curing_minutes IS NOT NULL
                              AND normal_curing_minutes<>0
                            """
                        )
                    ).scalar()
                    or 0
                )
            elif "curing_cycle" in curing_columns:
                curing_rows = int(
                    conn.execute(
                        text(
                            """
                            SELECT COUNT(*)
                            FROM smds
                            WHERE curing_cycle IS NOT NULL
                              AND BTRIM(CAST(curing_cycle AS TEXT))<>''
                              AND BTRIM(CAST(curing_cycle AS TEXT))<>'-'
                            """
                        )
                    ).scalar()
                    or 0
                )

            training_rows = {
                "MASTER_HEALTH": item_count,
                "SIMILAR_TYRE": item_count,
                "CURING_TIME": curing_rows,
                "LINE_OVEN_COMPATIBILITY": counts["line_oven_rows"],
                "WEIGHT_BASELINE": counts["weight_rows"],
                "SHIFT_PRODUCTIVITY": counts["shift_rows"],
                "PLAN_ACHIEVEMENT": counts["actual_rows"],
                "STOCK_PRODUCTION_RISK": counts["stock_rows"],
            }

            ready_count = 0

            for key, name, purpose, minimum in cls.MODULES:
                rows = int(training_rows.get(key, 0))
                history_days = counts["history_days"]

                history_requirement = 0
                if key in {
                    "CURING_TIME",
                    "LINE_OVEN_COMPATIBILITY",
                    "WEIGHT_BASELINE",
                    "SHIFT_PRODUCTIVITY",
                    "PLAN_ACHIEVEMENT",
                    "STOCK_PRODUCTION_RISK",
                }:
                    history_requirement = 10

                ready = (
                    rows >= minimum
                    and history_days >= history_requirement
                )

                if key == "MASTER_HEALTH":
                    ready = item_count > 0

                status = (
                    "TRAINED"
                    if ready
                    else "LEARNING"
                )
                ready_count += int(ready)

                row_score = min(1.0, rows / max(1, minimum))
                day_score = (
                    1.0
                    if history_requirement <= 0
                    else min(1.0, history_days / history_requirement)
                )
                readiness_score = 100.0 * min(row_score, day_score)

                conn.execute(
                    text(
                        """
                        INSERT INTO mpps_tyre_ml_registry (
                            module_key, module_name, purpose,
                            status, training_rows, history_days,
                            readiness_score, model_version,
                            last_trained_at, updated_at
                        )
                        VALUES (
                            :key, :name, :purpose,
                            :status, :training_rows, :history_days,
                            :readiness_score, 'V33',
                            CASE
                                WHEN :status='TRAINED'
                                THEN CURRENT_TIMESTAMP
                                ELSE NULL
                            END,
                            CURRENT_TIMESTAMP
                        )
                        ON CONFLICT (module_key) DO UPDATE
                        SET module_name=EXCLUDED.module_name,
                            purpose=EXCLUDED.purpose,
                            status=EXCLUDED.status,
                            training_rows=EXCLUDED.training_rows,
                            history_days=EXCLUDED.history_days,
                            readiness_score=EXCLUDED.readiness_score,
                            model_version='V33',
                            last_trained_at=CASE
                                WHEN EXCLUDED.status='TRAINED'
                                THEN CURRENT_TIMESTAMP
                                ELSE mpps_tyre_ml_registry.last_trained_at
                            END,
                            updated_at=CURRENT_TIMESTAMP
                        """
                    ),
                    {
                        "key": key,
                        "name": name,
                        "purpose": purpose,
                        "status": status,
                        "training_rows": rows,
                        "history_days": history_days,
                        "readiness_score": readiness_score,
                    },
                )

        return {
            "incremental": bool(incremental),
            "ready_count": ready_count,
            "module_count": len(cls.MODULES),
            "history_days": counts["history_days"],
            "master_health": health,
        }

    @classmethod
    def dashboard(cls) -> dict[str, Any]:
        cls.ensure_schema()

        with engine.connect() as conn:
            item_count, health = cls._master_health(conn)
            counts = cls._history_counts(conn)

            state = conn.execute(
                text(
                    """
                    SELECT *
                    FROM mpps_tyre_master_sync_state
                    WHERE id=1
                    """
                )
            ).mappings().first()

            modules = [
                dict(row)
                for row in conn.execute(
                    text(
                        """
                        SELECT
                            module_key, module_name, purpose,
                            status, training_rows, history_days,
                            readiness_score, model_version,
                            last_trained_at
                        FROM mpps_tyre_ml_registry
                        ORDER BY
                            CASE module_key
                                WHEN 'MASTER_HEALTH' THEN 1
                                WHEN 'SIMILAR_TYRE' THEN 2
                                WHEN 'CURING_TIME' THEN 3
                                WHEN 'LINE_OVEN_COMPATIBILITY' THEN 4
                                WHEN 'WEIGHT_BASELINE' THEN 5
                                WHEN 'SHIFT_PRODUCTIVITY' THEN 6
                                WHEN 'PLAN_ACHIEVEMENT' THEN 7
                                WHEN 'STOCK_PRODUCTION_RISK' THEN 8
                                ELSE 99
                            END
                        """
                    )
                ).mappings().all()
            ]

        ready_count = sum(
            1
            for row in modules
            if _clean(row.get("status")).upper() == "TRAINED"
        )

        return {
            "item_count": item_count,
            "health": health,
            "history_days": counts["history_days"],
            "observation_count": counts["observations"],
            "ready_count": ready_count,
            "module_count": len(modules),
            "modules": modules,
            "sync_state": dict(state or {}),
        }

    @classmethod
    def _master_columns(cls, conn) -> set[str]:
        return cls._smds_columns(conn)

    @classmethod
    def page_data(
        cls,
        view_key: str,
        search: str = "",
        page: int = 1,
        page_size: int = 200,
    ) -> dict[str, Any]:
        cls.ensure_schema()

        page = max(1, int(page))
        page_size = max(50, min(500, int(page_size)))
        offset = (page - 1) * page_size
        search_text = _clean(search)
        search_sql = ""
        params: dict[str, Any] = {
            "limit": page_size,
            "offset": offset,
        }

        if search_text:
            search_sql = """
                AND (
                    s.sap_code ILIKE :search
                    OR COALESCE(s.material_description, '') ILIKE :search
                    OR COALESCE(s.line, '') ILIKE :search
                    OR COALESCE(m.oven_no, '') ILIKE :search
                    OR COALESCE(s.key_code, '') ILIKE :search
                    OR COALESCE(s.casing_type, '') ILIKE :search
                )
            """
            params["search"] = f"%{search_text}%"

        with engine.connect() as conn:
            columns = cls._master_columns(conn)

            line_expr = (
                "COALESCE(NULLIF(s.line, ''), m.line, '')"
                if "line" in columns
                else "COALESCE(m.line, '')"
            )
            desc_expr = (
                "COALESCE(s.material_description, '')"
                if "material_description" in columns
                else "''"
            )
            key_expr = (
                "COALESCE(s.key_code, '')"
                if "key_code" in columns
                else "''"
            )
            casing_expr = (
                "COALESCE(s.casing_type, '')"
                if "casing_type" in columns
                else "''"
            )
            curing_expr = (
                "COALESCE(CAST(s.normal_curing_minutes AS TEXT), '')"
                if "normal_curing_minutes" in columns
                else (
                    "COALESCE(CAST(s.curing_cycle AS TEXT), '')"
                    if "curing_cycle" in columns
                    else "''"
                )
            )
            weight_expr = (
                "COALESCE(s.weight_per_tyre_kg, m.weight_kg, 0)"
                if "weight_per_tyre_kg" in columns
                else "COALESCE(m.weight_kg, 0)"
            )
            heel_expr = (
                "COALESCE(NULLIF(s.heel, ''), m.heel, '')"
                if "heel" in columns
                else "COALESCE(m.heel, '')"
            )
            soft_expr = (
                "COALESCE(NULLIF(s.soft, ''), m.soft, '')"
                if "soft" in columns
                else "COALESCE(m.soft, '')"
            )
            tread_column = "tred" if "tred" in columns else ("tread" if "tread" in columns else "")
            tread_expr = (
                f"COALESCE(NULLIF(s.{tread_column}, ''), m.tread, '')"
                if tread_column
                else "COALESCE(m.tread, '')"
            )

            latest_mapping = """
                LEFT JOIN LATERAL (
                    SELECT fm.*
                    FROM mpps_tyre_factory_mapping fm
                    WHERE fm.sap_code=s.sap_code
                    ORDER BY fm.last_plan_date DESC NULLS LAST,
                             fm.updated_at DESC,
                             fm.id DESC
                    LIMIT 1
                ) m ON TRUE
            """

            base_where = "WHERE 1=1"

            total = int(
                conn.execute(
                    text(
                        f"""
                        SELECT COUNT(*)
                        FROM smds s
                        {latest_mapping}
                        {base_where}
                        {search_sql}
                        """
                    ),
                    params,
                ).scalar()
                or 0
            )

            if view_key == "MASTER":
                rows = conn.execute(
                    text(
                        f"""
                        SELECT
                            s.sap_code,
                            {desc_expr} AS description,
                            {weight_expr} AS weight_kg,
                            {line_expr} AS line,
                            COALESCE(m.oven_no, '') AS oven_no,
                            {key_expr} AS key_code,
                            {casing_expr} AS casing_type,
                            m.last_plan_date,
                            COALESCE(m.last_workbook, '') AS last_workbook
                        FROM smds s
                        {latest_mapping}
                        {base_where}
                        {search_sql}
                        ORDER BY s.sap_code
                        LIMIT :limit OFFSET :offset
                        """
                    ),
                    params,
                ).mappings().all()

                output = []
                for row in rows:
                    item = dict(row)
                    item["tyre_size"] = _guess_size(item.get("description"))
                    missing = [
                        label
                        for label, value in (
                            ("Line", item.get("line")),
                            ("Oven", item.get("oven_no")),
                            ("Mold", item.get("key_code")),
                            ("Casing", item.get("casing_type")),
                            ("Weight", item.get("weight_kg")),
                        )
                        if not _clean(value) or _number(value) == 0 and label == "Weight"
                    ]
                    item["health"] = "OK" if not missing else f"{len(missing)} missing"
                    output.append(item)

            elif view_key == "PROCESS":
                rows = conn.execute(
                    text(
                        f"""
                        SELECT
                            s.sap_code,
                            {desc_expr} AS description,
                            {line_expr} AS line,
                            COALESCE(m.oven_no, '') AS oven_no,
                            {heel_expr} AS heel,
                            {soft_expr} AS soft,
                            {tread_expr} AS tread,
                            {curing_expr} AS curing,
                            COALESCE(m.day_plan, 0) AS day_plan,
                            COALESCE(m.night_plan, 0) AS night_plan,
                            COALESCE(m.day_produced, 0) AS day_produced,
                            COALESCE(m.night_produced, 0) AS night_produced
                        FROM smds s
                        {latest_mapping}
                        {base_where}
                        {search_sql}
                        ORDER BY s.sap_code
                        LIMIT :limit OFFSET :offset
                        """
                    ),
                    params,
                ).mappings().all()
                output = [dict(row) for row in rows]

            elif view_key == "QUALITY":
                rows = conn.execute(
                    text(
                        f"""
                        SELECT
                            s.sap_code,
                            {desc_expr} AS description,
                            {line_expr} AS line,
                            COALESCE(m.oven_no, '') AS oven_no,
                            {key_expr} AS key_code,
                            {casing_expr} AS casing_type,
                            {curing_expr} AS curing,
                            {weight_expr} AS weight_kg,
                            m.last_plan_date,
                            COALESCE(m.last_workbook, '') AS last_workbook
                        FROM smds s
                        {latest_mapping}
                        {base_where}
                        {search_sql}
                        ORDER BY s.sap_code
                        LIMIT :limit OFFSET :offset
                        """
                    ),
                    params,
                ).mappings().all()

                output = []
                for row in rows:
                    item = dict(row)
                    missing = []
                    for label, value in (
                        ("Line", item.get("line")),
                        ("Oven/Cavity", item.get("oven_no")),
                        ("Mold/Key", item.get("key_code")),
                        ("Casing", item.get("casing_type")),
                        ("Curing", item.get("curing")),
                    ):
                        if not _clean(value) or _clean(value) == "-":
                            missing.append(label)

                    if _number(item.get("weight_kg")) <= 0:
                        missing.append("Weight")

                    item["issue_count"] = len(missing)
                    item["issues"] = ", ".join(missing) if missing else "Healthy"
                    output.append(item)

            else:
                output = []

        return {
            "view_key": view_key,
            "rows": output,
            "total": total,
            "page": page,
            "page_size": page_size,
            "page_count": max(1, math.ceil(total / page_size)),
        }
