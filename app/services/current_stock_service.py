from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from sqlalchemy import text

from app.services.operational_source_service import OperationalSourceService


@dataclass(frozen=True)
class CurrentStockSource:
    import_run_id: int
    workbook_name: str
    workbook_hash: str
    plan_date: date | None
    source_path: str


class CurrentStockService:
    """Read the latest committed OVEN workbook's PROD HR:HV values directly.

    HR/HS/HT/HU/HV are treated as factory workbook truth.  The application does
    not recreate those Excel formulas.  Only the user-facing progress percentage
    is derived in Python as Current Stock / Total To be Shipped, capped at 100%.
    """

    SOURCE_SHEET = "PROD"
    SOURCE_COLUMNS = "HR:HV"

    FILTER_ALL = "All Items"
    FILTER_SHIPMENT = "Shipment Items"
    FILTER_COVERED = "Target Covered"
    FILTER_TO_PRODUCE = "To Be Produced"
    FILTER_OPTIONS = (
        FILTER_ALL,
        FILTER_SHIPMENT,
        FILTER_COVERED,
        FILTER_TO_PRODUCE,
    )

    @staticmethod
    def ensure_schema(session) -> None:
        statements = [
            """
            CREATE TABLE IF NOT EXISTS mpps_current_stock_snapshots (
                id BIGSERIAL PRIMARY KEY,
                import_run_id BIGINT NOT NULL,
                plan_date DATE,
                workbook_name TEXT NOT NULL DEFAULT '',
                workbook_hash TEXT NOT NULL DEFAULT '',
                sap_code TEXT NOT NULL,
                item_description TEXT NOT NULL DEFAULT '',
                total_to_be_shipped INTEGER NOT NULL DEFAULT 0,
                current_stock INTEGER NOT NULL DEFAULT 0,
                balance_to_produce INTEGER NOT NULL DEFAULT 0,
                total_plan INTEGER NOT NULL DEFAULT 0,
                total_to_be_plan INTEGER NOT NULL DEFAULT 0,
                source_sheet TEXT NOT NULL DEFAULT 'PROD',
                source_row INTEGER,
                created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(import_run_id, sap_code)
            )
            """,
            """
            CREATE INDEX IF NOT EXISTS ix_mpps_current_stock_snapshot_run
            ON mpps_current_stock_snapshots(import_run_id, sap_code)
            """,
            """
            CREATE INDEX IF NOT EXISTS ix_mpps_current_stock_snapshot_date
            ON mpps_current_stock_snapshots(plan_date DESC, import_run_id DESC)
            """,
        ]
        for statement in statements:
            session.execute(text(statement))

    @classmethod
    def latest_view(cls, session) -> dict[str, Any]:
        cls.ensure_schema(session)
        source = cls.ensure_latest_snapshot(session)
        if source is None:
            return {
                "source": {},
                "summary": {
                    "items": 0,
                    "total_to_be_shipped": 0,
                    "current_stock": 0,
                    "balance_to_produce": 0,
                    "total_plan": 0,
                    "total_to_be_plan": 0,
                },
                "rows": [],
            }

        rows = [
            dict(row)
            for row in session.execute(
                text(
                    """
                    SELECT
                        sap_code,
                        item_description,
                        total_to_be_shipped,
                        current_stock,
                        balance_to_produce,
                        total_plan,
                        total_to_be_plan,
                        source_row
                    FROM mpps_current_stock_snapshots
                    WHERE import_run_id = :run_id
                    ORDER BY sap_code
                    """
                ),
                {"run_id": source.import_run_id},
            ).mappings().all()
        ]

        for row in rows:
            row["progress_percent"] = cls.progress_percent(
                row.get("total_to_be_shipped"),
                row.get("current_stock"),
            )

        summary = cls.summarize_rows(rows)
        return {
            "source": {
                "import_run_id": source.import_run_id,
                "workbook_name": source.workbook_name,
                "workbook_hash": source.workbook_hash,
                "plan_date": source.plan_date.isoformat() if source.plan_date else None,
                "source_sheet": cls.SOURCE_SHEET,
                "source_columns": cls.SOURCE_COLUMNS,
            },
            "summary": summary,
            "rows": rows,
        }

    @classmethod
    def ensure_latest_snapshot(cls, session) -> CurrentStockSource | None:
        source = cls._latest_committed_source(session)
        if source is None:
            return None

        existing = session.execute(
            text(
                """
                SELECT COUNT(*)
                FROM mpps_current_stock_snapshots
                WHERE import_run_id = :run_id
                """
            ),
            {"run_id": source.import_run_id},
        ).scalar_one()
        if int(existing or 0) > 0:
            return source

        rows = cls.extract_workbook(source.source_path)
        if not rows:
            raise RuntimeError(
                "The latest committed OVEN workbook contains no readable PROD current-stock rows."
            )

        params = [
            {
                "import_run_id": source.import_run_id,
                "plan_date": source.plan_date,
                "workbook_name": source.workbook_name,
                "workbook_hash": source.workbook_hash,
                **row,
            }
            for row in rows
        ]
        session.execute(
            text(
                """
                INSERT INTO mpps_current_stock_snapshots (
                    import_run_id,
                    plan_date,
                    workbook_name,
                    workbook_hash,
                    sap_code,
                    item_description,
                    total_to_be_shipped,
                    current_stock,
                    balance_to_produce,
                    total_plan,
                    total_to_be_plan,
                    source_sheet,
                    source_row,
                    created_at,
                    updated_at
                ) VALUES (
                    :import_run_id,
                    :plan_date,
                    :workbook_name,
                    :workbook_hash,
                    :sap_code,
                    :item_description,
                    :total_to_be_shipped,
                    :current_stock,
                    :balance_to_produce,
                    :total_plan,
                    :total_to_be_plan,
                    'PROD',
                    :source_row,
                    CURRENT_TIMESTAMP,
                    CURRENT_TIMESTAMP
                )
                ON CONFLICT (import_run_id, sap_code) DO UPDATE SET
                    plan_date = EXCLUDED.plan_date,
                    workbook_name = EXCLUDED.workbook_name,
                    workbook_hash = EXCLUDED.workbook_hash,
                    item_description = EXCLUDED.item_description,
                    total_to_be_shipped = EXCLUDED.total_to_be_shipped,
                    current_stock = EXCLUDED.current_stock,
                    balance_to_produce = EXCLUDED.balance_to_produce,
                    total_plan = EXCLUDED.total_plan,
                    total_to_be_plan = EXCLUDED.total_to_be_plan,
                    source_row = EXCLUDED.source_row,
                    updated_at = CURRENT_TIMESTAMP
                """
            ),
            params,
        )
        session.flush()
        return source

    @classmethod
    def _latest_committed_source(cls, session) -> CurrentStockSource | None:
        live = OperationalSourceService.latest(session)
        if not live.import_run_id:
            return None
        row = session.execute(
            text(
                """
                SELECT
                    id,
                    workbook_name,
                    workbook_hash,
                    plan_date,
                    workbook_path,
                    archive_path
                FROM excel_import_runs
                WHERE id=:run_id
                  AND status LIKE 'COMMITTED%'
                  AND rollback_at IS NULL
                LIMIT 1
                """
            ),
            {"run_id": int(live.import_run_id)},
        ).mappings().first()
        if not row:
            return None

        candidates = [
            str(row.get("archive_path") or "").strip(),
            str(row.get("workbook_path") or "").strip(),
        ]
        project_root = Path(__file__).resolve().parents[2]
        source_path = ""
        for value in candidates:
            if not value:
                continue
            candidate = Path(value).expanduser()
            if not candidate.is_absolute():
                project_candidate = project_root / candidate
                if project_candidate.exists():
                    candidate = project_candidate
            if candidate.exists():
                source_path = str(candidate.resolve())
                break
        if not source_path:
            raise FileNotFoundError(
                "The latest committed OVEN workbook file is not available on this computer."
            )

        plan_date = row.get("plan_date")
        if isinstance(plan_date, datetime):
            plan_date = plan_date.date()
        return CurrentStockSource(
            import_run_id=int(row["id"]),
            workbook_name=str(row.get("workbook_name") or Path(source_path).name),
            workbook_hash=str(row.get("workbook_hash") or ""),
            plan_date=plan_date if isinstance(plan_date, date) else None,
            source_path=source_path,
        )

    @classmethod
    def extract_workbook(cls, workbook_path: str | Path) -> list[dict[str, Any]]:
        path = Path(workbook_path)
        if not path.exists():
            raise FileNotFoundError(path)

        workbook = load_workbook(
            path,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
        try:
            sheet_name = next(
                (name for name in workbook.sheetnames if name.strip().upper() == cls.SOURCE_SHEET),
                None,
            )
            if not sheet_name:
                raise RuntimeError("PROD sheet was not found in the OVEN workbook.")
            ws = workbook[sheet_name]

            hr = column_index_from_string("HR")
            hs = column_index_from_string("HS")
            ht = column_index_from_string("HT")
            hu = column_index_from_string("HU")
            hv = column_index_from_string("HV")
            min_col = 2  # B
            max_col = hv

            rows: list[dict[str, Any]] = []
            seen: set[str] = set()
            for row_number, values in enumerate(
                ws.iter_rows(
                    min_row=4,
                    min_col=min_col,
                    max_col=max_col,
                    values_only=True,
                ),
                start=4,
            ):
                def at(column: int):
                    index = column - min_col
                    return values[index] if 0 <= index < len(values) else None

                sap_code = cls._sap_code(at(2))
                description = cls._text(at(3))
                if not sap_code or not description or sap_code in seen:
                    continue
                seen.add(sap_code)

                rows.append(
                    {
                        "sap_code": sap_code,
                        "item_description": description,
                        "total_to_be_shipped": cls._to_int(at(hr)),
                        "current_stock": cls._to_int(at(hs)),
                        "balance_to_produce": cls._to_int(at(ht)),
                        "total_plan": cls._to_int(at(hu)),
                        "total_to_be_plan": cls._to_int(at(hv)),
                        "source_row": row_number,
                    }
                )
            return rows
        finally:
            workbook.close()

    @staticmethod
    def progress_percent(
        total_to_be_shipped: Any,
        current_stock: Any,
    ) -> float | None:
        """Return shipment coverage percentage, or None when no shipment exists."""
        target = CurrentStockService._to_int(total_to_be_shipped)
        stock = CurrentStockService._to_int(current_stock)
        if target <= 0:
            return None
        return round(max(0.0, min(100.0, (stock / target) * 100.0)), 1)

    @classmethod
    def filter_rows(
        cls,
        rows: list[dict[str, Any]],
        *,
        query: str = "",
        filter_mode: str = FILTER_ALL,
    ) -> list[dict[str, Any]]:
        """Filter the already-loaded latest snapshot in memory.

        HR/HS/HT values are never recalculated here.  Filtering only decides
        which source rows are visible.
        """
        needle = str(query or "").strip().lower()
        mode = str(filter_mode or cls.FILTER_ALL).strip()
        if mode not in cls.FILTER_OPTIONS:
            mode = cls.FILTER_ALL

        visible: list[dict[str, Any]] = []
        for row in rows:
            sap_code = str(row.get("sap_code") or "")
            description = str(row.get("item_description") or "")
            if needle and needle not in sap_code.lower() and needle not in description.lower():
                continue

            shipment = cls._to_int(row.get("total_to_be_shipped"))
            stock = cls._to_int(row.get("current_stock"))
            balance = cls._to_int(row.get("balance_to_produce"))

            if mode == cls.FILTER_SHIPMENT and shipment <= 0:
                continue
            if mode == cls.FILTER_COVERED and not (shipment > 0 and stock >= shipment):
                continue
            if mode == cls.FILTER_TO_PRODUCE and not (shipment > 0 and balance > 0):
                continue
            visible.append(row)
        return visible

    @classmethod
    def summarize_rows(cls, rows: list[dict[str, Any]]) -> dict[str, int]:
        """Build metric-card totals for the currently visible rows."""
        return {
            "items": len(rows),
            "total_to_be_shipped": sum(
                cls._to_int(row.get("total_to_be_shipped")) for row in rows
            ),
            "current_stock": sum(
                cls._to_int(row.get("current_stock")) for row in rows
            ),
            "balance_to_produce": sum(
                cls._to_int(row.get("balance_to_produce")) for row in rows
            ),
            "total_plan": sum(
                cls._to_int(row.get("total_plan")) for row in rows
            ),
            "total_to_be_plan": sum(
                cls._to_int(row.get("total_to_be_plan")) for row in rows
            ),
        }

    @staticmethod
    def _to_int(value: Any) -> int:
        try:
            return int(round(float(value or 0)))
        except Exception:
            return 0

    @staticmethod
    def _text(value: Any) -> str:
        return str(value or "").strip()

    @staticmethod
    def _sap_code(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        text_value = str(value).strip()
        if text_value.endswith(".0") and text_value[:-2].isdigit():
            return text_value[:-2]
        return text_value
