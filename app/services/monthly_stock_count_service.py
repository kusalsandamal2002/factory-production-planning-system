from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import Select, func, select
from sqlalchemy.orm import Session

from app.models import AuditLog, MonthlyStockCount, MonthlyStockCountLine


@dataclass(frozen=True)
class MonthlyStockSummary:
    total_materials: int
    total_fg: Decimal
    total_qc: Decimal
    total_balance_to_prd: Decimal
    total_final_stock: Decimal
    total_over_prd: Decimal


class MonthlyStockCountService:
    REQUIRED_HEADERS = {
        "material": "material_code",
        "material description": "material_description",
        "fg": "fg_qty",
        "qc": "qc_qty",
        "balance to prd": "balance_to_prd_qty",
        "over prd": "over_prd_qty",
    }

    OPTIONAL_HEADERS = {
        "final stock": "final_stock_qty",
    }

    EDITABLE_FIELDS = {
        "fg_qty": "FG",
        "qc_qty": "QC",
        "balance_to_prd_qty": "Balance to Over PRD",
        "over_prd_qty": "Over PRD",
    }

    def __init__(self, session: Session):
        self.session = session

    def import_excel(
        self,
        file_path: str | Path,
        month_key: str,
        stock_month_label: str,
        uploaded_by: str | None = None,
        sheet_name: str | None = None,
    ) -> MonthlyStockCount:
        path = Path(file_path)

        if not path.exists():
            raise FileNotFoundError(f"Stock Excel file not found: {path}")

        workbook = load_workbook(path, data_only=True)
        worksheet = workbook[sheet_name] if sheet_name else workbook.active

        header_row_index, header_map = self._find_header_row(worksheet)

        self._archive_existing_active_month(month_key)

        stock_count = MonthlyStockCount(
            stock_month_label=stock_month_label,
            month_key=month_key,
            file_name=path.name,
            sheet_name=worksheet.title,
            uploaded_at=datetime.utcnow(),
            uploaded_by=uploaded_by,
            total_rows=0,
            is_active=True,
            status="IMPORTED",
        )

        self.session.add(stock_count)
        self.session.flush()

        imported_rows = 0

        for row_index in range(header_row_index + 1, worksheet.max_row + 1):
            material_code = self._clean_text(
                worksheet.cell(row=row_index, column=header_map["material_code"]).value
            )

            if not material_code:
                continue

            material_description = self._clean_text(
                worksheet.cell(row=row_index, column=header_map["material_description"]).value
            )

            fg_qty = self._to_import_decimal(
                worksheet.cell(row=row_index, column=header_map["fg_qty"]).value
            )

            raw_qc_qty = self._to_import_decimal(
                worksheet.cell(row=row_index, column=header_map["qc_qty"]).value
            )

            balance_to_prd_qty = self._to_import_decimal(
                worksheet.cell(row=row_index, column=header_map["balance_to_prd_qty"]).value
            )

            over_prd_qty = self._to_import_decimal(
                worksheet.cell(row=row_index, column=header_map["over_prd_qty"]).value
            )

            if over_prd_qty > raw_qc_qty:
                over_prd_qty = raw_qc_qty

            qc_qty = raw_qc_qty - over_prd_qty

            line = MonthlyStockCountLine(
                stock_count_id=stock_count.id,
                material_code=material_code,
                material_description=material_description,
                fg_qty=fg_qty,
                qc_qty=qc_qty,
                balance_to_prd_qty=balance_to_prd_qty,
                over_prd_qty=over_prd_qty,
                source_row_number=row_index,
            )

            self.session.add(line)
            imported_rows += 1

        stock_count.total_rows = imported_rows
        self.session.flush()

        return stock_count

    def list_stock_counts(self) -> list[MonthlyStockCount]:
        statement: Select[tuple[MonthlyStockCount]] = (
            select(MonthlyStockCount)
            .order_by(
                MonthlyStockCount.month_key.desc(),
                MonthlyStockCount.uploaded_at.desc(),
            )
        )

        return list(self.session.scalars(statement).all())

    def get_active_stock_count_by_month(self, month_key: str) -> MonthlyStockCount | None:
        statement = (
            select(MonthlyStockCount)
            .where(MonthlyStockCount.month_key == month_key)
            .where(MonthlyStockCount.is_active.is_(True))
            .limit(1)
        )

        return self.session.scalar(statement)

    def get_latest_stock_count(self) -> MonthlyStockCount | None:
        statement = (
            select(MonthlyStockCount)
            .where(MonthlyStockCount.is_active.is_(True))
            .order_by(
                MonthlyStockCount.month_key.desc(),
                MonthlyStockCount.uploaded_at.desc(),
            )
            .limit(1)
        )

        return self.session.scalar(statement)

    def get_lines(
        self,
        stock_count_id: int,
        search_text: str | None = None,
    ) -> list[MonthlyStockCountLine]:
        statement = select(MonthlyStockCountLine).where(
            MonthlyStockCountLine.stock_count_id == stock_count_id
        )

        search_value = (search_text or "").strip()

        if search_value:
            pattern = f"%{search_value}%"
            statement = statement.where(
                MonthlyStockCountLine.material_code.ilike(pattern)
                | MonthlyStockCountLine.material_description.ilike(pattern)
            )

        statement = statement.order_by(MonthlyStockCountLine.material_code.asc())

        return list(self.session.scalars(statement).all())

    def update_stock_value(
        self,
        line_id: int,
        field_name: str,
        new_value: Decimal | int | float | str,
        username: str | None = None,
    ) -> MonthlyStockCountLine:
        if field_name not in self.EDITABLE_FIELDS:
            raise ValueError("This stock field cannot be edited.")

        line = self.session.get(MonthlyStockCountLine, line_id)
        if line is None:
            raise ValueError("Stock line not found.")

        stock_count = self.session.get(MonthlyStockCount, line.stock_count_id)
        if stock_count is None:
            raise ValueError("Stock count header not found.")

        value = self._to_edit_decimal(new_value)

        old_value = self._to_import_decimal(getattr(line, field_name))
        old_qc_qty = self._to_import_decimal(line.qc_qty)
        old_final_stock = self._calculate_final_stock(
            fg_qty=line.fg_qty,
            qc_qty=line.qc_qty,
            balance_to_prd_qty=line.balance_to_prd_qty,
        )

        if old_value == value:
            return line

        if field_name == "over_prd_qty":
            old_over_prd_qty = self._to_import_decimal(line.over_prd_qty)
            current_qc_qty = self._to_import_decimal(line.qc_qty)

            over_prd_difference = value - old_over_prd_qty
            new_qc_qty = current_qc_qty - over_prd_difference

            if new_qc_qty < 0:
                raise ValueError("Over PRD cannot be greater than available QC quantity.")

            line.over_prd_qty = value
            line.qc_qty = new_qc_qty
            line.over_prd_updated_at = datetime.utcnow()

        else:
            setattr(line, field_name, value)

        self.session.flush()
        self.session.refresh(line)

        new_qc_qty = self._to_import_decimal(line.qc_qty)
        new_final_stock = self._to_import_decimal(line.final_stock_qty)

        self._write_stock_edit_audit_log(
            username=username or "system",
            stock_count=stock_count,
            line=line,
            field_name=field_name,
            old_value=old_value,
            new_value=value,
            old_qc_qty=old_qc_qty,
            new_qc_qty=new_qc_qty,
            old_final_stock=old_final_stock,
            new_final_stock=new_final_stock,
        )

        self.session.flush()

        return line

    def update_over_prd(
        self,
        line_id: int,
        over_prd_qty: Decimal | int | float | str,
        username: str | None = None,
    ) -> MonthlyStockCountLine:
        return self.update_stock_value(
            line_id=line_id,
            field_name="over_prd_qty",
            new_value=over_prd_qty,
            username=username,
        )

    def get_summary(self, stock_count_id: int) -> MonthlyStockSummary:
        statement = (
            select(
                func.count(MonthlyStockCountLine.id),
                func.coalesce(func.sum(MonthlyStockCountLine.fg_qty), 0),
                func.coalesce(func.sum(MonthlyStockCountLine.qc_qty), 0),
                func.coalesce(func.sum(MonthlyStockCountLine.balance_to_prd_qty), 0),
                func.coalesce(func.sum(MonthlyStockCountLine.final_stock_qty), 0),
                func.coalesce(func.sum(MonthlyStockCountLine.over_prd_qty), 0),
            )
            .where(MonthlyStockCountLine.stock_count_id == stock_count_id)
        )

        row = self.session.execute(statement).one()

        return MonthlyStockSummary(
            total_materials=int(row[0] or 0),
            total_fg=self._to_import_decimal(row[1]),
            total_qc=self._to_import_decimal(row[2]),
            total_balance_to_prd=self._to_import_decimal(row[3]),
            total_final_stock=self._to_import_decimal(row[4]),
            total_over_prd=self._to_import_decimal(row[5]),
        )

    def is_latest_active_month(self, stock_count: MonthlyStockCount) -> bool:
        latest = self.get_latest_stock_count()

        return latest is not None and latest.id == stock_count.id

    def _archive_existing_active_month(self, month_key: str) -> None:
        existing_records = self.session.scalars(
            select(MonthlyStockCount)
            .where(MonthlyStockCount.month_key == month_key)
            .where(MonthlyStockCount.is_active.is_(True))
        ).all()

        for record in existing_records:
            record.is_active = False
            record.status = "ARCHIVED"
            record.updated_at = datetime.utcnow()

        self.session.flush()

    def _write_stock_edit_audit_log(
        self,
        username: str,
        stock_count: MonthlyStockCount,
        line: MonthlyStockCountLine,
        field_name: str,
        old_value: Decimal,
        new_value: Decimal,
        old_qc_qty: Decimal,
        new_qc_qty: Decimal,
        old_final_stock: Decimal,
        new_final_stock: Decimal,
    ) -> None:
        field_label = self.EDITABLE_FIELDS[field_name]

        old_payload = {
            "stock_month": stock_count.stock_month_label,
            "month_key": stock_count.month_key,
            "material_code": line.material_code,
            "material_description": line.material_description,
            "field": field_label,
            "field_old_value": self._decimal_to_text(old_value),
            "qc_qty": self._decimal_to_text(old_qc_qty),
            "final_stock": self._decimal_to_text(old_final_stock),
        }

        new_payload = {
            "stock_month": stock_count.stock_month_label,
            "month_key": stock_count.month_key,
            "material_code": line.material_code,
            "material_description": line.material_description,
            "field": field_label,
            "field_new_value": self._decimal_to_text(new_value),
            "qc_qty": self._decimal_to_text(new_qc_qty),
            "final_stock": self._decimal_to_text(new_final_stock),
        }

        if field_name == "over_prd_qty":
            note = (
                f"Over PRD changed for {line.material_code} "
                f"from {self._decimal_to_text(old_value)} "
                f"to {self._decimal_to_text(new_value)}. "
                f"QC auto adjusted from {self._decimal_to_text(old_qc_qty)} "
                f"to {self._decimal_to_text(new_qc_qty)}."
            )
        else:
            note = (
                f"{field_label} changed for {line.material_code} "
                f"from {self._decimal_to_text(old_value)} "
                f"to {self._decimal_to_text(new_value)}."
            )

        audit_log = AuditLog(
            action_timestamp=datetime.utcnow(),
            username=username,
            action_type="MONTHLY_STOCK_EDIT",
            table_name="monthly_stock_count_lines",
            record_id=str(line.id),
            old_values=json.dumps(old_payload, ensure_ascii=False),
            new_values=json.dumps(new_payload, ensure_ascii=False),
            note=note,
        )

        self.session.add(audit_log)

    def _find_header_row(self, worksheet: Any) -> tuple[int, dict[str, int]]:
        max_scan_rows = min(20, worksheet.max_row)

        for row_index in range(1, max_scan_rows + 1):
            header_map: dict[str, int] = {}

            for column_index in range(1, worksheet.max_column + 1):
                raw_value = worksheet.cell(row=row_index, column=column_index).value
                normalized = self._normalize_header(raw_value)

                if normalized in self.REQUIRED_HEADERS:
                    header_map[self.REQUIRED_HEADERS[normalized]] = column_index
                elif normalized in self.OPTIONAL_HEADERS:
                    header_map[self.OPTIONAL_HEADERS[normalized]] = column_index

            missing = [
                field_name
                for field_name in self.REQUIRED_HEADERS.values()
                if field_name not in header_map
            ]

            if not missing:
                return row_index, header_map

        required_names = ", ".join(self.REQUIRED_HEADERS.keys())

        raise ValueError(
            "Invalid stock Excel file. Required headers not found. "
            f"Required headers: {required_names}"
        )

    @staticmethod
    def _normalize_header(value: Any) -> str:
        text = str(value or "").strip().lower()

        return " ".join(text.replace("\n", " ").split())

    @staticmethod
    def _clean_text(value: Any) -> str:
        if value is None:
            return ""

        if isinstance(value, float) and value.is_integer():
            return str(int(value))

        return str(value).strip()

    @staticmethod
    def _to_import_decimal(value: Any) -> Decimal:
        if value is None or value == "":
            return Decimal("0.000")

        if isinstance(value, Decimal):
            number = value
        else:
            try:
                text = str(value).strip().replace(",", "")

                if text == "":
                    return Decimal("0.000")

                number = Decimal(text)
            except (InvalidOperation, ValueError):
                return Decimal("0.000")

        if number < 0:
            return Decimal("0.000")

        return number.quantize(Decimal("0.001"))

    @staticmethod
    def _to_edit_decimal(value: Any) -> Decimal:
        if value is None or value == "":
            return Decimal("0.000")

        try:
            text = str(value).strip().replace(",", "")

            if text == "":
                return Decimal("0.000")

            number = Decimal(text).quantize(Decimal("0.001"))
        except (InvalidOperation, ValueError) as exc:
            raise ValueError("Please enter a valid number.") from exc

        if number < 0:
            raise ValueError("Stock value cannot be negative.")

        return number

    @staticmethod
    def _calculate_final_stock(
        fg_qty: Decimal | int | float | str | None,
        qc_qty: Decimal | int | float | str | None,
        balance_to_prd_qty: Decimal | int | float | str | None,
    ) -> Decimal:
        fg = MonthlyStockCountService._to_import_decimal(fg_qty)
        qc = MonthlyStockCountService._to_import_decimal(qc_qty)
        balance = MonthlyStockCountService._to_import_decimal(balance_to_prd_qty)

        return fg + qc + balance

    @staticmethod
    def _decimal_to_text(value: Decimal | int | float | str | None) -> str:
        if value is None:
            return "0"

        try:
            decimal_value = Decimal(str(value)).quantize(Decimal("0.001"))
        except (InvalidOperation, ValueError):
            return "0"

        text = format(decimal_value, "f")

        if "." in text:
            text = text.rstrip("0").rstrip(".")

        return text or "0"