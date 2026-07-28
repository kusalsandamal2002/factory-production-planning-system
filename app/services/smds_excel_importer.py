from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import text

from app.database import engine
from app.services.master_data_normalization import (
    normalize_casing_type,
    normalize_line_name,
    normalize_mold_key,
    normalize_sap_code,
)
from app.services.smds_curing_time import refresh_smds_curing_time_columns
from app.services.smds_schema import ensure_smds_table


EXPECTED_HEADERS = {
    "sap code": "sap_code",
    "material description": "material_description",
    "line": "line",
    "heel": "heel",
    "soft": "soft",
    "tred": "tred",
    "remark": "remark",
    "weight per tyre (kg)": "weight_per_tyre_kg",
    "line-400": "line_400",
    "line-800": "line_800",
    "press -line": "press_line",
    "nancy press": "nancy_press",
    "400 t press": "press_400_t",
    "t 600 -01 press": "t_600_01_press",
    "t 600 -02 press": "t_600_02_press",
    "l-press-1250": "l_press_1250",
    "l-press-1500": "l_press_1500",
    "l-press-1800": "l_press_1800",
    "oring-press": "oring_press",
    "new press": "new_press",
    "key code": "key_code",
    "casing type": "casing_type",
    "curing cycle": "curing_cycle",
    "handling time": "handling_time",
    "day plan": "day_plan",
    "night plan": "night_plan",
    "total plan": "total_plan",
}

NUMERIC_COLUMNS = {
    "weight_per_tyre_kg",
    "handling_time",
    "day_plan",
    "night_plan",
    "total_plan",
}

INSERT_COLUMNS = [
    "sap_code",
    "material_description",
    "line",
    "heel",
    "soft",
    "tred",
    "remark",
    "weight_per_tyre_kg",
    "line_400",
    "line_800",
    "press_line",
    "nancy_press",
    "press_400_t",
    "t_600_01_press",
    "t_600_02_press",
    "l_press_1250",
    "l_press_1500",
    "l_press_1800",
    "oring_press",
    "new_press",
    "key_code",
    "casing_type",
    "curing_cycle",
    "handling_time",
    "day_plan",
    "night_plan",
    "total_plan",
    "planning_manager_approval_status",
    "source_file",
    "source_sheet",
    "source_row_number",
]


@dataclass(frozen=True)
class SMDSImportResult:
    file_path: str
    sheet_name: str
    imported_rows: int
    skipped_rows: int


def _normalise_header(value: Any) -> str:
    return " ".join(
        str(value or "").strip().lower().split()
    )


def _text_value(value: Any) -> str | None:
    if value is None:
        return None

    if isinstance(value, float) and value.is_integer():
        text_value = str(int(value))
    else:
        text_value = str(value).strip()

    return text_value or None


def _sap_code(value: Any) -> str | None:
    text_value = _text_value(value)

    if text_value is None:
        return None

    if (
        text_value.endswith(".0")
        and text_value[:-2].isdigit()
    ):
        return text_value[:-2]

    return normalize_sap_code(
        text_value
    )


def _decimal_value(value: Any) -> Decimal | None:
    text_value = _text_value(value)

    if text_value is None or text_value == "-":
        return None

    try:
        return Decimal(
            text_value.replace(",", "")
        )
    except (InvalidOperation, ValueError):
        return None


def _extract_header_map(
    headers: list[Any],
) -> dict[int, str]:
    header_map: dict[int, str] = {}

    for index, header in enumerate(headers):
        normalised = _normalise_header(header)
        column_name = EXPECTED_HEADERS.get(normalised)

        if column_name:
            header_map[index] = column_name

    missing_headers = sorted(
        set(EXPECTED_HEADERS.values())
        - set(header_map.values())
    )

    if missing_headers:
        raise ValueError(
            "SMDS file is missing required columns: "
            + ", ".join(missing_headers)
        )

    return header_map


def _row_to_record(
    row_values: tuple[Any, ...],
    header_map: dict[int, str],
    source_file: str,
    source_sheet: str,
    source_row_number: int,
) -> dict[str, Any] | None:
    record: dict[str, Any] = {
        column: None for column in INSERT_COLUMNS
    }

    for index, column_name in header_map.items():
        value = (
            row_values[index]
            if index < len(row_values)
            else None
        )

        if column_name == "sap_code":
            record[column_name] = _sap_code(value)
        elif column_name in NUMERIC_COLUMNS:
            record[column_name] = _decimal_value(
                value
            )
        else:
            record[column_name] = _text_value(
                value
            )

    if not record.get("sap_code"):
        return None

    record["sap_code"] = normalize_sap_code(
        record.get("sap_code")
    )
    record["key_code"] = normalize_mold_key(
        record.get("key_code")
    )
    record["casing_type"] = (
        normalize_casing_type(
            record.get("casing_type")
        )
    )
    record["line"] = normalize_line_name(
        record.get("line"),
        unknown_value="-",
    )
    record["material_description"] = (
        record.get("material_description") or ""
    )
    record["planning_manager_approval_status"] = (
        "Pending"
    )
    record["source_file"] = source_file
    record["source_sheet"] = source_sheet
    record["source_row_number"] = source_row_number
    return record


def _load_existing_approval_statuses() -> dict[str, str]:
    with engine.connect() as conn:
        rows = conn.execute(
            text(
                """
                SELECT
                    sap_code,
                    planning_manager_approval_status
                FROM smds
                """
            )
        ).mappings()

        return {
            normalize_sap_code(
                row["sap_code"]
            ): str(
                row[
                    "planning_manager_approval_status"
                ]
                or "Pending"
            ).strip()
            or "Pending"
            for row in rows
            if row["sap_code"]
        }


def import_smds_workbook(
    file_path: str | Path,
    sheet_name: str = "ALL",
    replace: bool = True,
) -> SMDSImportResult:
    """Import SMDS rows.

    New SAP codes always start as Pending. Existing manager approval
    decisions are preserved when the same SAP code is re-imported.
    """
    ensure_smds_table()
    existing_approvals = (
        _load_existing_approval_statuses()
    )

    path = Path(file_path).expanduser().resolve()

    if not path.exists():
        raise FileNotFoundError(
            f"SMDS file not found: {path}"
        )

    workbook = load_workbook(
        path,
        data_only=True,
        read_only=True,
    )

    if sheet_name not in workbook.sheetnames:
        available = ", ".join(
            workbook.sheetnames
        )
        raise ValueError(
            f"Sheet '{sheet_name}' not found. "
            f"Available sheets: {available}"
        )

    sheet = workbook[sheet_name]
    rows_iter = sheet.iter_rows(values_only=True)

    try:
        headers = list(next(rows_iter))
    except StopIteration as exc:
        raise ValueError(
            "SMDS sheet is empty."
        ) from exc

    header_map = _extract_header_map(headers)

    records: list[dict[str, Any]] = []
    skipped_rows = 0

    for row_number, row_values in enumerate(
        rows_iter,
        start=2,
    ):
        record = _row_to_record(
            row_values,
            header_map,
            path.name,
            sheet_name,
            row_number,
        )

        if record is None:
            skipped_rows += 1
            continue

        sap_code = normalize_sap_code(
            record["sap_code"]
        )
        record[
            "planning_manager_approval_status"
        ] = existing_approvals.get(
            sap_code,
            "Pending",
        )
        records.append(record)

    placeholders = ", ".join(
        f":{column}" for column in INSERT_COLUMNS
    )
    column_sql = ", ".join(INSERT_COLUMNS)

    update_sql = ",\n                    ".join(
        f"{column} = EXCLUDED.{column}"
        for column in INSERT_COLUMNS
        if column
        not in {
            "sap_code",
            "planning_manager_approval_status",
        }
    )

    insert_sql = text(
        f"""
        INSERT INTO smds ({column_sql})
        VALUES ({placeholders})
        ON CONFLICT (sap_code) DO UPDATE SET
                    {update_sql},
                    updated_at = CURRENT_TIMESTAMP,
                    imported_at = CURRENT_TIMESTAMP
        """
    )

    with engine.begin() as conn:
        if replace:
            conn.execute(
                text(
                    "TRUNCATE TABLE smds "
                    "RESTART IDENTITY"
                )
            )

        if records:
            conn.execute(insert_sql, records)

    refresh_smds_curing_time_columns()

    return SMDSImportResult(
        file_path=str(path),
        sheet_name=sheet_name,
        imported_rows=len(records),
        skipped_rows=skipped_rows,
    )

