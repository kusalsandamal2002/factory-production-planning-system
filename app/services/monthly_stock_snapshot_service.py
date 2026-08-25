from __future__ import annotations

from collections import defaultdict
from datetime import date, timedelta
from hashlib import sha256
import json
import math
from statistics import median
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from sqlalchemy import text


class MonthlyStockSnapshotService:
    """Build and read month-by-month stock truth from committed OVEN workbooks.

    Factory semantics:
    * PROD D/E/F in a workbook are the FINAL snapshot for the previous month.
    * PROD HS/E/F are the LIVE snapshot for the workbook's own month.
    * A FINAL snapshot always outranks LIVE for the same target month.
    * Same plan-date + same stock payload is ignored. A changed payload is kept as
      a new revision and the latest committed import wins while old evidence stays
      available for traceability.

    ML output is advisory only. Official stock values are never modified by the
    forecast/trend model.
    """

    MODEL_VERSION = "MONTHLY-STOCK-LINEAR-V1"

    @staticmethod
    def ensure_schema(session) -> None:
        statements = [
            """
            CREATE TABLE IF NOT EXISTS mpps_monthly_stock_snapshots (
                id BIGSERIAL PRIMARY KEY,
                import_run_id BIGINT NOT NULL,
                month_key VARCHAR(7) NOT NULL,
                source_kind VARCHAR(10) NOT NULL,
                source_plan_date DATE NOT NULL,
                import_mode VARCHAR(20) NOT NULL DEFAULT 'HISTORICAL',
                workbook_name TEXT NOT NULL DEFAULT '',
                workbook_hash TEXT NOT NULL DEFAULT '',
                data_hash TEXT NOT NULL,
                source_columns VARCHAR(30) NOT NULL,
                row_count INTEGER NOT NULL DEFAULT 0,
                created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                CHECK (source_kind IN ('LIVE', 'FINAL')),
                CHECK (month_key ~ '^[0-9]{4}-[0-9]{2}$'),
                UNIQUE(import_run_id, month_key, source_kind)
            )
            """,
            """
            CREATE INDEX IF NOT EXISTS ix_monthly_stock_snapshot_authority
            ON mpps_monthly_stock_snapshots(
                month_key,
                source_kind,
                source_plan_date DESC,
                import_run_id DESC,
                id DESC
            )
            """,
            """
            CREATE INDEX IF NOT EXISTS ix_monthly_stock_snapshot_dedupe
            ON mpps_monthly_stock_snapshots(
                month_key,
                source_kind,
                source_plan_date,
                data_hash
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS mpps_monthly_stock_snapshot_lines (
                id BIGSERIAL PRIMARY KEY,
                snapshot_id BIGINT NOT NULL
                    REFERENCES mpps_monthly_stock_snapshots(id)
                    ON DELETE CASCADE,
                sap_code TEXT NOT NULL,
                item_description TEXT NOT NULL DEFAULT '',
                total_stock INTEGER NOT NULL DEFAULT 0,
                scrap_qty INTEGER NOT NULL DEFAULT 0,
                blocked_qty INTEGER NOT NULL DEFAULT 0,
                source_row INTEGER,
                created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(snapshot_id, sap_code)
            )
            """,
            """
            CREATE INDEX IF NOT EXISTS ix_monthly_stock_lines_sap
            ON mpps_monthly_stock_snapshot_lines(sap_code, snapshot_id)
            """,
        ]
        for statement in statements:
            session.execute(text(statement))

    @staticmethod
    def _get(value: Any, key: str, default: Any = None) -> Any:
        if isinstance(value, dict):
            return value.get(key, default)
        return getattr(value, key, default)

    @staticmethod
    def _int(value: Any) -> int:
        try:
            return int(round(float(value or 0)))
        except Exception:
            return 0

    @staticmethod
    def previous_month_key(plan_date: date) -> str:
        first = plan_date.replace(day=1)
        return (first - timedelta(days=1)).strftime("%Y-%m")

    @staticmethod
    def current_month_key(plan_date: date) -> str:
        return plan_date.strftime("%Y-%m")

    @classmethod
    def _build_lines(
        cls,
        stock_rows: list[dict[str, Any]],
        *,
        source_kind: str,
    ) -> list[dict[str, Any]]:
        rows: dict[str, dict[str, Any]] = {}
        for source in stock_rows:
            sap = str(source.get("sap_code") or "").strip().upper()
            if not sap:
                continue
            if source_kind == "FINAL":
                total = cls._int(
                    source.get("opening_stock_qty", source.get("fg_stock", 0))
                )
            else:
                total = cls._int(source.get("total_available", 0))
            rows.setdefault(
                sap,
                {
                    "sap_code": sap,
                    "item_description": str(source.get("description") or "").strip(),
                    "total_stock": total,
                    "scrap_qty": cls._int(source.get("scrap_stock", 0)),
                    "blocked_qty": cls._int(source.get("blocked_stock", 0)),
                    "source_row": cls._int(source.get("source_row", 0)) or None,
                },
            )
        return [rows[key] for key in sorted(rows)]

    @staticmethod
    def _data_hash(lines: list[dict[str, Any]]) -> str:
        payload = [
            {
                "sap": row["sap_code"],
                "description": row["item_description"],
                "stock": int(row["total_stock"]),
                "scrap": int(row["scrap_qty"]),
                "block": int(row["blocked_qty"]),
            }
            for row in lines
        ]
        encoded = json.dumps(
            payload,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        ).encode("utf-8")
        return sha256(encoded).hexdigest()

    @classmethod
    def capture_import(
        cls,
        session,
        *,
        import_run_id: int,
        analysis: Any,
        import_mode: str,
    ) -> dict[str, Any]:
        cls.ensure_schema(session)
        raw_plan_date = cls._get(analysis, "plan_date")
        if isinstance(raw_plan_date, date):
            plan_date = raw_plan_date
        elif raw_plan_date:
            plan_date = date.fromisoformat(str(raw_plan_date)[:10])
        else:
            raise ValueError("Monthly stock capture requires a workbook plan date.")

        stock_rows = list(cls._get(analysis, "stock_rows", []) or [])
        workbook_name = str(cls._get(analysis, "workbook_name", "") or "")
        workbook_hash = str(cls._get(analysis, "workbook_hash", "") or "")

        specs = (
            (
                cls.previous_month_key(plan_date),
                "FINAL",
                "PROD D/E/F",
            ),
            (
                cls.current_month_key(plan_date),
                "LIVE",
                "PROD HS/E/F",
            ),
        )

        created = 0
        duplicates = 0
        saved_lines = 0
        skipped_empty = 0
        target_months: list[str] = []

        for month_key, source_kind, source_columns in specs:
            lines = cls._build_lines(stock_rows, source_kind=source_kind)
            target_months.append(month_key)
            if not lines:
                skipped_empty += 1
                continue
            data_hash = cls._data_hash(lines)

            duplicate_id = session.execute(
                text(
                    """
                    SELECT id
                    FROM mpps_monthly_stock_snapshots
                    WHERE month_key = :month_key
                      AND source_kind = :source_kind
                      AND source_plan_date = :source_plan_date
                      AND data_hash = :data_hash
                    ORDER BY id DESC
                    LIMIT 1
                    """
                ),
                {
                    "month_key": month_key,
                    "source_kind": source_kind,
                    "source_plan_date": plan_date,
                    "data_hash": data_hash,
                },
            ).scalar()
            if duplicate_id is not None:
                duplicates += 1
                continue

            snapshot_id = session.execute(
                text(
                    """
                    INSERT INTO mpps_monthly_stock_snapshots (
                        import_run_id,
                        month_key,
                        source_kind,
                        source_plan_date,
                        import_mode,
                        workbook_name,
                        workbook_hash,
                        data_hash,
                        source_columns,
                        row_count
                    ) VALUES (
                        :import_run_id,
                        :month_key,
                        :source_kind,
                        :source_plan_date,
                        :import_mode,
                        :workbook_name,
                        :workbook_hash,
                        :data_hash,
                        :source_columns,
                        :row_count
                    )
                    RETURNING id
                    """
                ),
                {
                    "import_run_id": int(import_run_id),
                    "month_key": month_key,
                    "source_kind": source_kind,
                    "source_plan_date": plan_date,
                    "import_mode": str(import_mode or "HISTORICAL").upper(),
                    "workbook_name": workbook_name,
                    "workbook_hash": workbook_hash,
                    "data_hash": data_hash,
                    "source_columns": source_columns,
                    "row_count": len(lines),
                },
            ).scalar_one()

            if lines:
                session.execute(
                    text(
                        """
                        INSERT INTO mpps_monthly_stock_snapshot_lines (
                            snapshot_id,
                            sap_code,
                            item_description,
                            total_stock,
                            scrap_qty,
                            blocked_qty,
                            source_row
                        ) VALUES (
                            :snapshot_id,
                            :sap_code,
                            :item_description,
                            :total_stock,
                            :scrap_qty,
                            :blocked_qty,
                            :source_row
                        )
                        """
                    ),
                    [
                        {"snapshot_id": int(snapshot_id), **row}
                        for row in lines
                    ],
                )
            created += 1
            saved_lines += len(lines)

        return {
            "monthly_stock_snapshots_created": created,
            "monthly_stock_snapshots_duplicate": duplicates,
            "monthly_stock_rows_saved": saved_lines,
            "monthly_stock_snapshots_skipped_empty": skipped_empty,
            "monthly_stock_target_months": target_months,
        }

    @classmethod
    def bootstrap_from_committed_imports(
        cls,
        session,
        *,
        max_months: int = 24,
    ) -> dict[str, int]:
        """Backfill representative committed workbooks already known by MPPS.

        Only the newest committed workbook for each plan month is considered. This
        makes the first Monthly Stock page useful without asking the user to
        re-upload old daily OVEN files, while keeping startup work bounded.
        """
        cls.ensure_schema(session)
        try:
            candidates = session.execute(
                text(
                    """
                    WITH monthly AS (
                        SELECT DISTINCT ON (TO_CHAR(plan_date, 'YYYY-MM'))
                               id,
                               plan_date,
                               workbook_name,
                               workbook_hash,
                               archive_path,
                               workbook_path,
                               status
                        FROM excel_import_runs
                        WHERE plan_date IS NOT NULL
                          AND status LIKE 'COMMITTED%'
                          AND rollback_at IS NULL
                        ORDER BY TO_CHAR(plan_date, 'YYYY-MM'), plan_date DESC, id DESC
                    )
                    SELECT m.*
                    FROM monthly m
                    WHERE NOT EXISTS (
                        SELECT 1
                        FROM mpps_monthly_stock_snapshots s
                        WHERE s.import_run_id = m.id
                    )
                    ORDER BY m.plan_date DESC, m.id DESC
                    LIMIT :limit
                    """
                ),
                {"limit": max(1, min(60, int(max_months)))},
            ).mappings().all()
        except Exception:
            return {"monthly_stock_bootstrap_files": 0, "monthly_stock_bootstrap_rows": 0}

        files = 0
        rows_saved = 0
        project_root = Path(__file__).resolve().parents[2]
        try:
            latest_plan_date = session.execute(
                text(
                    """
                    SELECT MAX(plan_date)
                    FROM excel_import_runs
                    WHERE plan_date IS NOT NULL
                      AND status LIKE 'COMMITTED%'
                      AND rollback_at IS NULL
                    """
                )
            ).scalar()
        except Exception:
            latest_plan_date = max(
                (row.get("plan_date") for row in candidates if row.get("plan_date")),
                default=None,
            )

        for candidate in reversed(list(candidates)):
            path = cls._resolve_workbook_path(candidate, project_root)
            if path is None:
                continue
            stock_rows = cls._read_prod_stock_rows(path)
            if not stock_rows:
                continue
            plan_date = candidate.get("plan_date")
            if not isinstance(plan_date, date):
                try:
                    plan_date = date.fromisoformat(str(plan_date)[:10])
                except Exception:
                    continue
            result = cls.capture_import(
                session,
                import_run_id=int(candidate["id"]),
                analysis={
                    "plan_date": plan_date.isoformat(),
                    "workbook_name": str(candidate.get("workbook_name") or path.name),
                    "workbook_hash": str(candidate.get("workbook_hash") or ""),
                    "stock_rows": stock_rows,
                },
                import_mode="LIVE" if plan_date == latest_plan_date else "HISTORICAL",
            )
            files += 1
            rows_saved += int(result.get("monthly_stock_rows_saved") or 0)

        return {
            "monthly_stock_bootstrap_files": files,
            "monthly_stock_bootstrap_rows": rows_saved,
        }

    @staticmethod
    def _resolve_workbook_path(candidate: Any, project_root: Path) -> Path | None:
        for key in ("archive_path", "workbook_path"):
            raw = str(candidate.get(key) or "").strip()
            if not raw:
                continue
            path = Path(raw).expanduser()
            attempts = [path]
            if not path.is_absolute():
                attempts.append(project_root / path)
            for attempt in attempts:
                try:
                    resolved = attempt.resolve()
                except Exception:
                    resolved = attempt
                if resolved.exists() and resolved.suffix.lower() in {".xlsx", ".xlsm"}:
                    return resolved
        return None

    @classmethod
    def _read_prod_stock_rows(cls, path: Path) -> list[dict[str, Any]]:
        workbook = load_workbook(
            path,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
        try:
            sheet_name = next(
                (name for name in workbook.sheetnames if str(name).strip().upper() == "PROD"),
                None,
            )
            if sheet_name is None:
                return []
            ws = workbook[sheet_name]
            hs_col = column_index_from_string("HS")
            min_col = 2  # B
            row2 = next(
                ws.iter_rows(
                    min_row=2,
                    max_row=2,
                    min_col=min_col,
                    max_col=hs_col,
                    values_only=True,
                ),
                tuple(),
            )
            p_indexes = [
                index
                for index, value in enumerate(row2)
                if str(value or "").strip().upper() == "P"
            ]
            hs_index = hs_col - min_col
            result: list[dict[str, Any]] = []
            for row_number, values in enumerate(
                ws.iter_rows(
                    min_row=4,
                    min_col=min_col,
                    max_col=hs_col,
                    values_only=True,
                ),
                start=4,
            ):
                sap = str(values[0] or "").strip()
                description = str(values[1] or "").strip() if len(values) > 1 else ""
                if not sap or not description:
                    continue
                if sap.endswith(".0") and sap[:-2].isdigit():
                    sap = sap[:-2]
                opening = cls._int(values[2] if len(values) > 2 else 0)
                scrap = cls._int(values[3] if len(values) > 3 else 0)
                blocked = cls._int(values[4] if len(values) > 4 else 0)
                hs_value = values[hs_index] if hs_index < len(values) else None
                if hs_value in (None, ""):
                    hs_value = (
                        sum(
                            cls._int(values[index])
                            for index in p_indexes
                            if index < len(values)
                        )
                        - scrap
                        - blocked
                    )
                result.append(
                    {
                        "sap_code": sap.upper(),
                        "description": description,
                        "opening_stock_qty": opening,
                        "fg_stock": opening,
                        "scrap_stock": scrap,
                        "blocked_stock": blocked,
                        "total_available": cls._int(hs_value),
                        "source_row": row_number,
                    }
                )
            return result
        finally:
            workbook.close()

    @classmethod
    def list_month_keys(cls, session) -> list[str]:
        cls.ensure_schema(session)
        values = session.execute(
            text(
                """
                SELECT DISTINCT month_key
                FROM mpps_monthly_stock_snapshots
                ORDER BY month_key DESC
                """
            )
        ).scalars().all()
        return [str(value) for value in values if value]

    @classmethod
    def _authoritative_snapshot(cls, session, month_key: str) -> dict[str, Any] | None:
        row = session.execute(
            text(
                """
                SELECT
                    id,
                    import_run_id,
                    month_key,
                    source_kind,
                    source_plan_date,
                    import_mode,
                    workbook_name,
                    workbook_hash,
                    data_hash,
                    source_columns,
                    row_count,
                    created_at
                FROM mpps_monthly_stock_snapshots
                WHERE month_key = :month_key
                ORDER BY
                    CASE WHEN source_kind = 'FINAL' THEN 0 ELSE 1 END,
                    source_plan_date DESC,
                    import_run_id DESC,
                    id DESC
                LIMIT 1
                """
            ),
            {"month_key": month_key},
        ).mappings().first()
        return dict(row) if row else None

    @staticmethod
    def predict_stock(history: list[int], current: int) -> dict[str, Any]:
        """Small local linear model + robust anomaly detector.

        It is intentionally dependency-free so monthly intelligence continues to
        work even when optional ML packages are not installed.
        """
        previous = [int(value) for value in history[-6:]]
        series = previous + [int(current)]
        n = len(series)

        if n < 2:
            return {
                "forecast": int(current),
                "trend": "LEARNING",
                "risk": "LEARNING",
                "anomaly": False,
                "confidence": 0.20,
                "samples": n,
            }

        x_values = list(range(n))
        x_mean = sum(x_values) / n
        y_mean = sum(series) / n
        denominator = sum((x - x_mean) ** 2 for x in x_values)
        slope = (
            sum((x - x_mean) * (y - y_mean) for x, y in zip(x_values, series))
            / denominator
            if denominator
            else 0.0
        )
        intercept = y_mean - slope * x_mean
        forecast = int(round(max(0.0, intercept + slope * n)))

        threshold = max(1.0, abs(y_mean) * 0.04)
        if slope > threshold:
            trend = "UP"
        elif slope < -threshold:
            trend = "DOWN"
        else:
            trend = "STABLE"

        anomaly = False
        downward_anomaly = False
        if len(previous) >= 3:
            baseline = float(median(previous))
            deviations = [abs(value - baseline) for value in previous]
            mad = float(median(deviations))
            anomaly_limit = max(5.0, 3.0 * mad, abs(baseline) * 0.45)
            delta = float(current) - baseline
            anomaly = abs(delta) > anomaly_limit
            downward_anomaly = anomaly and delta < 0
            if anomaly:
                trend = f"ANOMALY {trend}"

        if current <= 0 or forecast <= 0:
            risk = "HIGH"
        elif downward_anomaly:
            risk = "HIGH"
        elif "DOWN" in trend:
            if forecast <= max(5, int(round(max(current, 1) * 0.50))):
                risk = "HIGH"
            else:
                risk = "MEDIUM"
        else:
            risk = "LOW"

        confidence = min(0.95, 0.20 + 0.12 * n)
        if not math.isfinite(confidence):
            confidence = 0.20
        return {
            "forecast": forecast,
            "trend": trend,
            "risk": risk,
            "anomaly": anomaly,
            "confidence": round(confidence, 2),
            "samples": n,
        }

    @classmethod
    def month_view(cls, session, month_key: str) -> dict[str, Any]:
        cls.ensure_schema(session)
        snapshot = cls._authoritative_snapshot(session, month_key)
        if snapshot is None:
            return {
                "month_key": month_key,
                "source": None,
                "rows": [],
                "summary": {
                    "items": 0,
                    "total_stock": 0,
                    "scrap": 0,
                    "blocked": 0,
                    "high_risk": 0,
                },
                "model_version": cls.MODEL_VERSION,
            }

        lines = [
            dict(row)
            for row in session.execute(
                text(
                    """
                    SELECT
                        sap_code,
                        item_description,
                        total_stock,
                        scrap_qty,
                        blocked_qty,
                        source_row
                    FROM mpps_monthly_stock_snapshot_lines
                    WHERE snapshot_id = :snapshot_id
                    ORDER BY sap_code
                    """
                ),
                {"snapshot_id": int(snapshot["id"])},
            ).mappings().all()
        ]

        historical_rows = session.execute(
            text(
                """
                WITH ranked AS (
                    SELECT
                        id,
                        month_key,
                        ROW_NUMBER() OVER (
                            PARTITION BY month_key
                            ORDER BY
                                CASE WHEN source_kind = 'FINAL' THEN 0 ELSE 1 END,
                                source_plan_date DESC,
                                import_run_id DESC,
                                id DESC
                        ) AS authority_rank
                    FROM mpps_monthly_stock_snapshots
                    WHERE month_key < :month_key
                )
                SELECT
                    r.month_key,
                    l.sap_code,
                    l.total_stock
                FROM ranked r
                JOIN mpps_monthly_stock_snapshot_lines l
                  ON l.snapshot_id = r.id
                WHERE r.authority_rank = 1
                ORDER BY r.month_key, l.sap_code
                """
            ),
            {"month_key": month_key},
        ).mappings().all()

        history: dict[str, list[int]] = defaultdict(list)
        for row in historical_rows:
            history[str(row["sap_code"])].append(cls._int(row["total_stock"]))

        high_risk = 0
        total_stock = 0
        scrap = 0
        blocked = 0
        for row in lines:
            current = cls._int(row.get("total_stock"))
            prediction = cls.predict_stock(history.get(str(row["sap_code"]), []), current)
            row.update(prediction)
            row["model_status"] = "TRAINED" if prediction["samples"] >= 3 else "LEARNING"
            total_stock += current
            scrap += cls._int(row.get("scrap_qty"))
            blocked += cls._int(row.get("blocked_qty"))
            if prediction["risk"] == "HIGH":
                high_risk += 1

        source = dict(snapshot)
        source["status"] = (
            "FINAL"
            if source.get("source_kind") == "FINAL"
            else (
                "LIVE"
                if str(source.get("import_mode") or "").upper() == "LIVE"
                else "HISTORICAL LIVE SNAPSHOT"
            )
        )
        return {
            "month_key": month_key,
            "source": source,
            "rows": lines,
            "summary": {
                "items": len(lines),
                "total_stock": total_stock,
                "scrap": scrap,
                "blocked": blocked,
                "high_risk": high_risk,
            },
            "model_version": cls.MODEL_VERSION,
        }
