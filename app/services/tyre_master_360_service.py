from __future__ import annotations

from collections import defaultdict
from hashlib import sha256
import json
from pathlib import Path
import re
import threading
from typing import Any

from sqlalchemy import text

from app.database import engine, get_session
from app.services.tyre_master_auto_intelligence_service import TyreMasterAutoIntelligenceService

try:
    from app.services.factory_resource_intelligence_service import FactoryResourceIntelligenceService
except Exception:
    FactoryResourceIntelligenceService = None


_LOCK = threading.Lock()
_READY = False


def _clean(value: Any) -> str:
    if value is None:
        return ""
    value = str(value).strip()
    return "" if value.lower() in {"none", "nan", "null", "#n/a"} else value


def _sap(value: Any) -> str:
    value = _clean(value)
    if value.endswith(".0"):
        value = value[:-2]
    digits = re.sub(r"\D", "", value)
    return digits if 6 <= len(digits) <= 12 else ""


def _norm(value: Any) -> str:
    return " ".join(re.sub(r"[^A-Z0-9./+\- ]+", " ", _clean(value).upper()).split())


def _num(value: Any) -> float:
    try:
        return float(value or 0)
    except Exception:
        match = re.search(r"-?\d+(?:\.\d+)?", _clean(value).replace(",", ""))
        return float(match.group(0)) if match else 0.0


def _guess_size(description: str) -> str:
    parts = _clean(description).split()
    if not parts:
        return ""
    if len(parts) >= 2 and re.match(r"^\d+(?:\.\d+)?X\d+(?:\.\d+)?$", parts[0], re.I):
        return f"{parts[0]} {parts[1]}" if re.match(r"^\d+/\d+-\d+", parts[1]) else parts[0]
    return parts[0]


def _hash(path: Path) -> str:
    h = sha256()
    with path.open("rb") as f:
        for block in iter(lambda: f.read(1024 * 1024), b""):
            h.update(block)
    return h.hexdigest()


def _fingerprint(path: Path) -> str:
    s = path.stat()
    return f"{path.resolve()}|{s.st_size}|{s.st_mtime_ns}"


def _payload_rows(value: Any) -> list[dict[str, Any]]:
    if isinstance(value, list):
        return [dict(x) for x in value if isinstance(x, dict)]
    if isinstance(value, dict):
        for key in ("rows", "items", "data", "records", "resources", "result"):
            if key in value:
                rows = _payload_rows(value[key])
                if rows:
                    return rows
        for child in value.values():
            rows = _payload_rows(child)
            if rows:
                return rows
    return []


class TyreMaster360Service:
    MATERIAL_SHEETS = ("CORE", "BAND", "COMPOUND", "TOTAL BEAD", "BEAD", "WGT")

    @classmethod
    def ensure_schema(cls) -> None:
        global _READY
        if _READY:
            return
        with _LOCK:
            if _READY:
                return
            with engine.begin() as conn:
                for sql in (
                    """
                    CREATE TABLE IF NOT EXISTS mpps_tyre_line_compatibility (
                        sap_code TEXT NOT NULL,
                        line TEXT NOT NULL,
                        compatibility VARCHAR(24) NOT NULL DEFAULT 'UNKNOWN',
                        evidence_count INTEGER NOT NULL DEFAULT 0,
                        source VARCHAR(40) NOT NULL DEFAULT 'AUTO',
                        first_seen DATE,
                        last_seen DATE,
                        updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                        PRIMARY KEY (sap_code, line)
                    )
                    """,
                    """
                    CREATE TABLE IF NOT EXISTS mpps_tyre_cavity_compatibility (
                        sap_code TEXT NOT NULL,
                        line TEXT NOT NULL DEFAULT '',
                        cavity TEXT NOT NULL,
                        compatibility VARCHAR(24) NOT NULL DEFAULT 'UNKNOWN',
                        evidence_count INTEGER NOT NULL DEFAULT 0,
                        source VARCHAR(40) NOT NULL DEFAULT 'AUTO',
                        first_seen DATE,
                        last_seen DATE,
                        updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                        PRIMARY KEY (sap_code, line, cavity)
                    )
                    """,
                    """
                    CREATE TABLE IF NOT EXISTS mpps_tyre_material_component (
                        id BIGSERIAL PRIMARY KEY,
                        sap_code TEXT NOT NULL,
                        component_type VARCHAR(40) NOT NULL,
                        component_code TEXT NOT NULL DEFAULT '',
                        component_name TEXT NOT NULL DEFAULT '',
                        quantity NUMERIC(18,5) NOT NULL DEFAULT 0,
                        unit TEXT NOT NULL DEFAULT '',
                        source_sheet TEXT NOT NULL,
                        source_workbook TEXT NOT NULL DEFAULT '',
                        source_hash TEXT NOT NULL DEFAULT '',
                        match_method VARCHAR(40) NOT NULL DEFAULT '',
                        confidence_score NUMERIC(10,4) NOT NULL DEFAULT 0,
                        raw_json JSONB NOT NULL DEFAULT '{}'::jsonb,
                        last_plan_date DATE,
                        updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                        UNIQUE (sap_code, component_type, component_code, component_name, source_sheet)
                    )
                    """,
                    """
                    CREATE TABLE IF NOT EXISTS mpps_tyre_material_observation (
                        id BIGSERIAL PRIMARY KEY,
                        workbook_hash TEXT NOT NULL,
                        workbook_name TEXT NOT NULL,
                        plan_date DATE,
                        sap_code TEXT NOT NULL,
                        component_type VARCHAR(40) NOT NULL,
                        component_code TEXT NOT NULL DEFAULT '',
                        component_name TEXT NOT NULL DEFAULT '',
                        quantity NUMERIC(18,5) NOT NULL DEFAULT 0,
                        unit TEXT NOT NULL DEFAULT '',
                        match_method VARCHAR(40) NOT NULL DEFAULT '',
                        confidence_score NUMERIC(10,4) NOT NULL DEFAULT 0,
                        raw_json JSONB NOT NULL DEFAULT '{}'::jsonb,
                        created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                        UNIQUE (workbook_hash, sap_code, component_type, component_code, component_name)
                    )
                    """,
                    """
                    CREATE TABLE IF NOT EXISTS mpps_tyre_360_sync_state (
                        id INTEGER PRIMARY KEY,
                        source_fingerprint TEXT NOT NULL DEFAULT '',
                        workbook_name TEXT NOT NULL DEFAULT '',
                        material_rows INTEGER NOT NULL DEFAULT 0,
                        status VARCHAR(40) NOT NULL DEFAULT 'NEVER_SYNCED',
                        message TEXT NOT NULL DEFAULT '',
                        last_synced_at TIMESTAMP
                    )
                    """,
                    "CREATE INDEX IF NOT EXISTS ix_tyre_line_compat_sap ON mpps_tyre_line_compatibility (sap_code)",
                    "CREATE INDEX IF NOT EXISTS ix_tyre_cavity_compat_sap ON mpps_tyre_cavity_compatibility (sap_code)",
                    "CREATE INDEX IF NOT EXISTS ix_tyre_material_sap ON mpps_tyre_material_component (sap_code)",
                ):
                    conn.execute(text(sql))
                conn.execute(text("INSERT INTO mpps_tyre_360_sync_state (id) VALUES (1) ON CONFLICT (id) DO NOTHING"))

                # Extend the V33 registry with material intelligence.
                exists = bool(conn.execute(text("""
                    SELECT EXISTS (
                        SELECT 1 FROM information_schema.tables
                        WHERE table_schema='public' AND table_name='mpps_tyre_ml_registry'
                    )
                """)).scalar())
                if exists:
                    for key, name, purpose in (
                        ("MATERIAL_COMPONENT_INTELLIGENCE", "Material & Component Intelligence",
                         "Learn Core / BAND / Compound / Bead relationships per tyre."),
                        ("MATERIAL_REQUIREMENT_FORECAST", "Material Requirement Forecast",
                         "Learn material/component requirement baselines for future production."),
                    ):
                        conn.execute(text("""
                            INSERT INTO mpps_tyre_ml_registry (
                                module_key,module_name,purpose,status,training_rows,
                                history_days,readiness_score,model_version
                            )
                            VALUES (:key,:name,:purpose,'LEARNING',0,0,0,'V34')
                            ON CONFLICT (module_key) DO UPDATE
                            SET module_name=EXCLUDED.module_name,
                                purpose=EXCLUDED.purpose,
                                model_version='V34',
                                updated_at=CURRENT_TIMESTAMP
                        """), {"key": key, "name": name, "purpose": purpose})
            _READY = True

    @classmethod
    def _factory_resources(cls) -> tuple[list[str], dict[str, list[str]]]:
        lines: set[str] = set()
        cavities: dict[str, set[str]] = defaultdict(set)

        if FactoryResourceIntelligenceService is not None:
            try:
                with get_session() as session:
                    svc = FactoryResourceIntelligenceService()
                    for tab in ("lines", "production_lines"):
                        try:
                            for row in _payload_rows(svc.tab_snapshot(session, tab)):
                                line = _clean(row.get("line") or row.get("production_line") or row.get("line_name") or row.get("name"))
                                if line:
                                    lines.add(line)
                            if lines:
                                break
                        except Exception:
                            pass
                    for tab in ("cavities", "cavity"):
                        try:
                            rows = _payload_rows(svc.tab_snapshot(session, tab))
                            if not rows:
                                continue
                            for row in rows:
                                line = _clean(row.get("line") or row.get("production_line") or row.get("line_name"))
                                cavity = _clean(row.get("cavity") or row.get("oven_no") or row.get("oven_no_cavity") or row.get("oven") or row.get("name"))
                                if line:
                                    lines.add(line)
                                if cavity:
                                    cavities[line].add(cavity)
                            break
                        except Exception:
                            pass
            except Exception:
                pass

        with engine.connect() as conn:
            for table in ("mpps_tyre_factory_mapping", "mpps_tyre_workbook_observation"):
                try:
                    rows = conn.execute(text(f"""
                        SELECT DISTINCT line, oven_no
                        FROM {table}
                        WHERE COALESCE(line,'')<>'' OR COALESCE(oven_no,'')<>''
                    """)).mappings().all()
                    for row in rows:
                        line = _clean(row.get("line"))
                        cavity = _clean(row.get("oven_no"))
                        if line:
                            lines.add(line)
                        if cavity:
                            cavities[line].add(cavity)
                except Exception:
                    pass

        return sorted(lines), {k: sorted(v) for k, v in cavities.items()}

    @classmethod
    def rebuild_compatibility(cls) -> dict[str, int]:
        cls.ensure_schema()
        lines, cavities = cls._factory_resources()

        with engine.begin() as conn:
            saps = [_clean(v) for v in conn.execute(text(
                "SELECT sap_code FROM smds WHERE COALESCE(sap_code,'')<>''"
            )).scalars().all() if _clean(v)]

            # UNKNOWN means no proof yet; unseen never becomes a false X.
            for sap in saps:
                for line in lines:
                    conn.execute(text("""
                        INSERT INTO mpps_tyre_line_compatibility
                            (sap_code,line,compatibility,evidence_count,source)
                        VALUES (:sap,:line,'UNKNOWN',0,'FACTORY_RESOURCE')
                        ON CONFLICT (sap_code,line) DO NOTHING
                    """), {"sap": sap, "line": line})

            observed_lines = conn.execute(text("""
                SELECT sap_code,line,COUNT(*) evidence_count,
                       MIN(plan_date) first_seen,MAX(plan_date) last_seen
                FROM mpps_tyre_workbook_observation
                WHERE COALESCE(sap_code,'')<>'' AND COALESCE(line,'')<>''
                GROUP BY sap_code,line
            """)).mappings().all()

            for row in observed_lines:
                conn.execute(text("""
                    INSERT INTO mpps_tyre_line_compatibility
                        (sap_code,line,compatibility,evidence_count,source,first_seen,last_seen)
                    VALUES (:sap_code,:line,'CONFIRMED',:evidence_count,'EXCEL_HISTORY',:first_seen,:last_seen)
                    ON CONFLICT (sap_code,line) DO UPDATE
                    SET compatibility=CASE
                            WHEN mpps_tyre_line_compatibility.source='MANUAL'
                             AND mpps_tyre_line_compatibility.compatibility='INCOMPATIBLE'
                            THEN 'INCOMPATIBLE' ELSE 'CONFIRMED' END,
                        evidence_count=EXCLUDED.evidence_count,
                        last_seen=EXCLUDED.last_seen,
                        first_seen=COALESCE(mpps_tyre_line_compatibility.first_seen,EXCLUDED.first_seen),
                        source=CASE WHEN mpps_tyre_line_compatibility.source='MANUAL'
                                    THEN 'MANUAL' ELSE 'EXCEL_HISTORY' END,
                        updated_at=CURRENT_TIMESTAMP
                """), dict(row))

            observed_cavities = conn.execute(text("""
                SELECT sap_code,line,oven_no cavity,COUNT(*) evidence_count,
                       MIN(plan_date) first_seen,MAX(plan_date) last_seen
                FROM mpps_tyre_workbook_observation
                WHERE COALESCE(sap_code,'')<>'' AND COALESCE(oven_no,'')<>''
                GROUP BY sap_code,line,oven_no
            """)).mappings().all()

            for row in observed_cavities:
                conn.execute(text("""
                    INSERT INTO mpps_tyre_cavity_compatibility
                        (sap_code,line,cavity,compatibility,evidence_count,source,first_seen,last_seen)
                    VALUES (:sap_code,:line,:cavity,'CONFIRMED',:evidence_count,'EXCEL_HISTORY',:first_seen,:last_seen)
                    ON CONFLICT (sap_code,line,cavity) DO UPDATE
                    SET compatibility=CASE
                            WHEN mpps_tyre_cavity_compatibility.source='MANUAL'
                             AND mpps_tyre_cavity_compatibility.compatibility='INCOMPATIBLE'
                            THEN 'INCOMPATIBLE' ELSE 'CONFIRMED' END,
                        evidence_count=EXCLUDED.evidence_count,
                        last_seen=EXCLUDED.last_seen,
                        first_seen=COALESCE(mpps_tyre_cavity_compatibility.first_seen,EXCLUDED.first_seen),
                        source=CASE WHEN mpps_tyre_cavity_compatibility.source='MANUAL'
                                    THEN 'MANUAL' ELSE 'EXCEL_HISTORY' END,
                        updated_at=CURRENT_TIMESTAMP
                """), dict(row))

            confirmed: dict[str, list[str]] = defaultdict(list)
            for row in observed_lines:
                confirmed[_clean(row.get("line"))].append(_clean(row.get("sap_code")))
            for line, cavity_list in cavities.items():
                for sap in confirmed.get(line, []):
                    for cavity in cavity_list:
                        conn.execute(text("""
                            INSERT INTO mpps_tyre_cavity_compatibility
                                (sap_code,line,cavity,compatibility,evidence_count,source)
                            VALUES (:sap,:line,:cavity,'UNKNOWN',0,'FACTORY_RESOURCE')
                            ON CONFLICT (sap_code,line,cavity) DO NOTHING
                        """), {"sap": sap, "line": line, "cavity": cavity})

        return {"line_links": len(observed_lines), "cavity_links": len(observed_cavities)}

    @classmethod
    def _master_maps(cls) -> tuple[set[str], dict[str, str]]:
        sap_codes: set[str] = set()
        desc_map: dict[str, str] = {}
        duplicate_desc: set[str] = set()
        with engine.connect() as conn:
            rows = conn.execute(text("""
                SELECT sap_code,material_description
                FROM smds WHERE COALESCE(sap_code,'')<>''
            """)).mappings().all()
        for row in rows:
            sap = _clean(row.get("sap_code"))
            if not sap:
                continue
            sap_codes.add(sap)
            desc = _norm(row.get("material_description"))
            if desc:
                if desc in desc_map:
                    duplicate_desc.add(desc)
                else:
                    desc_map[desc] = sap
        for desc in duplicate_desc:
            desc_map.pop(desc, None)
        return sap_codes, desc_map

    @classmethod
    def _header_row(cls, ws) -> int:
        terms = ("SAP", "MATERIAL", "DESCRIPTION", "CORE", "BAND", "COMPOUND",
                 "BEAD", "QTY", "QUANTITY", "WEIGHT", "KG", "UNIT", "CODE")
        best = (1, -1)
        for r in range(1, min(ws.max_row, 35) + 1):
            values = [_norm(ws.cell(r, c).value) for c in range(1, min(ws.max_column, 70) + 1)]
            score = sum(4 for v in values if v and any(t in v for t in terms))
            score += sum(1 for v in values if v and not re.fullmatch(r"\d+(?:\.\d+)?", v))
            if score > best[1]:
                best = (r, score)
        return best[0]

    @classmethod
    def _material_sheet_rows(cls, ws, component_type, sap_codes, desc_map) -> list[dict[str, Any]]:
        header_row = cls._header_row(ws)
        max_col = min(ws.max_column, 70)
        headers = [_clean(ws.cell(header_row, c).value) or f"COL_{c}" for c in range(1, max_col + 1)]
        output = []

        for r in range(header_row + 1, ws.max_row + 1):
            values = [ws.cell(r, c).value for c in range(1, max_col + 1)]
            if not any(_clean(v) for v in values):
                continue

            sap = ""
            method = ""
            confidence = 0.0

            for value in values:
                candidate = _sap(value)
                if candidate in sap_codes:
                    sap, method, confidence = candidate, "SAP_DIRECT", 1.0
                    break

            if not sap:
                for value in values:
                    key = _norm(value)
                    if key and key in desc_map:
                        sap, method, confidence = desc_map[key], "DESCRIPTION_EXACT_UNIQUE", 0.96
                        break

            if not sap:
                continue

            raw = {headers[i]: values[i] for i in range(len(headers)) if _clean(values[i])}
            code = ""
            name = ""
            qty = 0.0
            unit = ""

            for header, value in raw.items():
                h = _norm(header)
                value_text = _clean(value)
                if not code and any(t in h for t in ("CODE", "MATERIAL", "CORE", "BAND", "COMPOUND", "BEAD")):
                    if _sap(value) != sap:
                        code = value_text
                if not name and any(t in h for t in ("DESCRIPTION", "NAME", "MATERIAL")):
                    if _norm(value_text) not in desc_map:
                        name = value_text
                if qty == 0 and any(t in h for t in ("QTY", "QUANTITY", "WEIGHT", "KG", "USAGE", "REQUIREMENT", "CONSUMPTION")):
                    qty = _num(value)
                if not unit and "UNIT" in h:
                    unit = value_text

            output.append({
                "sap_code": sap,
                "component_type": component_type,
                "component_code": (code or component_type)[:250],
                "component_name": (name or component_type)[:500],
                "quantity": qty,
                "unit": (unit or ("AUTO" if qty else ""))[:80],
                "match_method": method,
                "confidence_score": confidence,
                "raw_json": raw,
                "source_sheet": ws.title,
            })
        return output

    @classmethod
    def sync_latest(cls, force: bool = False) -> dict[str, Any]:
        cls.ensure_schema()
        location = TyreMasterAutoIntelligenceService.locate_latest_workbook()
        if location is None:
            return {"changed": False, "status": "NO_WORKBOOK"}

        path = Path(location.path)
        fingerprint = _fingerprint(path)

        with engine.connect() as conn:
            previous = _clean(conn.execute(text(
                "SELECT source_fingerprint FROM mpps_tyre_360_sync_state WHERE id=1"
            )).scalar())
        if not force and previous == fingerprint:
            return {"changed": False, "status": "UP_TO_DATE"}

        from openpyxl import load_workbook

        workbook_hash = _hash(path)
        sap_codes, desc_map = cls._master_maps()
        material_rows = []

        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet_name in wb.sheetnames:
                normalized = _norm(sheet_name)
                component_type = next((x for x in cls.MATERIAL_SHEETS if x in normalized), "")
                if not component_type:
                    continue
                material_rows.extend(
                    cls._material_sheet_rows(wb[sheet_name], component_type, sap_codes, desc_map)
                )
        finally:
            wb.close()

        plan_date = getattr(location, "plan_date", None)

        with engine.begin() as conn:
            for row in material_rows:
                params = {
                    **row,
                    "workbook": location.workbook_name,
                    "workbook_hash": workbook_hash,
                    "plan_date": plan_date,
                    "raw_json_text": json.dumps(row.get("raw_json") or {}, default=str),
                }
                conn.execute(text("""
                    INSERT INTO mpps_tyre_material_component (
                        sap_code,component_type,component_code,component_name,
                        quantity,unit,source_sheet,source_workbook,source_hash,
                        match_method,confidence_score,raw_json,last_plan_date
                    )
                    VALUES (
                        :sap_code,:component_type,:component_code,:component_name,
                        :quantity,:unit,:source_sheet,:workbook,:workbook_hash,
                        :match_method,:confidence_score,CAST(:raw_json_text AS JSONB),:plan_date
                    )
                    ON CONFLICT (sap_code,component_type,component_code,component_name,source_sheet)
                    DO UPDATE SET
                        quantity=EXCLUDED.quantity,unit=EXCLUDED.unit,
                        source_workbook=EXCLUDED.source_workbook,source_hash=EXCLUDED.source_hash,
                        match_method=EXCLUDED.match_method,confidence_score=EXCLUDED.confidence_score,
                        raw_json=EXCLUDED.raw_json,last_plan_date=EXCLUDED.last_plan_date,
                        updated_at=CURRENT_TIMESTAMP
                """), params)
                conn.execute(text("""
                    INSERT INTO mpps_tyre_material_observation (
                        workbook_hash,workbook_name,plan_date,sap_code,
                        component_type,component_code,component_name,quantity,
                        unit,match_method,confidence_score,raw_json
                    )
                    VALUES (
                        :workbook_hash,:workbook,:plan_date,:sap_code,
                        :component_type,:component_code,:component_name,:quantity,
                        :unit,:match_method,:confidence_score,CAST(:raw_json_text AS JSONB)
                    )
                    ON CONFLICT (workbook_hash,sap_code,component_type,component_code,component_name)
                    DO NOTHING
                """), params)

        compatibility = cls.rebuild_compatibility()
        cls.refresh_material_ml()

        with engine.begin() as conn:
            conn.execute(text("""
                UPDATE mpps_tyre_360_sync_state
                SET source_fingerprint=:fingerprint,workbook_name=:workbook,
                    material_rows=:material_rows,status='SYNCED',
                    message=:message,last_synced_at=CURRENT_TIMESTAMP
                WHERE id=1
            """), {
                "fingerprint": fingerprint,
                "workbook": location.workbook_name,
                "material_rows": len(material_rows),
                "message": f"V34 Tyre 360 enriched from {location.workbook_name}",
            })

        return {
            "changed": True,
            "status": "SYNCED",
            "workbook_name": location.workbook_name,
            "material_rows": len(material_rows),
            **compatibility,
        }

    @classmethod
    def refresh_material_ml(cls) -> None:
        cls.ensure_schema()
        with engine.begin() as conn:
            exists = bool(conn.execute(text("""
                SELECT EXISTS (
                    SELECT 1 FROM information_schema.tables
                    WHERE table_schema='public' AND table_name='mpps_tyre_ml_registry'
                )
            """)).scalar())
            if not exists:
                return

            rows = int(conn.execute(text("SELECT COUNT(*) FROM mpps_tyre_material_observation")).scalar() or 0)
            days = int(conn.execute(text("""
                SELECT COUNT(DISTINCT plan_date)
                FROM mpps_tyre_material_observation WHERE plan_date IS NOT NULL
            """)).scalar() or 0)

            for key, minimum in (
                ("MATERIAL_COMPONENT_INTELLIGENCE", 50),
                ("MATERIAL_REQUIREMENT_FORECAST", 250),
            ):
                ready = rows >= minimum and days >= 5
                score = min(100.0, min(1.0, rows / minimum) * min(1.0, days / 5) * 100)
                conn.execute(text("""
                    UPDATE mpps_tyre_ml_registry
                    SET status=:status,training_rows=:rows,history_days=:days,
                        readiness_score=:score,model_version='V34',
                        last_trained_at=CASE WHEN :status='TRAINED'
                                             THEN CURRENT_TIMESTAMP ELSE last_trained_at END,
                        updated_at=CURRENT_TIMESTAMP
                    WHERE module_key=:key
                """), {
                    "key": key, "status": "TRAINED" if ready else "LEARNING",
                    "rows": rows, "days": days, "score": score,
                })

    @classmethod
    def _symbol(cls, state: str) -> str:
        return "✓" if state == "CONFIRMED" else ("✕" if state == "INCOMPATIBLE" else "?")

    @classmethod
    def master_page(cls, search="", page=1, page_size=150) -> dict[str, Any]:
        cls.ensure_schema()
        page = max(1, int(page))
        page_size = max(50, min(300, int(page_size)))
        offset = (page - 1) * page_size
        params = {"limit": page_size, "offset": offset}
        search_sql = ""
        if _clean(search):
            search_sql = "AND (s.sap_code ILIKE :search OR COALESCE(s.material_description,'') ILIKE :search)"
            params["search"] = f"%{_clean(search)}%"

        with engine.connect() as conn:
            total = int(conn.execute(text(f"""
                SELECT COUNT(*) FROM smds s WHERE 1=1 {search_sql}
            """), params).scalar() or 0)
            rows = conn.execute(text(f"""
                SELECT
                    s.sap_code,
                    COALESCE(s.material_description,'') description,
                    (SELECT COUNT(*) FROM mpps_tyre_line_compatibility lc
                     WHERE lc.sap_code=s.sap_code AND lc.compatibility='CONFIRMED') confirmed_lines,
                    (SELECT COUNT(*) FROM mpps_tyre_line_compatibility lc
                     WHERE lc.sap_code=s.sap_code) total_lines,
                    (SELECT COUNT(*) FROM mpps_tyre_cavity_compatibility cc
                     WHERE cc.sap_code=s.sap_code AND cc.compatibility='CONFIRMED') confirmed_cavities,
                    (SELECT COUNT(*) FROM mpps_tyre_material_component mc
                     WHERE mc.sap_code=s.sap_code) material_rows,
                    (SELECT MAX(last_seen) FROM mpps_tyre_line_compatibility lc
                     WHERE lc.sap_code=s.sap_code) last_seen
                FROM smds s
                WHERE 1=1 {search_sql}
                ORDER BY s.sap_code
                LIMIT :limit OFFSET :offset
            """), params).mappings().all()

        output = []
        for row in rows:
            item = dict(row)
            item["tyre_size"] = _guess_size(item.get("description"))
            item["line_status"] = f"{int(item.get('confirmed_lines') or 0)} ✓ / {int(item.get('total_lines') or 0)}"
            item["cavity_status"] = f"{int(item.get('confirmed_cavities') or 0)} ✓"
            item["material_status"] = f"{int(item.get('material_rows') or 0)} linked"
            output.append(item)

        from math import ceil
        return {
            "rows": output, "total": total, "page": page,
            "page_size": page_size, "page_count": max(1, ceil(total / page_size)),
        }

    @classmethod
    def quality_page(cls, search="", page=1, page_size=150) -> dict[str, Any]:
        result = cls.master_page(search, page, page_size)
        rows = []
        for item in result["rows"]:
            issues = []
            if int(item.get("confirmed_lines") or 0) <= 0:
                issues.append("No confirmed line")
            if int(item.get("confirmed_cavities") or 0) <= 0:
                issues.append("No confirmed cavity")
            if int(item.get("material_rows") or 0) <= 0:
                issues.append("No material/component")
            row = dict(item)
            row["issues"] = ", ".join(issues) if issues else "Healthy"
            row["issue_count"] = len(issues)
            rows.append(row)
        result["rows"] = rows
        return result

    @classmethod
    def tyre_360(cls, sap_code: str) -> dict[str, Any]:
        cls.ensure_schema()
        sap_code = _clean(sap_code)

        with engine.connect() as conn:
            cols = {str(x) for x in conn.execute(text("""
                SELECT column_name FROM information_schema.columns
                WHERE table_schema='public' AND table_name='smds'
            """)).scalars().all()}

            def ex(candidates, alias, default="NULL"):
                col = next((c for c in candidates if c in cols), None)
                return f"{col} AS {alias}" if col else f"{default} AS {alias}"

            master = conn.execute(text(f"""
                SELECT
                    {ex(('sap_code',),'sap_code',"''")},
                    {ex(('material_description','description'),'description',"''")},
                    {ex(('line','production_line'),'line',"''")},
                    {ex(('key_code','mold_key_code'),'key_code',"''")},
                    {ex(('casing_type','casing_code'),'casing_type',"''")},
                    {ex(('curing_cycle',),'curing_cycle',"''")},
                    {ex(('normal_curing_minutes',),'curing_minutes','0')},
                    {ex(('heel',),'heel',"''")},
                    {ex(('soft',),'soft',"''")},
                    {ex(('tred','tread'),'tread',"''")},
                    {ex(('weight_per_tyre_kg','weight_kg','weight'),'weight_kg','0')},
                    {ex(('day_plan',),'day_plan','0')},
                    {ex(('night_plan',),'night_plan','0')},
                    {ex(('source_file',),'source_file',"''")}
                FROM smds WHERE sap_code=:sap LIMIT 1
            """), {"sap": sap_code}).mappings().first()
            master = dict(master or {})
            master["tyre_size"] = _guess_size(master.get("description"))

            line_rows = [dict(r) for r in conn.execute(text("""
                SELECT line,compatibility,evidence_count,source,first_seen,last_seen
                FROM mpps_tyre_line_compatibility
                WHERE sap_code=:sap ORDER BY line
            """), {"sap": sap_code}).mappings().all()]

            cavity_rows = [dict(r) for r in conn.execute(text("""
                SELECT line,cavity,compatibility,evidence_count,source,first_seen,last_seen
                FROM mpps_tyre_cavity_compatibility
                WHERE sap_code=:sap ORDER BY line,cavity
            """), {"sap": sap_code}).mappings().all()]

            materials = [dict(r) for r in conn.execute(text("""
                SELECT component_type,component_code,component_name,quantity,unit,
                       source_sheet,source_workbook,match_method,confidence_score,last_plan_date
                FROM mpps_tyre_material_component
                WHERE sap_code=:sap
                ORDER BY component_type,component_name,component_code
            """), {"sap": sap_code}).mappings().all()]

            latest = conn.execute(text("""
                SELECT line,oven_no,heel,soft,tread,weight_kg,day_plan,night_plan,
                       day_produced,night_produced,total_stock,current_stock,scrap,blocked,
                       plan_date,workbook_name
                FROM mpps_tyre_workbook_observation
                WHERE sap_code=:sap
                ORDER BY plan_date DESC NULLS LAST,id DESC
                LIMIT 1
            """), {"sap": sap_code}).mappings().first()

            observations = int(conn.execute(text("""
                SELECT COUNT(*) FROM mpps_tyre_workbook_observation WHERE sap_code=:sap
            """), {"sap": sap_code}).scalar() or 0)

            features = []
            try:
                features = [dict(r) for r in conn.execute(text("""
                    SELECT module_key,entity_key,feature_json,sample_count,updated_at
                    FROM mpps_tyre_ml_features
                    WHERE entity_key=:sap OR entity_key LIKE :prefix
                    ORDER BY module_key,sample_count DESC LIMIT 30
                """), {"sap": sap_code, "prefix": f"{sap_code}|%"}).mappings().all()]
            except Exception:
                pass

        latest = dict(latest or {})
        for key in ("line", "heel", "soft", "tread", "weight_kg", "day_plan", "night_plan"):
            if not _clean(master.get(key)) or (key in {"weight_kg","day_plan","night_plan"} and _num(master.get(key)) == 0):
                master[key] = latest.get(key)

        line_matrix = {
            r["line"]: {**r, "symbol": cls._symbol(_clean(r.get("compatibility")).upper())}
            for r in line_rows
        }
        cavity_matrix: dict[str, dict[str, dict[str, Any]]] = defaultdict(dict)
        for r in cavity_rows:
            cavity_matrix[_clean(r.get("line"))][_clean(r.get("cavity"))] = {
                **r, "symbol": cls._symbol(_clean(r.get("compatibility")).upper())
            }

        return {
            "sap_code": sap_code,
            "master": master,
            "latest": latest,
            "line_matrix": line_matrix,
            "cavity_matrix": dict(cavity_matrix),
            "materials": materials,
            "observation_count": observations,
            "ml_features": features,
        }
