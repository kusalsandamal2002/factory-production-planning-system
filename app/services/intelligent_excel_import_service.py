from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import asdict, dataclass, field
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher
from hashlib import sha256
import json
import math
import os
from pathlib import Path
import re
import shutil
import zipfile
import xml.etree.ElementTree as ET
from typing import Any, Callable, Iterable

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from sqlalchemy import text

from app.services.production_learning_service import (
    ProductionLearningService,
)
from app.services.ai_planning_service import AIPlanningService
from app.services.factory_intelligence_service import FactoryIntelligenceService
from app.services.factory_resource_intelligence_service import FactoryResourceIntelligenceService
from app.services.monthly_stock_snapshot_service import MonthlyStockSnapshotService
from app.services.workbook_continuous_sync_service import (
    WorkbookContinuousSyncService,
)


# INTELLIGENT CONTINUOUS EXCEL SYNC + LEARNING FOUNDATION V7.0
# DELIVERY DATE INTEGRITY V6.3: imported reviews never receive promise dates
ProgressCallback = Callable[[int, str], None]


def _fast_xlsx_sheet_stats(path: Path) -> dict[str, dict[str, int]]:
    """Count non-empty/formula/error cells directly from worksheet XML.

    Large OVEN workbooks contain millions of cells. Iterating every cell twice
    through openpyxl (data + formula workbooks) makes semantic profiling dominate
    import time. The XML stream already tells us whether a cell exists, whether it
    contains a formula and whether Excel cached an error. This path preserves exact
    diagnostics without materializing Python Cell objects.
    """
    stats: dict[str, dict[str, int]] = {}
    try:
        with zipfile.ZipFile(path) as zf:
            wb_root = ET.fromstring(zf.read('xl/workbook.xml'))
            rel_root = ET.fromstring(zf.read('xl/_rels/workbook.xml.rels'))
            rels = {}
            for rel in rel_root:
                rid = rel.attrib.get('Id')
                target = rel.attrib.get('Target', '')
                if rid and target:
                    target = target.lstrip('/')
                    if not target.startswith('xl/'):
                        target = 'xl/' + target.replace('\\', '/')
                    rels[rid] = target
            ns_rel = '{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id'
            sheets = []
            for node in wb_root.iter():
                if node.tag.endswith('}sheet'):
                    name = node.attrib.get('name', '')
                    rid = node.attrib.get(ns_rel, '')
                    target = rels.get(rid)
                    if name and target:
                        sheets.append((name, target))
            for name, target in sheets:
                nonempty = formulas = errors = 0
                try:
                    with zf.open(target) as stream:
                        for _event, elem in ET.iterparse(stream, events=('end',)):
                            if elem.tag.endswith('}c'):
                                has_formula = False
                                has_value = False
                                for child in list(elem):
                                    if child.tag.endswith('}f'):
                                        has_formula = True
                                    elif child.tag.endswith('}v') or child.tag.endswith('}is'):
                                        has_value = True
                                if has_formula:
                                    formulas += 1
                                if has_formula or has_value:
                                    nonempty += 1
                                if elem.attrib.get('t') == 'e':
                                    errors += 1
                                elem.clear()
                    stats[name] = {
                        'nonempty_cells': nonempty,
                        'formula_cells': formulas,
                        'cached_error_cells': errors,
                    }
                except KeyError:
                    continue
    except Exception:
        return {}
    return stats


ROLE_DEFINITIONS: dict[str, dict[str, Any]] = {
    "PRODUCTION_STOCK_SHIPMENTS": {
        "names": ["prod", "production", "stock shipment", "shipment demand"],
        "headers": [
            "sap code",
            "material description",
            "stock",
            "total shipment",
            "production required",
        ],
    },
    "DAILY_PRODUCTION_PLAN": {
        "names": ["daily plan", "daily production", "production plan"],
        "headers": ["tyre code", "daily plan", "production", "tomorrow plan"],
    },
    "OVEN_CAVITY_PLAN": {
        "names": ["oven", "oven plan", "cavity plan", "press plan"],
        "headers": [
            "line",
            "oven no",
            "tyre code",
            "day plan pcs",
            "night plan pcs",
            "next day plan",
        ],
    },
    "BEAD_REQUIREMENT": {
        "names": ["total bead", "bead", "bead requirement"],
        "headers": ["tyre size", "bead per tyre", "total bead requirement", "bead type"],
    },
    "COMPOUND_BOM": {
        "names": ["compound", "compound bom", "compound requirement"],
        "headers": ["compound", "sap", "material", "weight"],
    },
    "BAND_PLAN": {
        "names": ["band", "band plan", "band requirement"],
        "headers": ["material description", "next day plan"],
    },
    "CORE_PLAN": {
        "names": ["core", "core plan", "inner core"],
        "headers": ["core", "total plan", "day shift", "night shift", "closing stock"],
    },
    "WEIGHT_MASTER": {
        "names": ["wgt", "weight", "production weight"],
        "headers": ["material", "weight", "day", "night"],
    },
    "DAY_SHIFT_PLAN": {
        "names": ["day", "day shift"],
        "headers": ["day", "plan", "tyre"],
    },
    "NIGHT_SHIFT_PLAN": {
        "names": ["night", "night shift"],
        "headers": ["night", "plan", "tyre"],
    },
    "HOURLY_DAY_PLAN": {
        "names": ["hourly plan day", "hourly day"],
        "headers": ["hourly", "day", "plan"],
    },
    "HOURLY_NIGHT_PLAN": {
        "names": ["hourly plan night", "hourly night"],
        "headers": ["hourly", "night", "plan"],
    },
}


LIVE_TABLE_COLUMNS: dict[str, set[str]] = {
    "mpps_sap_stock_items": {
        "sap_code",
        "tyre_description",
        "item_description",
        "fg_stock",
        "qc_stock",
        "scrap_stock",
        "blocked_stock",
        "source_table",
        "source_note",
        "is_active",
        "updated_at",
    },
    "mpps_stock_items": {
        "material_code",
        "item_description",
        "fg_stock",
        "qc_stock",
        "scrap_stock",
        "blocked_stock",
        "average_weight",
        "source_workbook",
        "source_sheet",
        "source_row",
        "source_note",
        "is_active",
        "last_updated_date",
        "updated_at",
    },
    "mpps_daily_stock_entries": {
        "stock_date",
        "sap_code",
        "tyre_description",
        "production_qty",
        "fg_qty",
        "qc_qty",
        "scrap_qty",
        "blocked_qty",
        "note",
        "source_file",
        "updated_at",
    },
    "smds": {
        "sap_code",
        "material_description",
        "weight_per_tyre_kg",
        "source_file",
        "source_sheet",
        "source_row_number",
        "updated_at",
    },
    "mpps_oven_plan": {
        "plan_date",
        "oven_code",
        "shift_name",
        "material_code",
        "item_description",
        "planned_qty",
        "planned_weight_kg",
        "plan_status",
        "source_workbook",
        "source_sheet",
        "source_row",
        "source_note",
        "line_name",
        "cavity_code",
        "allocation_slot",
        "mold_code",
        "updated_at",
    },
    "mpps_compound_master": {
        "item_code",
        "compound_code",
        "compound_name",
        "compound_weight_per_unit",
        "stage",
        "is_active",
        "source_workbook",
        "source_sheet",
        "source_row",
        "source_note",
        "updated_at",
    },
    "mpps_bom_items": {
        "finished_item_code",
        "raw_material_code",
        "raw_material_name",
        "usage_per_unit",
        "unit",
        "wastage_percentage",
        "is_active",
        "source_workbook",
        "source_sheet",
        "source_row",
        "source_note",
        "updated_at",
    },
    "mpps_bead_master": {
        "item_code",
        "bead_type",
        "bead_per_tyre",
        "is_active",
        "source_workbook",
        "source_sheet",
        "source_row",
        "source_note",
        "updated_at",
    },
    "mpps_shipments": {
        "shipment_no",
        "shipment_id",
        "shipment_name",
        "customer_name",
        "shipment_date",
        "status",
        "note",
        "manager_order_date",
        "target_date",
        "plan_date",
        "factory_can_receive_date",
        "factory_out_date",
        "dispatch_buffer_days",
        "delivery_status",
        "delay_days",
        "early_days",
        "planning_status",
        "planning_note",
        "target_date_is_manual",
        "target_date_source",
        "last_replanned_at",
        "source_family",
        "source_identity_key",
        "source_latest_run_id",
        "source_latest_plan_date",
        "source_latest_workbook",
        "source_latest_column",
        "source_latest_status",
        "source_missing_from_latest",
        "source_revision_no",
        "source_sync_status",
        "source_sync_note",
        "updated_at",
    },
    "mpps_shipment_items": {
        "shipment_id",
        "sap_code",
        "item_description",
        "quantity",
        "item_status",
        "note",
        "stock_allocated_qty",
        "production_required_qty",
        "remaining_qty",
        "planning_note",
        "schedule_reason",
        "factory_out_reason",
        "source_item_key",
        "source_latest_run_id",
        "source_latest_plan_date",
        "source_latest_qty",
        "source_removed_from_latest",
        "source_revision_no",
        "source_sync_status",
        "source_sync_note",
        "source_manual_lock",
        "updated_at",
    },
    "excel_shipment_identities": {
        "source_family",
        "identity_key",
        "base_key",
        "display_name",
        "canonical_shipment_id",
        "first_seen_plan_date",
        "last_seen_plan_date",
        "latest_run_id",
        "latest_workbook_hash",
        "latest_workbook_name",
        "latest_column",
        "latest_status",
        "latest_item_fingerprint",
        "latest_total_qty",
        "latest_item_count",
        "is_active",
        "missing_since_plan_date",
        "updated_at",
    },
}


@dataclass
class ImportIssue:
    severity: str
    category: str
    sheet_name: str
    cell_address: str
    item_key: str
    message: str
    recommendation: str = ""


@dataclass
class SheetProfile:
    sheet_name: str
    role: str
    confidence: float
    max_row: int
    max_column: int
    nonempty_cells: int
    formula_cells: int
    cached_error_cells: int
    header_row: int | None
    evidence: str


@dataclass
class WorkbookAnalysis:
    workbook_path: str
    workbook_name: str
    workbook_hash: str
    workbook_size_bytes: int
    plan_date: str | None
    confidence_score: float
    detected_type: str
    sheet_profiles: list[SheetProfile] = field(default_factory=list)
    issues: list[ImportIssue] = field(default_factory=list)
    stock_rows: list[dict[str, Any]] = field(default_factory=list)
    shipment_rows: list[dict[str, Any]] = field(default_factory=list)
    oven_rows: list[dict[str, Any]] = field(default_factory=list)
    oven_resource_rows: list[dict[str, Any]] = field(default_factory=list)
    compound_rows: list[dict[str, Any]] = field(default_factory=list)
    bead_rows: list[dict[str, Any]] = field(default_factory=list)
    band_rows: list[dict[str, Any]] = field(default_factory=list)
    material_plan_rows: list[dict[str, Any]] = field(default_factory=list)
    production_history_rows: list[dict[str, Any]] = field(default_factory=list)
    actual_production_dates: list[dict[str, Any]] = field(default_factory=list)
    summary: dict[str, Any] = field(default_factory=dict)

    def to_dict(self, *, include_rows: bool = True) -> dict[str, Any]:
        payload = asdict(self)
        if not include_rows:
            for key in (
                "stock_rows",
                "shipment_rows",
                "oven_rows",
                "oven_resource_rows",
                "compound_rows",
                "bead_rows",
                "band_rows",
                "material_plan_rows",
                "production_history_rows",
                "actual_production_dates",
            ):
                payload.pop(key, None)
        return payload


class IntelligentExcelImportService:
    # R7.4.1 bulk-history caches: the training process stays on one verified DB
    # cluster, so additive DDL and serial-sequence repair only need one preflight.
    _r741_schema_ready = False
    _r741_sequences_repaired = False

    """Semantic, confidence-scored importer for MPPS/OVEN workbooks.

    The engine deliberately separates *source preservation* from *live master
    updates*.  Every workbook is archived byte-for-byte.  High-confidence stock,
    production, weight, oven, shipment snapshot and material records are mapped
    transactionally.  Ambiguous casing, mold, key-code and line-compatibility
    values are retained as evidence/issues instead of overwriting approved master
    data.
    """

    def __init__(self, project_root: str | Path | None = None):
        self.project_root = (
            Path(project_root).resolve()
            if project_root
            else Path(__file__).resolve().parents[2]
        )

    def analyze(
        self,
        workbook_path: str | Path,
        *,
        progress: ProgressCallback | None = None,
    ) -> WorkbookAnalysis:
        path = Path(workbook_path).expanduser().resolve()
        if not path.exists():
            raise FileNotFoundError(path)
        if path.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise ValueError("Only .xlsx and .xlsm workbooks are supported.")

        self._progress(progress, 2, "Calculating workbook checksum")
        digest = _sha256_file(path)

        self._progress(progress, 4, "Indexing workbook XML for fast formula/error statistics")
        xml_sheet_stats = _fast_xlsx_sheet_stats(path)

        self._progress(progress, 6, "Opening workbook in streaming mode")
        formula_wb = load_workbook(
            path,
            read_only=True,
            data_only=False,
            keep_links=False,
        )
        value_wb = load_workbook(
            path,
            read_only=True,
            data_only=True,
            keep_links=False,
        )

        try:
            roles: dict[str, str] = {}
            profiles: list[SheetProfile] = []
            issues: list[ImportIssue] = []

            total_sheets = max(1, len(value_wb.sheetnames))
            for index, sheet_name in enumerate(value_wb.sheetnames):
                pct = 8 + int((index / total_sheets) * 28)
                self._progress(
                    progress,
                    pct,
                    f"Profiling sheet {index + 1}/{total_sheets}: {sheet_name}",
                )
                vws = value_wb[sheet_name]
                fws = formula_wb[sheet_name]
                profile = self._profile_sheet(vws, fws, xml_sheet_stats.get(sheet_name))
                profiles.append(profile)
                roles[profile.role] = sheet_name

                if profile.cached_error_cells:
                    issues.append(
                        ImportIssue(
                            severity="WARNING",
                            category="FORMULA_ERRORS",
                            sheet_name=sheet_name,
                            cell_address="",
                            item_key="",
                            message=(
                                f"{profile.cached_error_cells:,} cached formula-error "
                                "cells were detected. Error cells are excluded from "
                                "live updates."
                            ),
                            recommendation=(
                                "Review the source formulas. The original workbook "
                                "is archived unchanged."
                            ),
                        )
                    )

            plan_date = self._detect_plan_date(value_wb, path)
            self._progress(progress, 38, "Extracting weight master")
            weights = self._extract_weights(value_wb, roles, issues)

            self._progress(progress, 46, "Reading production, stock and shipment headers")
            stock_rows, shipment_rows, history_rows, actual_dates = self._extract_prod(
                value_wb,
                formula_wb,
                roles,
                plan_date,
                weights,
                issues,
                progress=progress,
            )

            for stock_row in stock_rows:
                negative_fields = {
                    field_name: int(stock_row.get(field_name) or 0)
                    for field_name in (
                        "fg_stock",
                        "qc_stock",
                        "scrap_stock",
                        "blocked_stock",
                    )
                    if int(stock_row.get(field_name) or 0) < 0
                }
                if negative_fields:
                    details = ", ".join(
                        f"{field_name}={value}"
                        for field_name, value in negative_fields.items()
                    )
                    issues.append(
                        ImportIssue(
                            severity="WARNING",
                            category="NEGATIVE_STOCK_NORMALIZED",
                            sheet_name=stock_row.get("source_sheet", "PROD"),
                            cell_address=f"D{stock_row.get('source_row', '')}",
                            item_key=stock_row.get("sap_code", ""),
                            message=(
                                f"Negative source stock detected ({details}). "
                                "Live stock tables will store zero for negative "
                                "stock quantities so database integrity rules are "
                                "not violated."
                            ),
                            recommendation=(
                                "Review the exact archived workbook row. The "
                                "original negative value remains available in the "
                                "source workbook and import evidence."
                            ),
                        )
                    )

            self._progress(progress, 58, "Extracting oven/cavity day and night plan")
            oven_resource_rows = self._extract_oven_resource_structure(value_wb, roles)
            oven_rows = self._extract_oven(value_wb, roles, plan_date, issues)

            self._progress(progress, 68, "Extracting compound and BOM mappings")
            compound_rows = self._extract_compound(
                value_wb, roles, issues, progress=progress
            )

            self._progress(progress, 76, "Extracting bead requirements")
            bead_rows = self._extract_bead(value_wb, roles, issues)

            self._progress(progress, 79, "Extracting BAND mold-code master")
            band_rows = self._extract_band_master(value_wb, roles)

            self._progress(progress, 82, "Extracting band, core and shift material plans")
            material_rows = self._extract_material_plans(
                value_wb,
                roles,
                plan_date,
                issues,
                progress=progress,
            )

            self._add_cross_validation_issues(
                stock_rows,
                shipment_rows,
                oven_rows,
                weights,
                issues,
            )

            confidence = self._workbook_confidence(profiles)
            summary = self._build_summary(
                profiles=profiles,
                stock_rows=stock_rows,
                shipment_rows=shipment_rows,
                oven_rows=oven_rows,
                compound_rows=compound_rows,
                bead_rows=bead_rows,
                material_rows=material_rows,
                history_rows=history_rows,
                issues=issues,
            )
            # Factory rule confirmed from the OVEN workbook: PROD column D is
            # monthly opening-stock evidence, not a fresh daily stock count.
            summary["stock_source_semantics"] = "MONTHLY_OPENING_STOCK"
            summary["stock_source_sheet"] = roles.get("PRODUCTION_STOCK_SHIPMENTS", "PROD")
            summary["stock_source_column"] = "D"
            summary["physical_line_count"] = len({r.get("line_name") for r in oven_resource_rows if r.get("line_name")})
            summary["physical_cavity_position_count"] = len(oven_resource_rows)
            summary["band_mold_code_count"] = len(band_rows)

            self._progress(progress, 96, "Building professional import preview")
            analysis = WorkbookAnalysis(
                workbook_path=str(path),
                workbook_name=path.name,
                workbook_hash=digest,
                workbook_size_bytes=path.stat().st_size,
                plan_date=plan_date.isoformat() if plan_date else None,
                confidence_score=confidence,
                detected_type="MPPS OVEN PRODUCTION WORKBOOK",
                sheet_profiles=profiles,
                issues=issues,
                stock_rows=stock_rows,
                shipment_rows=shipment_rows,
                oven_rows=oven_rows,
                oven_resource_rows=oven_resource_rows,
                compound_rows=compound_rows,
                bead_rows=bead_rows,
                band_rows=band_rows,
                material_plan_rows=material_rows,
                production_history_rows=history_rows,
                actual_production_dates=actual_dates,
                summary=summary,
            )
            self._progress(progress, 100, "Analysis complete")
            return analysis
        finally:
            formula_wb.close()
            value_wb.close()

    def commit(
        self,
        analysis: WorkbookAnalysis | dict[str, Any],
        *,
        options: dict[str, bool] | None = None,
        imported_by: str = "",
        progress: ProgressCallback | None = None,
    ) -> dict[str, Any]:
        if isinstance(analysis, dict):
            analysis = _analysis_from_dict(analysis)

        options = {
            "archive_source": True,
            # PROD column D is monthly opening-stock evidence, not a daily/live
            # stock snapshot.  V10 captures it through FactoryIntelligenceService;
            # legacy stock-cache writes stay disabled unless an old integration
            # explicitly opts in.
            "update_stock": False,
            "update_daily_stock": False,
            "update_blank_weights": True,
            "overwrite_existing_weights": False,
            "import_oven_plan": True,
            "import_materials": True,
            "import_shipment_snapshots": True,
            "sync_live_shipments": True,
            "create_draft_shipments": False,
            "auto_detect_import_mode": True,
            "force_historical_snapshot": False,
            "force_live_revision": False,
            "mark_missing_shipments": True,
            "sync_deferred_shipments": True,
            "authoritative_latest_shipments": True,
            "protect_manual_fields": True,
            "capture_learning_observations": True,
            "rebuild_learning_models": True,
            "defer_factory_intelligence_training": False,
            "import_production_history": True,
            **(options or {}),
        }
        # Backward-compatible alias used by V6 screens/installations.
        if options.get("create_draft_shipments"):
            options["sync_live_shipments"] = True
        if options.get("authoritative_latest_shipments", False):
            # Business rule: the newest workbook is the FINAL shipment truth.
            # Every non-zero shipment column participates in the live snapshot.
            options["sync_live_shipments"] = True
            options["mark_missing_shipments"] = True
            options["sync_deferred_shipments"] = True

        if analysis.confidence_score < 0.55:
            raise RuntimeError(
                "Workbook confidence is below the safe-import threshold. "
                "Use Analyze Only and resolve sheet/header issues first."
            )

        workbook_path = Path(analysis.workbook_path)
        if not workbook_path.exists():
            raise FileNotFoundError(workbook_path)
        if _sha256_file(workbook_path) != analysis.workbook_hash:
            raise RuntimeError(
                "The workbook changed after analysis. Analyze the file again before importing."
            )

        archive_path = ""
        plan_date = (
            date.fromisoformat(analysis.plan_date)
            if analysis.plan_date
            else date.today()
        )

        bulk_history = str(os.environ.get("MPPS_R741_BULK_HISTORY") or "").strip().lower() in {
            "1", "true", "yes", "on"
        }

        with _get_session() as session:
            sync_service = WorkbookContinuousSyncService(self.project_root)
            learning_service = ProductionLearningService()
            ai_planning_service = AIPlanningService()
            factory_intelligence = FactoryIntelligenceService()
            resource_intelligence = FactoryResourceIntelligenceService()
            monthly_stock = MonthlyStockSnapshotService()

            # R7.4.1: all historical workbooks in this process target the same
            # already-verified C: PostgreSQL cluster. Replaying dozens of CREATE /
            # ALTER / index checks for every workbook costs far more than the
            # actual data write. Run the complete additive preflight once; normal
            # app imports (without MPPS_R741_BULK_HISTORY) retain old behavior.
            if (not bulk_history) or (not self.__class__._r741_schema_ready):
                self._progress(progress, 2, "Creating intelligent import schema")
                self.ensure_schema(session)
                sync_service.ensure_schema(session)
                learning_service.ensure_schema(session)
                ai_planning_service.ensure_schema(session)
                factory_intelligence.ensure_schema(session)
                resource_intelligence.ensure_schema(session)
                monthly_stock.ensure_schema(session)
                if bulk_history:
                    self.__class__._r741_schema_ready = True

            # Resolve recoverable SAP/description mismatches before shipment-sync
            # preview.  Only very-high-confidence matches are auto-corrected;
            # ambiguous rows stay unchanged and are retained for review/history.
            identity_started = datetime.now()
            identity_preview = factory_intelligence.resolve_analysis(
                session, analysis, import_run_id=None, persist=False
            )
            if bulk_history:
                identity_seconds = (datetime.now() - identity_started).total_seconds()
                print(
                    f"[R7.4.4 ID MAX] {analysis.workbook_name}: "
                    f"{int(identity_preview.get('identity_unique_pairs') or 0)} unique pairs "
                    f"in {identity_seconds:.1f}s",
                    flush=True,
                )

            # A few long-lived local MPPS databases were restored/imported
            # from backups over time.  PostgreSQL row data can then be ahead
            # of the SERIAL/BIGSERIAL sequence even though the table itself is
            # healthy.  Advancing a sequence to the current MAX(id) is safe
            # and prevents false duplicate-primary-key IntegrityErrors during
            # the transactional import.
            # Repairing every BIGSERIAL with MAX(id) on every workbook becomes
            # progressively slower as the historical corpus grows. Sequences only
            # need this recovery pass once per verified training process; normal
            # inserts advance them automatically afterwards.
            if (not bulk_history) or (not self.__class__._r741_sequences_repaired):
                self._repair_serial_sequences(session)
                if bulk_history:
                    self.__class__._r741_sequences_repaired = True

            sync_preview = sync_service.preview_with_session(
                session,
                analysis,
                options,
            )
            effective_options = dict(options)
            effective_options["resolved_import_mode"] = sync_preview.mode
            effective_options["resolved_import_reason"] = sync_preview.reason
            if sync_preview.mode == "HISTORICAL":
                # Older workbooks are ML/history evidence only. They may enrich
                # plan-vs-actual learning and dated snapshots, but they must never
                # move the operational factory state backwards or overwrite current
                # material/weight masters from the newest live OVEN workbook.
                effective_options["update_stock"] = False
                effective_options["update_daily_stock"] = False
                effective_options["sync_live_shipments"] = False
                effective_options["update_blank_weights"] = False
                effective_options["overwrite_existing_weights"] = False
                effective_options["import_materials"] = False

            duplicate = session.execute(
                text(
                    """
                    SELECT id, status
                    FROM excel_import_runs
                    WHERE workbook_hash = :workbook_hash
                      AND status IN ('COMMITTED', 'COMMITTED WITH WARNINGS')
                      AND rollback_at IS NULL
                    ORDER BY id DESC
                    LIMIT 1
                    """
                ),
                {"workbook_hash": analysis.workbook_hash},
            ).mappings().first()
            if duplicate:
                raise RuntimeError(
                    "This exact workbook was already committed as import run "
                    f"#{duplicate['id']}. Roll it back or choose another file."
                )

            run_id = session.execute(
                text(
                    """
                    INSERT INTO excel_import_runs (
                        workbook_name,
                        workbook_path,
                        workbook_hash,
                        workbook_size_bytes,
                        detected_type,
                        confidence_score,
                        plan_date,
                        status,
                        imported_by,
                        options_json,
                        analysis_json,
                        started_at
                    ) VALUES (
                        :workbook_name,
                        :workbook_path,
                        :workbook_hash,
                        :workbook_size_bytes,
                        :detected_type,
                        :confidence_score,
                        :plan_date,
                        'COMMITTING',
                        :imported_by,
                        CAST(:options_json AS JSONB),
                        CAST(:analysis_json AS JSONB),
                        CURRENT_TIMESTAMP
                    )
                    RETURNING id
                    """
                ),
                {
                    "workbook_name": analysis.workbook_name,
                    "workbook_path": str(workbook_path),
                    "workbook_hash": analysis.workbook_hash,
                    "workbook_size_bytes": analysis.workbook_size_bytes,
                    "detected_type": analysis.detected_type,
                    "confidence_score": analysis.confidence_score,
                    "plan_date": plan_date,
                    "imported_by": imported_by or "Local User",
                    "options_json": json.dumps(
                        {
                            **options,
                            "resolved_import_mode": sync_preview.mode,
                            "resolved_import_reason": sync_preview.reason,
                        },
                        default=str,
                    ),
                    "analysis_json": json.dumps(
                        analysis.to_dict(include_rows=False),
                        default=str,
                    ),
                },
            ).scalar_one()

            try:
                if options["archive_source"]:
                    self._progress(progress, 8, "Archiving the exact original workbook")
                    archive_path = self._archive_workbook(workbook_path, analysis.workbook_hash)

                # Persist the identity-resolution evidence inside the same
                # transaction as the workbook import.  A rollback therefore also
                # rolls back learned aliases from this workbook.
                if bulk_history:
                    # R7.4.3: preview already performed the expensive deterministic
                    # identity resolution and mutated high-confidence rows. Persist
                    # that exact result without resolving the workbook a second time.
                    identity_result = dict(identity_preview)
                    identity_result.update(
                        factory_intelligence.persist_identity_resolutions(
                            session,
                            resolutions=dict(identity_preview.get("_resolutions") or {}),
                            import_run_id=run_id,
                            analysis=analysis,
                            plan_date=plan_date,
                        )
                    )
                else:
                    identity_result = factory_intelligence.resolve_analysis(
                        session, analysis, import_run_id=run_id, persist=True
                    )

                self._save_profiles_and_issues(session, run_id, analysis)
                self._save_workbook_registry(
                    session,
                    run_id,
                    analysis,
                    archive_path or str(workbook_path),
                    imported_by,
                )

                counters = Counter()
                shipment_sync_result: dict[str, Any] = {
                    "sync_mode": sync_preview.mode,
                    "sync_reason": sync_preview.reason,
                    **sync_preview.summary,
                }
                learning_result: dict[str, Any] = {}
                if effective_options["update_stock"] or effective_options["update_daily_stock"]:
                    self._progress(progress, 18, "Updating stock snapshot transactionally")
                    self._commit_stock(
                        session,
                        run_id,
                        analysis,
                        plan_date,
                        effective_options,
                        counters,
                    )

                if options["import_shipment_snapshots"]:
                    self._progress(progress, 38, "Importing shipment demand snapshots")
                    shipment_sync_result = self._commit_shipment_snapshots(
                        session,
                        run_id,
                        analysis,
                        plan_date,
                        effective_options,
                        counters,
                    )

                if options["import_oven_plan"]:
                    # V11.2 adds first-class mold_code/line/cavity columns used by
                    # the import itself, so migrate before writing mpps_oven_plan.
                    resource_intelligence.ensure_schema(session)
                    self._progress(progress, 53, "Importing oven, cavity, day and night plans")
                    self._commit_oven_plan(
                        session,
                        run_id,
                        analysis,
                        counters,
                    )
                    # V11 stores the resource configuration before aggregation so
                    # future ML can learn line/cavity/mold/casing effects without
                    # reparsing 5 years of raw Excel on every model run.
                    self._progress(progress, 58, "Capturing lossless factory resource allocations")
                    resource_result = resource_intelligence.capture_workbook_resources(
                        session,
                        import_run_id=run_id,
                        analysis=analysis,
                        import_mode=sync_preview.mode,
                    )
                    # _commit_oven_plan already writes the first-class line /
                    # cavity / mold columns. The legacy backfill UPDATE is redundant
                    # during historical bulk ingestion and is retained for normal
                    # interactive/live imports.
                    if not bulk_history:
                        resource_intelligence.sync_operational_oven_columns(
                            session,
                            import_run_id=run_id,
                        )
                    counters.update({
                        k: v for k, v in resource_result.items()
                        if isinstance(v, int)
                    })
                    learning_result["factory_resources"] = resource_result

                if options["import_materials"]:
                    self._progress(progress, 68, "Updating compound, BOM, bead, band and core data")
                    self._commit_materials(
                        session,
                        run_id,
                        analysis,
                        counters,
                    )

                if options["import_production_history"]:
                    self._progress(progress, 82, "Importing verified day/night actual production history")
                    self._commit_production_history(
                        session,
                        run_id,
                        analysis,
                        counters,
                    )

                # PROD column D is the factory's monthly opening-stock source.
                # Capture it as dated evidence for every workbook; only a LIVE
                # workbook can update the month's operational opening authority.
                self._progress(progress, 86, "Capturing monthly opening stock from PROD STOCK")
                opening_result = factory_intelligence.capture_opening_stock(
                    session,
                    import_run_id=run_id,
                    analysis=analysis,
                    import_mode=sync_preview.mode,
                )
                for key, value in opening_result.items():
                    if isinstance(value, int):
                        counters[key] += value

                self._progress(progress, 88, "Capturing Monthly Stock LIVE / FINAL snapshots")
                monthly_stock_result = monthly_stock.capture_import(
                    session,
                    import_run_id=run_id,
                    analysis=analysis,
                    import_mode=sync_preview.mode,
                )
                learning_result["monthly_stock"] = monthly_stock_result
                for key, value in monthly_stock_result.items():
                    if isinstance(value, int):
                        counters[key] += value

                if options.get("capture_learning_observations", True):
                    learning_started = datetime.now()
                    self._progress(
                        progress,
                        90,
                        "Capturing local learning observations",
                    )
                    observation_result = (
                        learning_service.capture_import_observations(
                            session,
                            import_run_id=run_id,
                            analysis=analysis,
                        )
                    )
                    learning_result.update(observation_result)
                    counters.update(observation_result)
                    # Save the human FINAL Excel planning decision against the
                    # workbook's own shipment/production requirement. This becomes
                    # training data for the V10 planner-policy model.
                    reconciliation_result = learning_service.save_reconciliation(
                        session,
                        import_run_id=run_id,
                        analysis=analysis,
                    )
                    learning_result.update(reconciliation_result)
                    counters.update(reconciliation_result)
                    if options.get("rebuild_learning_models", True):
                        model_result = learning_service.rebuild_models(session)
                        learning_result.update(model_result)
                        counters.update(model_result)
                    if bulk_history:
                        learning_seconds = (datetime.now() - learning_started).total_seconds()
                        print(
                            f"[R7.4.3 LEARN BATCH] {analysis.workbook_name}: "
                            f"{learning_seconds:.1f}s",
                            flush=True,
                        )

                self._progress(
                    progress,
                    96,
                    "Reconciling final Excel plan, verified actuals and AI shadow plan",
                )
                ai_result = ai_planning_service.post_excel_import(
                    session,
                    import_run_id=run_id,
                    analysis=analysis,
                    import_mode=sync_preview.mode,
                )
                learning_result["ai_planner"] = ai_result
                for key, value in ai_result.items():
                    if isinstance(value, int):
                        counters[key] += value

                if options.get("defer_factory_intelligence_training", False):
                    self._progress(
                        progress,
                        98,
                        "Deferring factory-intelligence retraining until historical ingestion completes",
                    )
                    learning_result["factory_intelligence"] = {
                        "training_deferred": True,
                        **{k: v for k, v in identity_result.items() if not k.startswith("_")},
                        **opening_result,
                    }
                else:
                    self._progress(progress, 98, "Training real factory capacity and resource-intelligence models")
                    capacity_result = factory_intelligence.train_capacity_models(session)
                    planner_policy_result = factory_intelligence.train_planner_policy(session)
                    resource_capacity_result = resource_intelligence.train_profiles(session)
                    intelligence_state = factory_intelligence.refresh_state(session)
                    learning_result["factory_intelligence"] = {
                        **capacity_result,
                        **planner_policy_result,
                        **resource_capacity_result,
                        **intelligence_state,
                        **{k: v for k, v in identity_result.items() if not k.startswith("_")},
                        **opening_result,
                    }
                    for key, value in capacity_result.items():
                        if isinstance(value, int):
                            counters[key] += value
                    for key, value in planner_policy_result.items():
                        if isinstance(value, int):
                            counters[key] += value
                    for key, value in resource_capacity_result.items():
                        if isinstance(value, int):
                            counters[key] += value
                for key, value in identity_result.items():
                    if isinstance(value, int):
                        counters[key] += value

                warning_count = sum(
                    1 for issue in analysis.issues if issue.severity != "INFO"
                )
                status = (
                    "COMMITTED WITH WARNINGS"
                    if warning_count
                    else "COMMITTED"
                )
                result = {
                    "run_id": run_id,
                    "status": status,
                    "workbook": analysis.workbook_name,
                    "archive_path": archive_path,
                    "plan_date": plan_date.isoformat(),
                    "warnings": warning_count,
                    "import_mode": sync_preview.mode,
                    "import_mode_reason": sync_preview.reason,
                    "shipment_sync": shipment_sync_result,
                    "learning": learning_result,
                    "changes": dict(counters),
                }
                session.execute(
                    text(
                        """
                        UPDATE excel_import_runs
                        SET status = :status,
                            archive_path = :archive_path,
                            result_json = CAST(:result_json AS JSONB),
                            completed_at = CURRENT_TIMESTAMP
                        WHERE id = :run_id
                        """
                    ),
                    {
                        "status": status,
                        "archive_path": archive_path,
                        "result_json": json.dumps(result, default=str),
                        "run_id": run_id,
                    },
                )
                self._progress(progress, 100, "Safe Excel import committed")
                return result
            except Exception:
                # PostgreSQL marks the transaction as aborted after any statement
                # failure. Do not issue another UPDATE here because that masks the
                # original database error with "current transaction is aborted".
                # The outer get_session() context rolls the full import back.
                raise

    def preview_shipment_sync(
        self,
        analysis: WorkbookAnalysis | dict[str, Any],
        *,
        options: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        if isinstance(analysis, dict):
            analysis = _analysis_from_dict(analysis)
        return WorkbookContinuousSyncService(
            self.project_root
        ).preview(analysis, options)

    def finalize_post_plan(
        self,
        *,
        import_run_id: int,
        analysis: WorkbookAnalysis | dict[str, Any],
    ) -> dict[str, Any]:
        if isinstance(analysis, dict):
            analysis = _analysis_from_dict(analysis)
        from app.database import get_session

        with get_session() as session:
            learning = ProductionLearningService()
            result = learning.save_reconciliation(
                session,
                import_run_id=int(import_run_id),
                analysis=analysis,
            )
            return result

    def list_history(self, limit: int = 50) -> list[dict[str, Any]]:
        with _get_session() as session:
            self.ensure_schema(session)
            rows = session.execute(
                text(
                    """
                    SELECT
                        id,
                        workbook_name,
                        workbook_hash,
                        detected_type,
                        confidence_score,
                        plan_date,
                        status,
                        imported_by,
                        started_at,
                        completed_at,
                        rollback_at,
                        archive_path,
                        result_json,
                        error_text
                    FROM excel_import_runs
                    ORDER BY id DESC
                    LIMIT :limit
                    """
                ),
                {"limit": max(1, min(500, int(limit)))},
            ).mappings().all()
            return [dict(row) for row in rows]

    def rollback(self, run_id: int, *, rolled_back_by: str = "") -> dict[str, Any]:
        with _get_session() as session:
            self.ensure_schema(session)
            MonthlyStockSnapshotService.ensure_schema(session)
            run = session.execute(
                text(
                    """
                    SELECT id, status, rollback_at
                    FROM excel_import_runs
                    WHERE id = :run_id
                    FOR UPDATE
                    """
                ),
                {"run_id": int(run_id)},
            ).mappings().first()
            if not run:
                raise RuntimeError(f"Import run #{run_id} was not found.")
            if run["rollback_at"] is not None:
                raise RuntimeError(f"Import run #{run_id} was already rolled back.")
            if not str(run["status"]).startswith("COMMITTED"):
                raise RuntimeError("Only committed import runs can be rolled back.")

            changes = session.execute(
                text(
                    """
                    SELECT id, table_name, action, key_json, before_json, after_json
                    FROM excel_import_changes
                    WHERE run_id = :run_id
                    ORDER BY id DESC
                    """
                ),
                {"run_id": int(run_id)},
            ).mappings().all()

            restored = 0
            removed = 0
            for change in changes:
                table_name = change["table_name"]
                if table_name not in LIVE_TABLE_COLUMNS:
                    continue
                key_data = _json_object(change["key_json"])
                before_data = _json_object(change["before_json"])
                if change["action"] == "INSERT":
                    removed += self._delete_by_key(session, table_name, key_data)
                elif change["action"] in {"UPDATE", "DELETE"} and before_data:
                    restored += self._restore_by_key(
                        session,
                        table_name,
                        key_data,
                        before_data,
                    )

            for table in (
                "excel_import_shipment_snapshots",
                "excel_import_material_plans",
                "excel_import_production_history",
            ):
                session.execute(
                    text(f"DELETE FROM {table} WHERE run_id = :run_id"),
                    {"run_id": int(run_id)},
                )
            session.execute(
                text(
                    "DELETE FROM mpps_monthly_stock_snapshots "
                    "WHERE import_run_id = :run_id"
                ),
                {"run_id": int(run_id)},
            )
            session.execute(
                text(
                    "DELETE FROM excel_learning_observations "
                    "WHERE import_run_id = :run_id"
                ),
                {"run_id": int(run_id)},
            )
            session.execute(
                text(
                    "DELETE FROM excel_plan_reconciliation "
                    "WHERE import_run_id = :run_id"
                ),
                {"run_id": int(run_id)},
            )
            session.execute(
                text(
                    "UPDATE excel_shipment_sync_runs "
                    "SET status = 'ROLLED BACK', rollback_at = CURRENT_TIMESTAMP "
                    "WHERE import_run_id = :run_id"
                ),
                {"run_id": int(run_id)},
            )

            session.execute(
                text(
                    """
                    UPDATE excel_import_runs
                    SET status = 'ROLLED BACK',
                        rollback_at = CURRENT_TIMESTAMP,
                        rollback_by = :rollback_by
                    WHERE id = :run_id
                    """
                ),
                {
                    "run_id": int(run_id),
                    "rollback_by": rolled_back_by or "Local User",
                },
            )
            return {
                "run_id": int(run_id),
                "status": "ROLLED BACK",
                "restored_rows": restored,
                "removed_rows": removed,
            }

    @staticmethod
    def _repair_serial_sequences(session) -> None:
        tables = (
            "excel_import_runs",
            "excel_import_sheet_profiles",
            "excel_import_issues",
            "excel_import_changes",
            "excel_import_shipment_snapshots",
            "excel_import_material_plans",
            "excel_import_production_history",
            "excel_shipment_identities",
            "excel_shipment_sync_runs",
            "excel_shipment_sync_rows",
            "excel_shipment_item_revisions",
            "excel_authoritative_shipment_archive",
            "mpps_sap_stock_items",
            "mpps_stock_items",
            "mpps_daily_stock_entries",
            "mpps_oven_plan",
            "mpps_compound_master",
            "mpps_bom_items",
            "mpps_bead_master",
            "mpps_shipments",
            "mpps_shipment_items",
            "mpps_final_plan_history",
            "mpps_actual_production",
            "mpps_plan_actual_reconciliation",
            "mpps_ai_model_state",
            "mpps_ai_plan_runs",
            "mpps_ai_plan_items",
            "mpps_ai_plan_evaluation",
        )
        for table_name in tables:
            # Keep every optional probe behind a SAVEPOINT.  If an old install
            # does not have a table/sequence, only this probe is rolled back;
            # the outer import transaction remains healthy.
            try:
                with session.begin_nested():
                    sequence_name = session.execute(
                        text(
                            "SELECT pg_get_serial_sequence(:table_name, 'id')"
                        ),
                        {"table_name": table_name},
                    ).scalar()
                    if not sequence_name:
                        continue
                    max_id = session.execute(
                        text(f"SELECT COALESCE(MAX(id), 0) FROM {table_name}")
                    ).scalar()
                    max_id = int(max_id or 0)
                    if max_id > 0:
                        session.execute(
                            text(
                                "SELECT setval(CAST(:sequence_name AS regclass), "
                                ":max_id, TRUE)"
                            ),
                            {
                                "sequence_name": str(sequence_name),
                                "max_id": max_id,
                            },
                        )
                    else:
                        session.execute(
                            text(
                                "SELECT setval(CAST(:sequence_name AS regclass), "
                                "1, FALSE)"
                            ),
                            {"sequence_name": str(sequence_name)},
                        )
            except Exception:
                continue

    @staticmethod
    def ensure_schema(session) -> None:
        statements = [
            """
            CREATE TABLE IF NOT EXISTS excel_import_runs (
                id BIGSERIAL PRIMARY KEY,
                workbook_name TEXT NOT NULL,
                workbook_path TEXT NOT NULL DEFAULT '',
                archive_path TEXT NOT NULL DEFAULT '',
                workbook_hash VARCHAR(64) NOT NULL,
                workbook_size_bytes BIGINT NOT NULL DEFAULT 0,
                detected_type TEXT NOT NULL DEFAULT '',
                confidence_score NUMERIC(8,5) NOT NULL DEFAULT 0,
                plan_date DATE,
                status VARCHAR(50) NOT NULL DEFAULT 'ANALYZED',
                imported_by TEXT NOT NULL DEFAULT '',
                options_json JSONB NOT NULL DEFAULT '{}'::jsonb,
                analysis_json JSONB NOT NULL DEFAULT '{}'::jsonb,
                result_json JSONB NOT NULL DEFAULT '{}'::jsonb,
                error_text TEXT NOT NULL DEFAULT '',
                started_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                completed_at TIMESTAMP,
                rollback_at TIMESTAMP,
                rollback_by TEXT NOT NULL DEFAULT ''
            )
            """,
            """
            CREATE INDEX IF NOT EXISTS ix_excel_import_runs_hash
            ON excel_import_runs(workbook_hash, status)
            """,
            """
            CREATE TABLE IF NOT EXISTS excel_import_sheet_profiles (
                id BIGSERIAL PRIMARY KEY,
                run_id BIGINT NOT NULL REFERENCES excel_import_runs(id) ON DELETE CASCADE,
                sheet_name TEXT NOT NULL,
                detected_role TEXT NOT NULL,
                confidence_score NUMERIC(8,5) NOT NULL DEFAULT 0,
                max_row INTEGER NOT NULL DEFAULT 0,
                max_column INTEGER NOT NULL DEFAULT 0,
                nonempty_cells BIGINT NOT NULL DEFAULT 0,
                formula_cells BIGINT NOT NULL DEFAULT 0,
                cached_error_cells BIGINT NOT NULL DEFAULT 0,
                header_row INTEGER,
                evidence TEXT NOT NULL DEFAULT ''
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS excel_import_issues (
                id BIGSERIAL PRIMARY KEY,
                run_id BIGINT NOT NULL REFERENCES excel_import_runs(id) ON DELETE CASCADE,
                severity VARCHAR(20) NOT NULL,
                category VARCHAR(80) NOT NULL,
                sheet_name TEXT NOT NULL DEFAULT '',
                cell_address TEXT NOT NULL DEFAULT '',
                item_key TEXT NOT NULL DEFAULT '',
                message TEXT NOT NULL,
                recommendation TEXT NOT NULL DEFAULT '',
                resolution_status VARCHAR(30) NOT NULL DEFAULT 'OPEN',
                created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
            """,
            """
            CREATE INDEX IF NOT EXISTS ix_excel_import_issues_run
            ON excel_import_issues(run_id, severity, category)
            """,
            """
            CREATE TABLE IF NOT EXISTS excel_import_changes (
                id BIGSERIAL PRIMARY KEY,
                run_id BIGINT NOT NULL REFERENCES excel_import_runs(id) ON DELETE CASCADE,
                table_name TEXT NOT NULL,
                action VARCHAR(20) NOT NULL,
                key_json JSONB NOT NULL DEFAULT '{}'::jsonb,
                before_json JSONB NOT NULL DEFAULT '{}'::jsonb,
                after_json JSONB NOT NULL DEFAULT '{}'::jsonb,
                created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
            """,
            """
            CREATE INDEX IF NOT EXISTS ix_excel_import_changes_run
            ON excel_import_changes(run_id, id)
            """,
            """
            CREATE TABLE IF NOT EXISTS excel_import_shipment_snapshots (
                id BIGSERIAL PRIMARY KEY,
                run_id BIGINT NOT NULL REFERENCES excel_import_runs(id) ON DELETE CASCADE,
                shipment_column TEXT NOT NULL,
                shipment_name TEXT NOT NULL,
                source_status TEXT NOT NULL DEFAULT '',
                source_item_code TEXT NOT NULL,
                item_description TEXT NOT NULL DEFAULT '',
                quantity INTEGER NOT NULL DEFAULT 0,
                plan_date DATE,
                live_shipment_id BIGINT,
                created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(run_id, shipment_column, source_item_code)
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS excel_import_material_plans (
                id BIGSERIAL PRIMARY KEY,
                run_id BIGINT NOT NULL REFERENCES excel_import_runs(id) ON DELETE CASCADE,
                plan_date DATE,
                material_type VARCHAR(30) NOT NULL,
                material_key TEXT NOT NULL,
                material_description TEXT NOT NULL DEFAULT '',
                day_qty NUMERIC(18,5) NOT NULL DEFAULT 0,
                night_qty NUMERIC(18,5) NOT NULL DEFAULT 0,
                total_qty NUMERIC(18,5) NOT NULL DEFAULT 0,
                produced_qty NUMERIC(18,5) NOT NULL DEFAULT 0,
                stock_qty NUMERIC(18,5) NOT NULL DEFAULT 0,
                next_day_qty NUMERIC(18,5) NOT NULL DEFAULT 0,
                unit VARCHAR(20) NOT NULL DEFAULT 'PCS',
                source_sheet TEXT NOT NULL,
                source_row INTEGER NOT NULL,
                source_json JSONB NOT NULL DEFAULT '{}'::jsonb,
                created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS excel_import_production_history (
                id BIGSERIAL PRIMARY KEY,
                run_id BIGINT NOT NULL REFERENCES excel_import_runs(id) ON DELETE CASCADE,
                production_date DATE NOT NULL,
                sap_code TEXT NOT NULL,
                item_description TEXT NOT NULL DEFAULT '',
                production_qty INTEGER NOT NULL DEFAULT 0,
                source_sheet TEXT NOT NULL,
                source_row INTEGER NOT NULL,
                created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(run_id, production_date, sap_code)
            )
            """,
            """
            ALTER TABLE excel_raw_cells
            ADD COLUMN IF NOT EXISTS number_value NUMERIC
            """,
            """
            ALTER TABLE excel_raw_cells
            ADD COLUMN IF NOT EXISTS date_value TIMESTAMP
            """,
            """
            ALTER TABLE excel_raw_cells
            ADD COLUMN IF NOT EXISTS boolean_value BOOLEAN
            """,
            """
            ALTER TABLE excel_raw_cells
            ADD COLUMN IF NOT EXISTS cached_value TEXT
            """,
            """
            ALTER TABLE excel_raw_cells
            ADD COLUMN IF NOT EXISTS error_value TEXT
            """,
            """
            ALTER TABLE excel_raw_cells
            ADD COLUMN IF NOT EXISTS style_id INTEGER
            """,
            """
            ALTER TABLE excel_raw_cells
            ADD COLUMN IF NOT EXISTS import_run_id BIGINT
            """,
        ]
        for statement in statements:
            try:
                session.execute(text(statement))
            except Exception:
                # The raw Excel foundation may not exist in a very old install.
                if "ALTER TABLE excel_raw_cells" not in statement:
                    raise
                session.rollback()
                session.execute(
                    text(
                        """
                        CREATE TABLE IF NOT EXISTS excel_raw_cells (
                            id BIGSERIAL PRIMARY KEY,
                            sheet_id BIGINT,
                            row_number INTEGER NOT NULL,
                            column_number INTEGER NOT NULL,
                            column_letter TEXT NOT NULL,
                            cell_address TEXT NOT NULL,
                            raw_value TEXT,
                            display_value TEXT,
                            formula_value TEXT,
                            is_formula BOOLEAN NOT NULL DEFAULT FALSE,
                            data_type TEXT,
                            number_value NUMERIC,
                            date_value TIMESTAMP,
                            boolean_value BOOLEAN,
                            cached_value TEXT,
                            error_value TEXT,
                            style_id INTEGER,
                            import_run_id BIGINT,
                            created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
                        )
                        """
                    )
                )

        WorkbookContinuousSyncService.ensure_schema(session)
        ProductionLearningService.ensure_schema(session)

    def _profile_sheet(self, value_ws, formula_ws, fast_stats: dict[str, int] | None = None) -> SheetProfile:
        sample_values: list[str] = []
        nonempty = 0
        errors = 0
        formula_count = 0
        header_row = None
        best_header_score = 0

        for row_index, row in enumerate(
            value_ws.iter_rows(min_row=1, max_row=min(value_ws.max_row, 40)),
            start=1,
        ):
            row_words: list[str] = []
            for cell in row:
                value = cell.value
                if value not in (None, ""):
                    text_value = _text(value)
                    sample_values.append(text_value)
                    row_words.append(text_value)
            score = sum(
                1
                for word in row_words
                if _normalize(word)
                in {
                    "sap code",
                    "material description",
                    "tyre code",
                    "description",
                    "line",
                    "oven no",
                    "daily plan",
                    "weight",
                    "material",
                    "bead type",
                }
            )
            if score > best_header_score:
                best_header_score = score
                header_row = row_index

        role, confidence, evidence = self._classify_sheet(
            value_ws.title,
            sample_values,
        )

        if fast_stats:
            nonempty = int(fast_stats.get("nonempty_cells", 0) or 0)
            formula_count = int(fast_stats.get("formula_cells", 0) or 0)
            errors = int(fast_stats.get("cached_error_cells", 0) or 0)
        else:
            # Fallback for unusual non-OOXML inputs. Keep the scan bounded on large
            # sheets so diagnostics can never freeze the import UI.
            max_profile_rows = min(value_ws.max_row, 250)
            for row in value_ws.iter_rows(min_row=1, max_row=max_profile_rows):
                for cell in row:
                    value = cell.value
                    if value not in (None, ""):
                        nonempty += 1
                        if isinstance(value, str) and value.startswith("#"):
                            errors += 1
            for row in formula_ws.iter_rows(min_row=1, max_row=max_profile_rows):
                for cell in row:
                    value = cell.value
                    if isinstance(value, str) and value.startswith("="):
                        formula_count += 1

        return SheetProfile(
            sheet_name=value_ws.title,
            role=role,
            confidence=confidence,
            max_row=value_ws.max_row,
            max_column=value_ws.max_column,
            nonempty_cells=nonempty,
            formula_cells=formula_count,
            cached_error_cells=errors,
            header_row=header_row,
            evidence=evidence,
        )

    def _classify_sheet(
        self,
        sheet_name: str,
        sample_values: Iterable[str],
    ) -> tuple[str, float, str]:
        normalized_name = _normalize(sheet_name)
        normalized_sample = " | ".join(
            _normalize(value) for value in sample_values[:500]
        )
        best_role = "UNMAPPED_SOURCE_DATA"
        best_score = 0.0
        best_evidence = "No high-confidence semantic role matched."

        for role, definition in ROLE_DEFINITIONS.items():
            name_score = max(
                SequenceMatcher(None, normalized_name, _normalize(alias)).ratio()
                for alias in definition["names"]
            )
            direct_name = any(
                _normalize(alias) in normalized_name
                or normalized_name in _normalize(alias)
                for alias in definition["names"]
            )
            header_hits = [
                header
                for header in definition["headers"]
                if _normalize(header) in normalized_sample
            ]
            header_score = len(header_hits) / max(1, len(definition["headers"]))
            score = (0.58 * name_score) + (0.42 * header_score)
            if direct_name:
                score += 0.18
            score = min(1.0, score)
            if score > best_score:
                best_role = role
                best_score = score
                best_evidence = (
                    f"Sheet-name similarity {name_score:.2f}; "
                    f"header matches: {', '.join(header_hits) or 'none'}."
                )

        if best_score < 0.48:
            return "UNMAPPED_SOURCE_DATA", best_score, best_evidence
        return best_role, best_score, best_evidence

    def _detect_plan_date(self, workbook, path: Path) -> date | None:
        for sheet_name in workbook.sheetnames:
            if _normalize(sheet_name) == "daily plan":
                value = workbook[sheet_name]["C3"].value
                parsed = _as_date(value)
                if parsed:
                    return parsed
        month_names = {
            "january": 1,
            "february": 2,
            "march": 3,
            "april": 4,
            "may": 5,
            "june": 6,
            "july": 7,
            "august": 8,
            "auguest": 8,  # common spelling in the factory workbook filenames
            "september": 9,
            "october": 10,
            "november": 11,
            "december": 12,
        }
        match = re.search(
            r"(?i)(january|february|march|april|may|june|july|august|auguest|"
            r"september|october|november|december)\D+(\d{1,2})\D+(20\d{2})",
            path.stem,
        )
        if match:
            return date(
                int(match.group(3)),
                month_names[match.group(1).lower()],
                int(match.group(2)),
            )
        return None

    def _extract_weights(self, workbook, roles, issues) -> dict[str, float]:
        sheet_name = roles.get("WEIGHT_MASTER")
        if not sheet_name:
            issues.append(
                ImportIssue(
                    "BLOCKER",
                    "MISSING_WEIGHT_SHEET",
                    "",
                    "",
                    "",
                    "Weight master sheet was not detected.",
                    "Add a WGT/Weight sheet or map the sheet role manually.",
                )
            )
            return {}
        ws = workbook[sheet_name]
        weight_col = column_index_from_string("BR")
        weights: dict[str, float] = {}
        seen_descriptions: dict[str, str] = {}
        for row_number, row in enumerate(
            ws.iter_rows(min_row=4, values_only=True),
            start=4,
        ):
            code = _code(_value_at(row, 2))
            if not code:
                continue
            description = _text(_value_at(row, 3))
            weight = _number(_value_at(row, weight_col))
            if weight is None or weight <= 0:
                continue
            if code in weights and not math.isclose(weights[code], weight, rel_tol=1e-6):
                issues.append(
                    ImportIssue(
                        "WARNING",
                        "DUPLICATE_WEIGHT",
                        sheet_name,
                        f"B{row_number}",
                        code,
                        f"Conflicting weights {weights[code]} and {weight} kg.",
                        "The first positive weight is retained for safe import.",
                    )
                )
                continue
            weights[code] = weight
            seen_descriptions[code] = description
        return weights

    def _extract_prod(
        self,
        workbook,
        formula_workbook,
        roles,
        plan_date: date | None,
        weights: dict[str, float],
        issues: list[ImportIssue],
        *,
        progress: ProgressCallback | None = None,
    ) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
        """Extract PROD data without random-access scans on read-only worksheets.

        openpyxl read-only worksheets are optimized for sequential iteration. Calling
        ``ws.cell()`` repeatedly can replay the XML stream for every coordinate and
        made the 46% analysis stage appear frozen on large OVEN workbooks.  This
        implementation reads the header rows once, resolves any referenced shipment
        dates in one sequential pass, then processes the production rows once.
        """
        sheet_name = roles.get("PRODUCTION_STOCK_SHIPMENTS")
        if not sheet_name:
            issues.append(
                ImportIssue(
                    "BLOCKER",
                    "MISSING_PROD_SHEET",
                    "",
                    "",
                    "",
                    "Production/stock/shipment sheet was not detected.",
                    "Map the sheet containing SAP, stock and shipment columns.",
                )
            )
            return [], [], [], []

        ws = workbook[sheet_name]
        formula_ws = formula_workbook[sheet_name]
        shipment_start = column_index_from_string("BY")
        shipment_end = min(column_index_from_string("HP"), ws.max_column)
        history_start = column_index_from_string("G")
        history_end = min(column_index_from_string("BT"), ws.max_column)
        max_header_col = max(shipment_end, history_end)

        self._progress(progress, 47, "Caching PROD header rows")
        value_headers = list(
            ws.iter_rows(
                min_row=1,
                max_row=3,
                min_col=1,
                max_col=max_header_col,
                values_only=True,
            )
        )
        formula_headers = list(
            formula_ws.iter_rows(
                min_row=1,
                max_row=3,
                min_col=1,
                max_col=max_header_col,
                values_only=True,
            )
        )
        while len(value_headers) < 3:
            value_headers.append(tuple())
        while len(formula_headers) < 3:
            formula_headers.append(tuple())

        def header_value(rows, row_number: int, column: int):
            row = rows[row_number - 1]
            index = column - 1
            if 0 <= index < len(row):
                return row[index]
            return None

        # Build shipment metadata from cached rows.  Formula references to rows
        # below the header are collected first and resolved in one sequential read.
        raw_shipment_headers: list[dict[str, Any]] = []
        referenced_cells: dict[int, set[int]] = defaultdict(set)
        for column in range(shipment_start, shipment_end + 1):
            status = _text(header_value(value_headers, 1, column)).upper()
            name = _text(header_value(value_headers, 3, column))
            if not name:
                continue
            formula_text = _text(header_value(formula_headers, 1, column))
            column_letter = get_column_letter(column)
            referenced_rows = [
                int(value)
                for value in re.findall(
                    rf"{column_letter}(\d+)",
                    formula_text,
                    flags=re.IGNORECASE,
                )
            ]
            direct_date = _as_date(header_value(value_headers, 2, column))
            for referenced_row in referenced_rows:
                if 3 < referenced_row <= ws.max_row:
                    referenced_cells[referenced_row].add(column)
            raw_shipment_headers.append(
                {
                    "column": column,
                    "column_letter": column_letter,
                    "name": name,
                    "status": status,
                    "direct_date": direct_date,
                    "referenced_rows": referenced_rows,
                }
            )

        referenced_dates: dict[tuple[int, int], date] = {}
        if referenced_cells:
            self._progress(progress, 48, "Resolving shipment target-date references")
            min_ref_row = min(referenced_cells)
            max_ref_row = max(referenced_cells)
            # Reading BY:HP sequentially is much faster than hundreds of ws.cell()
            # calls on a read-only worksheet.
            for row_number, values in enumerate(
                ws.iter_rows(
                    min_row=min_ref_row,
                    max_row=max_ref_row,
                    min_col=shipment_start,
                    max_col=shipment_end,
                    values_only=True,
                ),
                start=min_ref_row,
            ):
                wanted_columns = referenced_cells.get(row_number)
                if not wanted_columns:
                    continue
                for column in wanted_columns:
                    index = column - shipment_start
                    value = values[index] if 0 <= index < len(values) else None
                    parsed = _as_date(value)
                    if parsed:
                        referenced_dates[(row_number, column)] = parsed

        shipment_headers: list[tuple[int, str, str, str, date | None, str]] = []
        for meta in raw_shipment_headers:
            source_target_date = meta["direct_date"]
            if source_target_date is None:
                for referenced_row in meta["referenced_rows"]:
                    source_target_date = referenced_dates.get(
                        (referenced_row, meta["column"])
                    )
                    if source_target_date:
                        break
            if source_target_date and 2020 <= source_target_date.year <= 2035:
                source_date_class = "EXCEL_APPROVED"
            elif source_target_date and source_target_date.year == 2060:
                source_date_class = "AUTO_TARGET_REQUIRED"
            elif source_target_date and source_target_date.year >= 2061:
                source_date_class = "DEFERRED_CONTROL"
            else:
                source_date_class = "DATE_MISSING"
            shipment_headers.append(
                (
                    meta["column"],
                    meta["column_letter"],
                    meta["name"],
                    meta["status"],
                    source_target_date,
                    source_date_class,
                )
            )

        # PROD production history is stored as paired shift columns. The first
        # column carries the date header and represents DAY; the following column
        # represents NIGHT. Only dates before the workbook plan date are verified
        # actual production.
        self._progress(progress, 49, "Mapping verified DAY/NIGHT actual-production columns")
        history_pairs: list[tuple[int, int | None, date]] = []
        column = history_start
        while column <= history_end:
            history_date = _as_date(header_value(value_headers, 3, column))
            if not history_date:
                column += 1
                continue
            next_column = column + 1 if column + 1 <= history_end else None
            next_date = (
                _as_date(header_value(value_headers, 3, next_column))
                if next_column is not None
                else None
            )
            night_column = (
                next_column
                if next_column is not None and next_date is None
                else None
            )
            if plan_date is None or history_date < plan_date:
                history_pairs.append((column, night_column, history_date))
            column += 2 if night_column is not None else 1

        actual_dates = [
            {
                "production_date": history_date.isoformat(),
                "source_day_column": get_column_letter(day_column),
                "source_night_column": (
                    get_column_letter(night_column)
                    if night_column is not None
                    else ""
                ),
                "source_sheet": sheet_name,
                "is_complete": True,
            }
            for day_column, night_column, history_date in history_pairs
        ]

        stock_rows: list[dict[str, Any]] = []
        shipment_rows: list[dict[str, Any]] = []
        history_rows: list[dict[str, Any]] = []
        duplicate_rows: dict[str, dict[str, Any]] = {}

        self._progress(progress, 50, "Scanning PROD items, shipments and actual production")
        total_data_rows = max(1, ws.max_row - 3)
        last_progress = -1
        for row_offset, row in enumerate(
            ws.iter_rows(min_row=4, values_only=True),
            start=0,
        ):
            row_number = row_offset + 4
            # Give the UI genuine activity through the formerly long 46% plateau.
            local_progress = 50 + min(7, int((row_offset / total_data_rows) * 8))
            if local_progress != last_progress:
                self._progress(
                    progress,
                    local_progress,
                    f"Scanning PROD rows {row_offset + 1:,}/{total_data_rows:,}",
                )
                last_progress = local_progress

            code = _code(_value_at(row, 2))
            description = _text(_value_at(row, 3))
            if not code or not description:
                continue
            if not _looks_like_item_code(code):
                issues.append(
                    ImportIssue(
                        "INFO",
                        "NON_STANDARD_ITEM_CODE",
                        sheet_name,
                        f"B{row_number}",
                        code,
                        "Non-standard item code retained in import evidence.",
                        "Confirm whether this is a SAP code before promoting it to master data.",
                    )
                )

            record = {
                "sap_code": code,
                "description": description,
                "fg_stock": _integer(_value_at(row, 4)),
                "opening_stock_qty": _integer(_value_at(row, 4)),
                "stock_source_semantics": "MONTHLY_OPENING_STOCK",
                "scrap_stock": _integer(_value_at(row, 5)),
                "blocked_stock": _integer(_value_at(row, 6)),
                "qc_stock": 0,
                "total_shipment": _integer(_value_at(row, column_index_from_string("HR"))),
                "total_available": _integer(_value_at(row, column_index_from_string("HS"))),
                "production_required": _integer(_value_at(row, column_index_from_string("HT"))),
                "planned_today": _integer(_value_at(row, column_index_from_string("HU"))),
                "remaining_to_plan": _integer(_value_at(row, column_index_from_string("HV"))),
                "weight_kg": weights.get(code),
                "source_sheet": sheet_name,
                "source_row": row_number,
            }
            if code in duplicate_rows:
                previous = duplicate_rows[code]
                if _normalize(previous["description"]) != _normalize(description):
                    issues.append(
                        ImportIssue(
                            "WARNING",
                            "DUPLICATE_ITEM_DESCRIPTION",
                            sheet_name,
                            f"B{row_number}",
                            code,
                            "Duplicate item code has a conflicting description.",
                            "The first row is used for master update; all source rows remain archived.",
                        )
                    )
                continue
            duplicate_rows[code] = record
            stock_rows.append(record)

            for (
                shipment_column,
                column_letter,
                shipment_name,
                status,
                source_target_date,
                source_date_class,
            ) in shipment_headers:
                qty = _integer(_value_at(row, shipment_column))
                if qty <= 0:
                    continue
                shipment_rows.append(
                    {
                        "shipment_column": column_letter,
                        "shipment_name": shipment_name,
                        "source_status": status,
                        "source_target_date": (
                            source_target_date.isoformat()
                            if source_target_date
                            else None
                        ),
                        "source_date_class": source_date_class,
                        "sap_code": code,
                        "description": description,
                        "quantity": qty,
                        "source_sheet": sheet_name,
                        "source_row": row_number,
                    }
                )

            for day_column, night_column, production_date in history_pairs:
                day_qty = max(0, _integer(_value_at(row, day_column)))
                night_qty = (
                    max(0, _integer(_value_at(row, night_column)))
                    if night_column is not None
                    else 0
                )
                qty = day_qty + night_qty
                if qty <= 0:
                    continue
                history_rows.append(
                    {
                        "production_date": production_date.isoformat(),
                        "sap_code": code,
                        "description": description,
                        "production_qty": qty,
                        "day_actual_qty": day_qty,
                        "night_actual_qty": night_qty,
                        "source_day_column": get_column_letter(day_column),
                        "source_night_column": (
                            get_column_letter(night_column)
                            if night_column is not None
                            else ""
                        ),
                        "source_semantics": "VERIFIED_ACTUAL_PRODUCTION",
                        "source_sheet": sheet_name,
                        "source_row": row_number,
                    }
                )

        self._progress(
            progress,
            57,
            f"PROD extraction complete: {len(stock_rows):,} items, "
            f"{len(shipment_rows):,} shipment rows, {len(history_rows):,} actual rows",
        )
        return stock_rows, shipment_rows, history_rows, actual_dates

    def _extract_oven_resource_structure(self, workbook, roles) -> list[dict[str, Any]]:
        """Capture the physical line/cavity-position skeleton losslessly.

        The OVEN workbook repeats each physical position across several allocation
        rows, including blank rows. Positive SAP rows alone therefore undercount
        the factory resource registry. This extractor keeps one record per exact
        (line, oven/cavity) identity plus the number of allocation slots printed
        for that position.
        """
        sheet_name = roles.get("OVEN_CAVITY_PLAN")
        if not sheet_name:
            return []
        ws = workbook[sheet_name]
        seen: dict[tuple[str, str], dict[str, Any]] = {}
        for row_number, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            line_name = _text(_value_at(row, 2))
            cavity_code = _text(_value_at(row, 3))
            if not line_name or not cavity_code:
                continue
            key = (line_name, cavity_code)
            rec = seen.setdefault(
                key,
                {
                    "line_name": line_name,
                    "cavity_code": cavity_code,
                    "first_source_row": row_number,
                    "last_source_row": row_number,
                    "allocation_slot_capacity": 0,
                    "source_sheet": sheet_name,
                },
            )
            rec["last_source_row"] = row_number
            rec["allocation_slot_capacity"] += 1
        return list(seen.values())

    def _extract_oven(self, workbook, roles, plan_date, issues) -> list[dict[str, Any]]:
        sheet_name = roles.get("OVEN_CAVITY_PLAN")
        if not sheet_name:
            issues.append(
                ImportIssue(
                    "BLOCKER",
                    "MISSING_OVEN_SHEET",
                    "",
                    "",
                    "",
                    "Oven/cavity plan sheet was not detected.",
                    "Map the sheet containing line, oven, tyre code and shift quantities.",
                )
            )
            return []
        ws = workbook[sheet_name]
        result: list[dict[str, Any]] = []
        for row_number, row in enumerate(
            ws.iter_rows(min_row=3, values_only=True),
            start=3,
        ):
            code = _code(_value_at(row, 4))
            if not code:
                continue
            description = _text(_value_at(row, 5))
            line_name = _text(_value_at(row, 2))
            oven_code = _text(_value_at(row, 3))
            total_to_produce_qty = _integer(_value_at(row, 10))
            today_qty = _integer(_value_at(row, 11))
            day_qty = _integer(_value_at(row, 12))
            night_qty = _integer(_value_at(row, 13))
            next_qty = _integer(_value_at(row, 15))
            unit_weight = _number(_value_at(row, 17)) or 0.0
            balance_qty = _integer(_value_at(row, 21))
            casing_evidence = _text(_value_at(row, 22))
            # Factory rule: BAND!A Material Description is the Mold Code.
            # OVEN column AA carries that BAND material description for the
            # planned SAP/resource row. Keep it as a separate identity from the
            # older tyre-size/key-code field.
            mold_code = _text(_value_at(row, 27))
            if mold_code.startswith("#"):
                mold_code = ""

            if not oven_code:
                issues.append(
                    ImportIssue(
                        "WARNING",
                        "MISSING_OVEN_CODE",
                        sheet_name,
                        f"C{row_number}",
                        code,
                        "Planned item has no oven/cavity identifier.",
                        "The row is retained in preview but not written to the live oven plan.",
                    )
                )
                continue

            for shift, qty, offset in (
                ("DAY", day_qty, 0),
                ("NIGHT", night_qty, 0),
                ("NEXT DAY", next_qty, 1),
            ):
                if qty <= 0:
                    continue
                row_date = (plan_date or date.today()) + timedelta(days=offset)
                result.append(
                    {
                        "plan_date": row_date.isoformat(),
                        "line_name": line_name,
                        "oven_code": oven_code,
                        "shift_name": shift,
                        "sap_code": code,
                        "description": description,
                        "planned_qty": qty,
                        # Row-level resource evidence retained for V11 capacity
                        # learning. TODAY/total/balance are not treated as verified
                        # actuals; verified actual production still comes from PROD.
                        "today_qty": today_qty,
                        "total_to_produce_qty": total_to_produce_qty,
                        "next_day_qty": next_qty,
                        "balance_qty": balance_qty,
                        "planned_weight_kg": round(qty * unit_weight, 5),
                        "unit_weight_kg": unit_weight,
                        "casing_evidence": casing_evidence,
                        "mold_code": mold_code,
                        "source_sheet": sheet_name,
                        "source_row": row_number,
                    }
                )
        return result

    def _extract_compound(
        self,
        workbook,
        roles,
        issues,
        *,
        progress: ProgressCallback | None = None,
    ) -> list[dict[str, Any]]:
        """Extract the per-tyre compound master using sequential streaming reads."""
        sheet_name = roles.get("COMPOUND_BOM")
        if not sheet_name:
            return []
        ws = workbook[sheet_name]

        max_material_column = min(column_index_from_string("CV"), ws.max_column)
        header_scan_last = min(ws.max_row, 12)
        self._progress(progress, 69, "Reading compound master headers")
        header_rows = list(
            ws.iter_rows(
                min_row=1,
                max_row=header_scan_last,
                min_col=1,
                max_col=max_material_column,
                values_only=True,
            )
        )
        header_row = None
        header_values: tuple[Any, ...] | None = None
        for row_number, values in enumerate(header_rows, start=1):
            material_description = _normalize(_value_at(values, 2))
            candidate_headers = [
                _text(_value_at(values, column))
                for column in range(4, max_material_column + 1)
            ]
            compound_header_count = sum(
                bool(value)
                and (
                    "compound" in _normalize(value)
                    or "friction cord" in _normalize(value)
                )
                for value in candidate_headers
            )
            if material_description == "material description" and compound_header_count >= 3:
                header_row = row_number
                header_values = values
                break

        if header_row is None or header_values is None:
            issues.append(
                ImportIssue(
                    "WARNING",
                    "COMPOUND_HEADERS_NOT_FOUND",
                    sheet_name,
                    "",
                    "",
                    "Compound sheet was detected but the per-tyre master header was not recognized.",
                    "The exact workbook is archived; review its header layout before promoting compound data.",
                )
            )
            return []

        ignored_headers = {
            "material",
            "material description",
            "altbom",
            "total to be produced",
            "day",
            "night",
            "key 01",
            "key 02",
            "total prodution",
            "total production",
            "inner core",
            "compound weight",
            "bead wire weight",
            "total tyre weight",
            "total tire weight",
            "band",
            "bead weight",
        }
        headers: list[tuple[int, str]] = []
        for column in range(column_index_from_string("D"), max_material_column + 1):
            header = _text(_value_at(header_values, column))
            normalized = _normalize(header)
            if not header or normalized in ignored_headers:
                continue
            headers.append((column, header))

        if not headers:
            issues.append(
                ImportIssue(
                    "WARNING",
                    "COMPOUND_HEADERS_NOT_FOUND",
                    sheet_name,
                    f"A{header_row}",
                    "",
                    "Compound master header row was found but no material columns were usable.",
                    "Review the workbook header names before import.",
                )
            )
            return []

        result: list[dict[str, Any]] = []
        max_scan_row = min(ws.max_row, 7000)
        total_rows = max(1, max_scan_row - header_row)
        last_progress = -1
        self._progress(progress, 70, "Scanning per-tyre compound usage")
        for offset, values in enumerate(
            ws.iter_rows(
                min_row=header_row + 1,
                max_row=max_scan_row,
                min_col=1,
                max_col=max_material_column,
                values_only=True,
            ),
            start=0,
        ):
            row_number = header_row + 1 + offset
            local_progress = 70 + min(5, int((offset / total_rows) * 6))
            if local_progress != last_progress:
                self._progress(
                    progress,
                    local_progress,
                    f"Scanning compound rows {offset + 1:,}/{total_rows:,}",
                )
                last_progress = local_progress

            first = _normalize(_value_at(values, 1))
            day_label = _normalize(_value_at(values, 4))
            night_label = _normalize(_value_at(values, 5))
            if first == "material" and day_label == "day" and night_label == "night":
                break

            code = _code(_value_at(values, 1))
            if not code or not _looks_like_item_code(code):
                continue
            for column, compound_name in headers:
                usage = _number(_value_at(values, column))
                if usage is None or usage <= 0:
                    continue
                compound_code = "CMP-" + sha256(
                    _normalize(compound_name).encode("utf-8")
                ).hexdigest()[:12].upper()
                result.append(
                    {
                        "sap_code": code,
                        "compound_code": compound_code,
                        "compound_name": compound_name,
                        "usage_per_unit": usage,
                        "stage": "MAIN",
                        "source_sheet": sheet_name,
                        "source_row": row_number,
                    }
                )
        self._progress(
            progress,
            75,
            f"Compound extraction complete: {len(result):,} BOM usage rows",
        )
        return result

    def _extract_bead(self, workbook, roles, issues) -> list[dict[str, Any]]:
        sheet_name = roles.get("BEAD_REQUIREMENT")
        if not sheet_name:
            return []
        ws = workbook[sheet_name]
        result: list[dict[str, Any]] = []
        for row_number, row in enumerate(
            ws.iter_rows(min_row=5, values_only=True),
            start=5,
        ):
            item_code = _text(_value_at(row, 1))
            bead_per_tyre = _number(_value_at(row, 2))
            bead_type = _text(_value_at(row, 5))
            if not item_code or not bead_type or bead_per_tyre is None:
                continue
            result.append(
                {
                    "item_code": item_code,
                    "bead_type": bead_type,
                    "bead_per_tyre": bead_per_tyre,
                    "planned_tyre_qty": _integer(_value_at(row, 3)),
                    "total_requirement": _number(_value_at(row, 4)) or 0.0,
                    "source_sheet": sheet_name,
                    "source_row": row_number,
                }
            )
        return result

    def _extract_band_master(self, workbook, roles) -> list[dict[str, Any]]:
        """Read every unique BAND Material Description as a learned Mold Code.

        Unlike the daily material-plan extractor, zero-plan/STOP rows are retained
        because they are still valid mold identities that may reappear later.
        """
        sheet_name = roles.get("BAND_PLAN")
        if not sheet_name:
            return []
        ws = workbook[sheet_name]
        result: list[dict[str, Any]] = []
        seen: set[str] = set()
        for row_number, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
            description = _text(_value_at(row, 1))
            if not description:
                continue
            key = _normalize(description)
            if not key or key in seen:
                continue
            seen.add(key)
            result.append({
                "mold_code": description,
                "day_qty": _number(_value_at(row, 2)) or 0.0,
                "night_qty": _number(_value_at(row, 3)) or 0.0,
                "total_qty": _number(_value_at(row, 4)) or 0.0,
                "stop_flag": _text(_value_at(row, 5)),
                "stock_qty": _number(_value_at(row, 6)) or 0.0,
                "produced_qty": _number(_value_at(row, 7)) or 0.0,
                "next_day_qty": _number(_value_at(row, 8)) or 0.0,
                "source_sheet": sheet_name,
                "source_row": row_number,
            })
        return result

    def _extract_material_plans(
        self,
        workbook,
        roles,
        plan_date,
        issues,
        *,
        progress: ProgressCallback | None = None,
    ) -> list[dict[str, Any]]:
        result: list[dict[str, Any]] = []
        default_date = plan_date or date.today()

        # The OVEN workbook contains a compact daily compound summary beside the
        # second calculated material block. Read DD:DI once sequentially; repeated
        # random ws.cell() access is very slow in openpyxl read-only mode.
        compound_sheet = roles.get("COMPOUND_BOM")
        if compound_sheet:
            ws = workbook[compound_sheet]
            compound_type_col = column_index_from_string("DD")
            balance_col = column_index_from_string("DI")
            max_scan_row = min(ws.max_row, 7000)
            self._progress(progress, 83, "Scanning compound daily requirement summary")
            header_found = False
            started = False
            blank_streak = 0
            for row_number, values in enumerate(
                ws.iter_rows(
                    min_row=1,
                    max_row=max_scan_row,
                    min_col=compound_type_col,
                    max_col=balance_col,
                    values_only=True,
                ),
                start=1,
            ):
                material_key = _text(_value_at(values, 1))
                requirement_text = _text(_value_at(values, 2))
                if not header_found:
                    if (
                        _normalize(material_key) == "compound type"
                        and "requirement" in _normalize(requirement_text)
                    ):
                        header_found = True
                    continue

                requirement = _number(_value_at(values, 2))
                if not material_key or requirement is None:
                    if started:
                        blank_streak += 1
                        if blank_streak >= 12:
                            break
                    continue
                if requirement < 0:
                    continue
                started = True
                blank_streak = 0
                first_stage = _number(_value_at(values, 3)) or 0.0
                batch = _number(_value_at(values, 4)) or 0.0
                stock = _number(_value_at(values, 5)) or 0.0
                balance = _number(_value_at(values, 6))
                description_bits = []
                if first_stage:
                    description_bits.append(f"1st stage eq. {first_stage:.3f} kg")
                if batch:
                    description_bits.append(f"Batch {batch:g}")
                if balance is not None:
                    description_bits.append(f"Excel balance {balance:.3f}")
                result.append(
                    {
                        "plan_date": default_date.isoformat(),
                        "material_type": "COMPOUND",
                        "material_key": material_key,
                        "material_description": "; ".join(description_bits),
                        "day_qty": 0.0,
                        "night_qty": 0.0,
                        "total_qty": requirement,
                        "produced_qty": 0.0,
                        "stock_qty": stock,
                        "next_day_qty": 0.0,
                        "unit": "KG",
                        "source_sheet": compound_sheet,
                        "source_row": row_number,
                    }
                )
            self._progress(progress, 89, "Compound daily summary mapped")

        band_sheet = roles.get("BAND_PLAN")
        if band_sheet:
            ws = workbook[band_sheet]
            for row_number, row in enumerate(
                ws.iter_rows(min_row=4, values_only=True),
                start=4,
            ):
                description = _text(_value_at(row, 1))
                if not description:
                    continue
                values = {
                    "day_qty": _number(_value_at(row, 2)) or 0.0,
                    "night_qty": _number(_value_at(row, 3)) or 0.0,
                    "total_qty": _number(_value_at(row, 4)) or 0.0,
                    "stock_qty": _number(_value_at(row, 6)) or 0.0,
                    "produced_qty": _number(_value_at(row, 7)) or 0.0,
                    "next_day_qty": _number(_value_at(row, 8)) or 0.0,
                }
                if not any(values.values()) and _normalize(_value_at(row, 5)) in {"", "stop"}:
                    continue
                result.append(
                    {
                        "plan_date": default_date.isoformat(),
                        "material_type": "BAND",
                        "material_key": description,
                        "material_description": description,
                        "unit": "PCS",
                        "source_sheet": band_sheet,
                        "source_row": row_number,
                        **values,
                    }
                )

        core_sheet = roles.get("CORE_PLAN")
        if core_sheet:
            ws = workbook[core_sheet]
            for row_number, row in enumerate(
                ws.iter_rows(min_row=6, values_only=True),
                start=6,
            ):
                material_key = _text(_value_at(row, 1)) or _text(_value_at(row, 2))
                if not material_key:
                    continue
                values = {
                    "total_qty": _number(_value_at(row, 2)) or 0.0,
                    "day_qty": _number(_value_at(row, 3)) or 0.0,
                    "night_qty": _number(_value_at(row, 6)) or 0.0,
                    "stock_qty": _number(_value_at(row, 9)) or 0.0,
                    "produced_qty": 0.0,
                    "next_day_qty": 0.0,
                }
                if not any(values.values()):
                    continue
                result.append(
                    {
                        "plan_date": default_date.isoformat(),
                        "material_type": "CORE",
                        "material_key": material_key,
                        "material_description": _text(_value_at(row, 10)),
                        "unit": "PCS",
                        "source_sheet": core_sheet,
                        "source_row": row_number,
                        **values,
                    }
                )

        bead_sheet = roles.get("BEAD_REQUIREMENT")
        if bead_sheet:
            for bead in self._extract_bead(workbook, roles, issues=[]):
                if bead["total_requirement"] <= 0:
                    continue
                result.append(
                    {
                        "plan_date": default_date.isoformat(),
                        "material_type": "BEAD",
                        "material_key": bead["bead_type"],
                        "material_description": bead["item_code"],
                        "day_qty": 0.0,
                        "night_qty": 0.0,
                        "total_qty": bead["total_requirement"],
                        "produced_qty": 0.0,
                        "stock_qty": 0.0,
                        "next_day_qty": 0.0,
                        "unit": "PCS",
                        "source_sheet": bead_sheet,
                        "source_row": bead["source_row"],
                    }
                )
        return result

    def _add_cross_validation_issues(
        self,
        stock_rows,
        shipment_rows,
        oven_rows,
        weights,
        issues,
    ) -> None:
        stock_codes = {row["sap_code"] for row in stock_rows}
        oven_codes = {row["sap_code"] for row in oven_rows}
        shipment_codes = {row["sap_code"] for row in shipment_rows}
        for code in sorted(oven_codes - stock_codes):
            issues.append(
                ImportIssue(
                    "WARNING",
                    "OVEN_ITEM_NOT_IN_PROD",
                    "OVEN",
                    "",
                    code,
                    "Oven plan item was not found in the production/stock item list.",
                    "Retained in the oven plan with its source description; review master data.",
                )
            )
        for code in sorted((shipment_codes | oven_codes) - set(weights)):
            issues.append(
                ImportIssue(
                    "INFO",
                    "MISSING_WEIGHT",
                    "WGT",
                    "",
                    code,
                    "No positive weight was found for an active demand/plan item.",
                    "Add or approve the item weight before weight-based reporting.",
                )
            )

    def _workbook_confidence(self, profiles: list[SheetProfile]) -> float:
        essential = {
            "PRODUCTION_STOCK_SHIPMENTS": 0.28,
            "OVEN_CAVITY_PLAN": 0.24,
            "WEIGHT_MASTER": 0.18,
            "DAILY_PRODUCTION_PLAN": 0.10,
            "COMPOUND_BOM": 0.08,
            "BEAD_REQUIREMENT": 0.05,
            "BAND_PLAN": 0.04,
            "CORE_PLAN": 0.03,
        }
        best = defaultdict(float)
        for profile in profiles:
            best[profile.role] = max(best[profile.role], profile.confidence)
        return round(
            min(1.0, sum(best[role] * weight for role, weight in essential.items())),
            5,
        )

    def _build_summary(
        self,
        *,
        profiles,
        stock_rows,
        shipment_rows,
        oven_rows,
        compound_rows,
        bead_rows,
        material_rows,
        history_rows,
        issues,
    ) -> dict[str, Any]:
        shipments = defaultdict(int)
        for row in shipment_rows:
            shipments[row["shipment_column"]] += row["quantity"]
        day_plan = sum(row["planned_qty"] for row in oven_rows if row["shift_name"] == "DAY")
        night_plan = sum(row["planned_qty"] for row in oven_rows if row["shift_name"] == "NIGHT")
        next_plan = sum(row["planned_qty"] for row in oven_rows if row["shift_name"] == "NEXT DAY")
        return {
            "sheet_count": len(profiles),
            "mapped_sheet_count": sum(p.role != "UNMAPPED_SOURCE_DATA" for p in profiles),
            "nonempty_cell_count": sum(p.nonempty_cells for p in profiles),
            "formula_cell_count": sum(p.formula_cells for p in profiles),
            "cached_error_cell_count": sum(p.cached_error_cells for p in profiles),
            "stock_item_count": len(stock_rows),
            "total_fg_stock": sum(row["fg_stock"] for row in stock_rows),
            "negative_stock_row_count": sum(
                any(
                    int(row.get(field_name) or 0) < 0
                    for field_name in (
                        "fg_stock",
                        "qc_stock",
                        "scrap_stock",
                        "blocked_stock",
                    )
                )
                for row in stock_rows
            ),
            "total_scrap_stock": sum(row["scrap_stock"] for row in stock_rows),
            "total_blocked_stock": sum(row["blocked_stock"] for row in stock_rows),
            "shipment_count": len(shipments),
            "shipment_item_count": len(shipment_rows),
            "total_shipment_qty": sum(row["quantity"] for row in shipment_rows),
            "production_required_qty": sum(row["production_required"] for row in stock_rows),
            "oven_plan_rows": len(oven_rows),
            "day_plan_qty": day_plan,
            "night_plan_qty": night_plan,
            "next_day_plan_qty": next_plan,
            "compound_bom_rows": len(compound_rows),
            "bead_master_rows": len(bead_rows),
            "material_plan_rows": len(material_rows),
            "production_history_rows": len(history_rows),
            "blocker_count": sum(i.severity == "BLOCKER" for i in issues),
            "warning_count": sum(i.severity == "WARNING" for i in issues),
            "info_count": sum(i.severity == "INFO" for i in issues),
        }

    def _archive_workbook(self, path: Path, digest: str) -> str:
        # R7.3 rolling-NVMe mode keeps the authoritative historical source on D:\n        # and uses C: only as a disposable processing cache.  The updater verifies
        # the relative file map before training, so the import can register that
        # already-preserved D: source without copying every workbook back to D:.
        fast_root = str(os.environ.get("MPPS_HISTORICAL_INBOX") or "").strip()
        archive_root = str(os.environ.get("MPPS_HISTORICAL_ARCHIVE_ROOT") or "").strip()
        if fast_root and archive_root:
            try:
                relative = path.resolve().relative_to(Path(fast_root).resolve())
                return str(Path(archive_root) / relative)
            except (OSError, ValueError):
                pass

        archive_dir = self.project_root / "data_sources" / "import_archive"
        archive_dir.mkdir(parents=True, exist_ok=True)
        safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", path.name).strip("_")
        destination = archive_dir / f"{digest[:16]}_{safe_name}"
        if not destination.exists():
            shutil.copy2(path, destination)
        try:
            return str(destination.relative_to(self.project_root))
        except Exception:
            return str(destination)

    def _save_profiles_and_issues(self, session, run_id, analysis) -> None:
        profile_sql = text(
            """
            INSERT INTO excel_import_sheet_profiles (
                run_id, sheet_name, detected_role, confidence_score,
                max_row, max_column, nonempty_cells, formula_cells,
                cached_error_cells, header_row, evidence
            ) VALUES (
                :run_id, :sheet_name, :role, :confidence,
                :max_row, :max_column, :nonempty, :formula,
                :errors, :header_row, :evidence
            )
            """
        )
        profile_params = [
            {
                "run_id": run_id,
                "sheet_name": profile.sheet_name,
                "role": profile.role,
                "confidence": profile.confidence,
                "max_row": profile.max_row,
                "max_column": profile.max_column,
                "nonempty": profile.nonempty_cells,
                "formula": profile.formula_cells,
                "errors": profile.cached_error_cells,
                "header_row": profile.header_row,
                "evidence": profile.evidence,
            }
            for profile in analysis.sheet_profiles
        ]
        if profile_params:
            session.execute(profile_sql, profile_params)

        issue_sql = text(
            """
            INSERT INTO excel_import_issues (
                run_id, severity, category, sheet_name, cell_address,
                item_key, message, recommendation
            ) VALUES (
                :run_id, :severity, :category, :sheet_name, :cell_address,
                :item_key, :message, :recommendation
            )
            """
        )
        issue_params = [{"run_id": run_id, **asdict(issue)} for issue in analysis.issues]
        if issue_params:
            session.execute(issue_sql, issue_params)

    def _save_workbook_registry(
        self,
        session,
        run_id,
        analysis,
        archive_path,
        imported_by,
    ) -> None:
        session.execute(
            text(
                """
                INSERT INTO excel_workbooks (
                    workbook_key, original_file_name, file_path, file_hash,
                    imported_by, remarks
                ) VALUES (
                    :key, :name, :path, :hash, :imported_by, :remarks
                )
                ON CONFLICT (workbook_key)
                DO UPDATE SET
                    original_file_name = EXCLUDED.original_file_name,
                    file_path = EXCLUDED.file_path,
                    file_hash = EXCLUDED.file_hash,
                    imported_at = CURRENT_TIMESTAMP,
                    imported_by = EXCLUDED.imported_by,
                    remarks = EXCLUDED.remarks
                """
            ),
            {
                "key": f"INTELLIGENT-{analysis.workbook_hash}",
                "name": analysis.workbook_name,
                "path": archive_path,
                "hash": analysis.workbook_hash,
                "imported_by": imported_by or "Local User",
                "remarks": f"Intelligent Excel Import run #{run_id}",
            },
        )

    def _commit_stock(
        self,
        session,
        run_id,
        analysis,
        plan_date,
        options,
        counters,
    ) -> None:
        for row in analysis.stock_rows:
            if not _looks_like_item_code(row["sap_code"]):
                counters["stock_nonstandard_skipped"] += 1
                continue
            code = row["sap_code"]
            description = row["description"]

            source_stock = {
                "fg_stock": int(row.get("fg_stock") or 0),
                "qc_stock": int(row.get("qc_stock") or 0),
                "scrap_stock": int(row.get("scrap_stock") or 0),
                "blocked_stock": int(row.get("blocked_stock") or 0),
            }
            live_stock = {
                field_name: max(0, value)
                for field_name, value in source_stock.items()
            }
            negative_stock = {
                field_name: value
                for field_name, value in source_stock.items()
                if value < 0
            }
            negative_note = ""
            if negative_stock:
                counters["negative_stock_rows_normalized"] += 1
                negative_note = (
                    "; negative source stock normalized to zero: "
                    + ", ".join(
                        f"{field_name}={value}"
                        for field_name, value in negative_stock.items()
                    )
                )

            if options["update_stock"]:
                self._upsert_with_change(
                    session,
                    run_id,
                    "mpps_sap_stock_items",
                    {"sap_code": code},
                    {
                        "sap_code": code,
                        "tyre_description": description,
                        "item_description": description,
                        "fg_stock": live_stock["fg_stock"],
                        "qc_stock": live_stock["qc_stock"],
                        "scrap_stock": live_stock["scrap_stock"],
                        "blocked_stock": live_stock["blocked_stock"],
                        "is_active": True,
                        "source_table": "INTELLIGENT_EXCEL_IMPORT",
                        "source_note": (
                            f"Run #{run_id}; {analysis.workbook_name}; "
                            f"{row['source_sheet']} row {row['source_row']}"
                            f"{negative_note}"
                        ),
                        "updated_at": datetime.now(),
                    },
                )
                stock_item = {
                    "material_code": code,
                    "item_description": description,
                    "fg_stock": live_stock["fg_stock"],
                    "qc_stock": live_stock["qc_stock"],
                    "scrap_stock": live_stock["scrap_stock"],
                    "blocked_stock": live_stock["blocked_stock"],
                    "is_active": True,
                    "last_updated_date": plan_date,
                    "source_workbook": analysis.workbook_name,
                    "source_sheet": row["source_sheet"],
                    "source_row": row["source_row"],
                    "source_note": (
                        f"Intelligent import run #{run_id}{negative_note}"
                    ),
                    "updated_at": datetime.now(),
                }
                weight = row.get("weight_kg")
                if weight and (
                    options["overwrite_existing_weights"]
                    or options["update_blank_weights"]
                ):
                    current = self._fetch_existing(
                        session,
                        "mpps_stock_items",
                        {"material_code": code},
                    )
                    current_weight = _number(
                        current.get("average_weight") if current else None
                    )
                    if options["overwrite_existing_weights"] or not current_weight:
                        stock_item["average_weight"] = weight
                self._upsert_with_change(
                    session,
                    run_id,
                    "mpps_stock_items",
                    {"material_code": code},
                    stock_item,
                )
                counters["stock_items_updated"] += 1

                if row.get("weight_kg") and options["update_blank_weights"]:
                    existing_smds = self._fetch_existing(
                        session,
                        "smds",
                        {"sap_code": code},
                    )
                    if existing_smds:
                        current_weight = _number(existing_smds.get("weight_per_tyre_kg"))
                        if options["overwrite_existing_weights"] or not current_weight:
                            self._upsert_with_change(
                                session,
                                run_id,
                                "smds",
                                {"sap_code": code},
                                {
                                    "sap_code": code,
                                    "material_description": (
                                        existing_smds.get("material_description")
                                        or description
                                    ),
                                    "weight_per_tyre_kg": row["weight_kg"],
                                    "source_file": analysis.workbook_name,
                                    "source_sheet": "WGT",
                                    "source_row_number": row["source_row"],
                                    "updated_at": datetime.now(),
                                },
                            )
                            counters["blank_smds_weights_filled"] += 1

            if options["update_daily_stock"]:
                production_qty = sum(
                    h["production_qty"]
                    for h in analysis.production_history_rows
                    if h["sap_code"] == code
                    and h["production_date"] == plan_date.isoformat()
                )
                self._upsert_with_change(
                    session,
                    run_id,
                    "mpps_daily_stock_entries",
                    {"stock_date": plan_date, "sap_code": code},
                    {
                        "stock_date": plan_date,
                        "sap_code": code,
                        "tyre_description": description,
                        "production_qty": production_qty,
                        "fg_qty": live_stock["fg_stock"],
                        "qc_qty": live_stock["qc_stock"],
                        "scrap_qty": live_stock["scrap_stock"],
                        "blocked_qty": live_stock["blocked_stock"],
                        "note": (
                            f"Intelligent import run #{run_id}{negative_note}"
                        ),
                        "source_file": analysis.workbook_name,
                        "updated_at": datetime.now(),
                    },
                )
                counters["daily_stock_rows_updated"] += 1

    def _commit_shipment_snapshots(
        self,
        session,
        run_id,
        analysis,
        plan_date,
        options,
        counters,
    ) -> dict[str, Any]:
        grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
        params = []
        for row in analysis.shipment_rows:
            grouped[row["shipment_column"]].append(row)
            source_target_date = row.get("source_target_date")
            if isinstance(source_target_date, str) and source_target_date:
                source_target_date = date.fromisoformat(source_target_date)
            params.append(
                {
                    "run_id": run_id,
                    "plan_date": plan_date,
                    **row,
                    "source_target_date": source_target_date,
                }
            )

        if params:
            session.execute(
                text(
                    """
                    INSERT INTO excel_import_shipment_snapshots (
                        run_id, shipment_column, shipment_name, source_status,
                        source_target_date, source_date_class, source_item_code,
                        item_description, quantity, plan_date
                    ) VALUES (
                        :run_id, :shipment_column, :shipment_name, :source_status,
                        :source_target_date, :source_date_class, :sap_code,
                        :description, :quantity, :plan_date
                    )
                    ON CONFLICT (run_id, shipment_column, source_item_code)
                    DO UPDATE SET
                        shipment_name = EXCLUDED.shipment_name,
                        source_status = EXCLUDED.source_status,
                        source_target_date = EXCLUDED.source_target_date,
                        source_date_class = EXCLUDED.source_date_class,
                        item_description = EXCLUDED.item_description,
                        quantity = EXCLUDED.quantity,
                        plan_date = EXCLUDED.plan_date
                    """
                ),
                params,
            )
            counters["shipment_snapshot_items"] += len(params)

        counters["shipment_snapshots"] = len(grouped)
        if not options.get("sync_live_shipments", False):
            return {
                "sync_mode": options.get("resolved_import_mode", "SNAPSHOT_ONLY"),
                "sync_reason": options.get(
                    "resolved_import_reason",
                    "Shipment data was retained as a dated snapshot only.",
                ),
                "live_change_count": 0,
                "shipment_snapshots": len(grouped),
            }

        sync_service = WorkbookContinuousSyncService(self.project_root)
        return sync_service.sync(
            session,
            import_run_id=run_id,
            analysis=analysis,
            options=options,
            counters=counters,
            ledger=self,
        )

    def _commit_oven_plan(self, session, run_id, analysis, counters) -> None:
        for row in analysis.oven_rows:
            key = {
                "plan_date": date.fromisoformat(row["plan_date"]),
                "oven_code": row["oven_code"],
                "shift_name": row["shift_name"],
                "material_code": row["sap_code"],
                "source_workbook": analysis.workbook_name,
                "source_row": row["source_row"],
            }
            values = {
                **key,
                "item_description": row["description"],
                "planned_qty": row["planned_qty"],
                "planned_weight_kg": row["planned_weight_kg"],
                "plan_status": "IMPORTED",
                "source_sheet": row["source_sheet"],
                "source_note": (
                    f"Intelligent import run #{run_id}; line={row['line_name']}; "
                    f"casing evidence retained={row['casing_evidence'] or '-'}"
                ),
                "line_name": row.get("line_name") or "",
                "cavity_code": row.get("oven_code") or "",
                "allocation_slot": 1,
                "mold_code": row.get("mold_code") or "",
                "updated_at": datetime.now(),
            }
            self._upsert_with_change(session, run_id, "mpps_oven_plan", key, values)
            counters["oven_plan_rows_updated"] += 1

    def _commit_materials(self, session, run_id, analysis, counters) -> None:
        for row in analysis.compound_rows:
            master_key = {
                "item_code": row["sap_code"],
                "compound_code": row["compound_code"],
            }
            self._upsert_with_change(
                session,
                run_id,
                "mpps_compound_master",
                master_key,
                {
                    **master_key,
                    "compound_name": row["compound_name"],
                    "compound_weight_per_unit": row["usage_per_unit"],
                    "stage": row["stage"],
                    "is_active": True,
                    "source_workbook": analysis.workbook_name,
                    "source_sheet": row["source_sheet"],
                    "source_row": row["source_row"],
                    "source_note": f"Intelligent import run #{run_id}",
                    "updated_at": datetime.now(),
                },
            )
            bom_key = {
                "finished_item_code": row["sap_code"],
                "raw_material_code": row["compound_code"],
            }
            self._upsert_with_change(
                session,
                run_id,
                "mpps_bom_items",
                bom_key,
                {
                    **bom_key,
                    "raw_material_name": row["compound_name"],
                    "usage_per_unit": row["usage_per_unit"],
                    "unit": "KG",
                    "wastage_percentage": 0,
                    "is_active": True,
                    "source_workbook": analysis.workbook_name,
                    "source_sheet": row["source_sheet"],
                    "source_row": row["source_row"],
                    "source_note": f"Intelligent import run #{run_id}",
                    "updated_at": datetime.now(),
                },
            )
            counters["compound_bom_rows_updated"] += 1

        for row in analysis.bead_rows:
            key = {"item_code": row["item_code"], "bead_type": row["bead_type"]}
            self._upsert_with_change(
                session,
                run_id,
                "mpps_bead_master",
                key,
                {
                    **key,
                    "bead_per_tyre": row["bead_per_tyre"],
                    "is_active": True,
                    "source_workbook": analysis.workbook_name,
                    "source_sheet": row["source_sheet"],
                    "source_row": row["source_row"],
                    "source_note": f"Intelligent import run #{run_id}",
                    "updated_at": datetime.now(),
                },
            )
            counters["bead_master_rows_updated"] += 1

        for row in analysis.material_plan_rows:
            session.execute(
                text(
                    """
                    INSERT INTO excel_import_material_plans (
                        run_id, plan_date, material_type, material_key,
                        material_description, day_qty, night_qty, total_qty,
                        produced_qty, stock_qty, next_day_qty, unit,
                        source_sheet, source_row, source_json
                    ) VALUES (
                        :run_id, :plan_date, :material_type, :material_key,
                        :material_description, :day_qty, :night_qty, :total_qty,
                        :produced_qty, :stock_qty, :next_day_qty, :unit,
                        :source_sheet, :source_row, CAST(:source_json AS JSONB)
                    )
                    """
                ),
                {
                    "run_id": run_id,
                    **row,
                    "plan_date": date.fromisoformat(row["plan_date"]),
                    "source_json": json.dumps(row, default=str),
                },
            )
            counters["material_plan_rows_imported"] += 1

    def _commit_production_history(self, session, run_id, analysis, counters) -> None:
        params = [
            {
                "run_id": run_id,
                **row,
                "production_date": date.fromisoformat(row["production_date"]),
            }
            for row in analysis.production_history_rows
        ]
        if not params:
            return
        session.execute(
            text(
                """
                INSERT INTO excel_import_production_history (
                    run_id, production_date, sap_code, item_description,
                    production_qty, source_sheet, source_row
                ) VALUES (
                    :run_id, :production_date, :sap_code, :description,
                    :production_qty, :source_sheet, :source_row
                )
                ON CONFLICT (run_id, production_date, sap_code)
                DO UPDATE SET
                    item_description = EXCLUDED.item_description,
                    production_qty = EXCLUDED.production_qty,
                    source_sheet = EXCLUDED.source_sheet,
                    source_row = EXCLUDED.source_row
                """
            ),
            params,
        )
        counters["production_history_rows_imported"] += len(params)

    def _fetch_existing(self, session, table_name, key_fields):
        self._validate_table_and_columns(table_name, key_fields)
        where = " AND ".join(f"{column} = :key_{column}" for column in key_fields)
        params = {f"key_{column}": value for column, value in key_fields.items()}
        return session.execute(
            text(f"SELECT * FROM {table_name} WHERE {where} ORDER BY id LIMIT 1"),
            params,
        ).mappings().first()

    _NATIVE_UPSERT_KEYS = {
        "mpps_sap_stock_items": ("sap_code",),
        "mpps_stock_items": ("material_code",),
        "mpps_daily_stock_entries": ("stock_date", "sap_code"),
        # Shipment identities have a database UNIQUE(source_family, identity_key)
        # constraint.  Treat them as an atomic natural-key entity as well so
        # legacy/partial V7->V8 registries can self-heal without SELECT/INSERT
        # races or stale preview assumptions.
        "excel_shipment_identities": ("source_family", "identity_key"),
    }

    def _upsert_with_change(self, session, run_id, table_name, key_fields, values):
        conflict_keys = self._NATIVE_UPSERT_KEYS.get(table_name)
        if conflict_keys and set(conflict_keys) == set(key_fields):
            return self._native_upsert_with_change(
                session,
                run_id,
                table_name,
                key_fields,
                values,
                conflict_keys=conflict_keys,
            )

        existing = self._fetch_existing(session, table_name, key_fields)
        if existing:
            return self._update_existing_by_id(
                session,
                run_id,
                table_name,
                int(existing["id"]),
                values,
                key_fields=key_fields,
                existing=dict(existing),
            )
        return self._insert_with_change(
            session,
            run_id,
            table_name,
            values,
            key_fields=key_fields,
        )

    def _native_upsert_with_change(
        self,
        session,
        run_id,
        table_name,
        key_fields,
        values,
        *,
        conflict_keys,
    ):
        """Atomic PostgreSQL UPSERT for stable natural-key tables.

        The old SELECT-then-INSERT path could lose a race against another
        stock synchronizer (or a legacy trigger/workflow) and raise a unique
        violation even though the desired operation was simply to refresh the
        existing SAP row.  ON CONFLICT makes that operation atomic.
        """
        self._validate_table_and_columns(table_name, key_fields)
        existing = self._fetch_existing(session, table_name, key_fields)
        before = dict(existing) if existing else {}

        clean_values = self._clean_values(table_name, values)
        for key, value in key_fields.items():
            clean_values.setdefault(key, value)
        if not clean_values:
            return

        columns = list(clean_values)
        placeholders = [f":value_{column}" for column in columns]
        params = {f"value_{column}": clean_values[column] for column in columns}
        update_columns = [
            column for column in columns
            if column not in conflict_keys and column != "id"
        ]
        conflict_clause = ", ".join(conflict_keys)

        if update_columns:
            update_clause = ", ".join(
                f"{column} = EXCLUDED.{column}" for column in update_columns
            )
            sql = (
                f"INSERT INTO {table_name} ({', '.join(columns)}) "
                f"VALUES ({', '.join(placeholders)}) "
                f"ON CONFLICT ({conflict_clause}) DO UPDATE SET {update_clause} "
                "RETURNING *"
            )
            returned = session.execute(text(sql), params).mappings().first()
            if returned is None:
                returned = self._fetch_existing(session, table_name, key_fields)
            if returned is None:
                raise RuntimeError(
                    f"Import consistency error: PostgreSQL UPSERT for {table_name} "
                    f"returned no row for key {key_fields}. Re-analyze the workbook "
                    "and retry the import."
                )
            after = dict(returned)
        else:
            sql = (
                f"INSERT INTO {table_name} ({', '.join(columns)}) "
                f"VALUES ({', '.join(placeholders)}) "
                f"ON CONFLICT ({conflict_clause}) DO NOTHING "
                "RETURNING *"
            )
            returned = session.execute(text(sql), params).mappings().first()
            after = dict(returned) if returned else dict(
                self._fetch_existing(session, table_name, key_fields) or {}
            )

        if before:
            comparable_after = {key: after.get(key) for key in after}
            if _json_safe(before) == _json_safe(comparable_after):
                return after
            action = "UPDATE"
        else:
            action = "INSERT"

        self._record_change(
            session,
            run_id,
            table_name,
            action,
            key_fields,
            before,
            after,
        )
        return after

    def _update_existing_by_id(
        self,
        session,
        run_id,
        table_name,
        row_id,
        values,
        *,
        key_fields=None,
        existing=None,
    ):
        key_fields = key_fields or {"id": row_id}
        if existing is None:
            current = session.execute(
                text(f"SELECT * FROM {table_name} WHERE id = :id"),
                {"id": row_id},
            ).mappings().first()
            # A preview can legitimately become stale when a legacy repair,
            # trigger, or another sync path has re-keyed a natural-key row.
            # Recover by its stable key instead of raising NoResultFound.
            if current is None and key_fields and set(key_fields) != {"id"}:
                current = self._fetch_existing(session, table_name, key_fields)
                if current is not None and current.get("id") is not None:
                    row_id = int(current["id"])
            if current is None:
                raise RuntimeError(
                    f"Import consistency error: {table_name} row id={row_id} "
                    f"was not found for key {key_fields}. The transaction was "
                    "left unchanged; re-analyze and retry."
                )
            existing = dict(current)
        else:
            existing = dict(existing)
        clean_values = self._clean_values(table_name, values)
        if not clean_values:
            return existing
        changed_values = {
            key: value
            for key, value in clean_values.items()
            if _json_safe(existing.get(key)) != _json_safe(value)
        }
        if not changed_values:
            return existing
        set_clause = ", ".join(f"{key} = :value_{key}" for key in changed_values)
        params = {f"value_{key}": value for key, value in changed_values.items()}
        params["id"] = row_id
        session.execute(
            text(f"UPDATE {table_name} SET {set_clause} WHERE id = :id"),
            params,
        )
        after = {**existing, **changed_values}
        self._record_change(
            session,
            run_id,
            table_name,
            "UPDATE",
            key_fields,
            existing,
            after,
        )
        return after

    def _insert_with_change(
        self,
        session,
        run_id,
        table_name,
        values,
        *,
        key_fields,
    ):
        clean_values = self._clean_values(table_name, values)
        columns = list(clean_values)
        placeholders = [f":value_{column}" for column in columns]
        params = {f"value_{column}": clean_values[column] for column in columns}
        inserted = session.execute(
            text(
                f"INSERT INTO {table_name} ({', '.join(columns)}) "
                f"VALUES ({', '.join(placeholders)}) RETURNING *"
            ),
            params,
        ).mappings().first()
        if inserted is None:
            inserted = self._fetch_existing(session, table_name, key_fields)
        if inserted is None:
            raise RuntimeError(
                f"Import consistency error: INSERT into {table_name} returned "
                f"no row for key {key_fields}."
            )
        inserted_dict = dict(inserted)
        # Prefer the actual returned primary key for rollback. This remains
        # correct even when a legacy PostgreSQL trigger normalizes a natural
        # key such as an SAP code during INSERT.
        change_key = (
            {"id": int(inserted_dict["id"])}
            if inserted_dict.get("id") is not None
            else key_fields
        )
        self._record_change(
            session,
            run_id,
            table_name,
            "INSERT",
            change_key,
            {},
            inserted_dict,
        )
        return inserted_dict

    def _record_change(
        self,
        session,
        run_id,
        table_name,
        action,
        key_fields,
        before,
        after,
    ):
        session.execute(
            text(
                """
                INSERT INTO excel_import_changes (
                    run_id, table_name, action, key_json, before_json, after_json
                ) VALUES (
                    :run_id, :table_name, :action,
                    CAST(:key_json AS JSONB),
                    CAST(:before_json AS JSONB),
                    CAST(:after_json AS JSONB)
                )
                """
            ),
            {
                "run_id": run_id,
                "table_name": table_name,
                "action": action,
                "key_json": json.dumps(_json_safe(key_fields), default=str),
                "before_json": json.dumps(_json_safe(before), default=str),
                "after_json": json.dumps(_json_safe(after), default=str),
            },
        )

    def _delete_by_key(self, session, table_name, key_fields) -> int:
        self._validate_table_and_columns(table_name, key_fields)
        where = " AND ".join(f"{column} = :key_{column}" for column in key_fields)
        params = {f"key_{column}": value for column, value in key_fields.items()}
        result = session.execute(text(f"DELETE FROM {table_name} WHERE {where}"), params)
        return int(result.rowcount or 0)

    def _restore_by_key(self, session, table_name, key_fields, before_data) -> int:
        clean_before = self._clean_values(table_name, before_data)
        if not clean_before:
            return 0
        self._validate_table_and_columns(table_name, key_fields)
        set_clause = ", ".join(f"{column} = :value_{column}" for column in clean_before)
        where = " AND ".join(f"{column} = :key_{column}" for column in key_fields)
        params = {f"value_{column}": value for column, value in clean_before.items()}
        params.update({f"key_{column}": value for column, value in key_fields.items()})
        result = session.execute(
            text(f"UPDATE {table_name} SET {set_clause} WHERE {where}"),
            params,
        )
        return int(result.rowcount or 0)

    def _clean_values(self, table_name, values):
        allowed = LIVE_TABLE_COLUMNS.get(table_name)
        if allowed is None:
            raise RuntimeError(f"Table is not approved for intelligent import: {table_name}")
        return {key: value for key, value in values.items() if key in allowed}

    def _validate_table_and_columns(self, table_name, values):
        allowed = LIVE_TABLE_COLUMNS.get(table_name)
        if allowed is None:
            raise RuntimeError(f"Unsupported table: {table_name}")
        unknown = set(values) - (allowed | {"id"})
        if unknown:
            raise RuntimeError(f"Unsupported columns for {table_name}: {sorted(unknown)}")

    @staticmethod
    def _progress(callback, percent, message):
        if callback:
            callback(int(percent), str(message))


def _get_session():
    # Lazy import keeps workbook analysis available even in offline review tools.
    from app.database import get_session

    return get_session()


def _analysis_from_dict(payload: dict[str, Any]) -> WorkbookAnalysis:
    return WorkbookAnalysis(
        workbook_path=payload["workbook_path"],
        workbook_name=payload["workbook_name"],
        workbook_hash=payload["workbook_hash"],
        workbook_size_bytes=int(payload["workbook_size_bytes"]),
        plan_date=payload.get("plan_date"),
        confidence_score=float(payload["confidence_score"]),
        detected_type=payload.get("detected_type", ""),
        sheet_profiles=[SheetProfile(**row) for row in payload.get("sheet_profiles", [])],
        issues=[ImportIssue(**row) for row in payload.get("issues", [])],
        stock_rows=list(payload.get("stock_rows", [])),
        shipment_rows=list(payload.get("shipment_rows", [])),
        oven_rows=list(payload.get("oven_rows", [])),
        oven_resource_rows=list(payload.get("oven_resource_rows", [])),
        compound_rows=list(payload.get("compound_rows", [])),
        bead_rows=list(payload.get("bead_rows", [])),
        band_rows=list(payload.get("band_rows", [])),
        material_plan_rows=list(payload.get("material_plan_rows", [])),
        production_history_rows=list(payload.get("production_history_rows", [])),
        actual_production_dates=list(payload.get("actual_production_dates", [])),
        summary=dict(payload.get("summary", {})),
    )


def _sha256_file(path: Path) -> str:
    digest = sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _normalize(value: Any) -> str:
    text_value = _text(value).lower()
    text_value = re.sub(r"[^a-z0-9]+", " ", text_value)
    return re.sub(r"\s+", " ", text_value).strip()


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _code(value: Any) -> str:
    value = _text(value)
    if re.fullmatch(r"\d+\.0", value):
        value = value[:-2]
    return re.sub(r"\s+", " ", value).strip()


def _looks_like_item_code(value: str) -> bool:
    return bool(
        re.fullmatch(r"\d{6,}(?:\s+[A-Z]{1,5})?", value, flags=re.IGNORECASE)
        or re.fullmatch(r"[A-Z][A-Z0-9 -]{4,30}", value, flags=re.IGNORECASE)
    )


def _number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return float(int(value))
    if isinstance(value, (int, float, Decimal)):
        value = float(value)
        if math.isnan(value) or math.isinf(value):
            return None
        return value
    text_value = _text(value).replace(",", "")
    if text_value.startswith("#"):
        return None
    try:
        return float(Decimal(text_value))
    except (InvalidOperation, ValueError):
        return None


def _integer(value: Any) -> int:
    number = _number(value)
    if number is None:
        return 0
    return int(round(number))


def _as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    number = _number(value)
    if number is not None and 1 <= number <= 100000:
        try:
            return date(1899, 12, 30) + timedelta(days=int(number))
        except ValueError:
            pass
    text_value = _text(value)
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text_value, pattern).date()
        except ValueError:
            continue
    return None


def _value_at(row: tuple[Any, ...], one_based_column: int) -> Any:
    index = one_based_column - 1
    if index < 0 or index >= len(row):
        return None
    return row[index]


def _json_object(value: Any) -> dict[str, Any]:
    if value is None:
        return {}
    if isinstance(value, dict):
        return value
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
            return parsed if isinstance(parsed, dict) else {}
        except json.JSONDecodeError:
            return {}
    return dict(value)


def _json_safe(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_safe(item) for item in value]
    if isinstance(value, (date, datetime, Decimal)):
        return str(value)
    return value
