from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from PySide6.QtCore import QDate, Qt
from PySide6.QtGui import QColor
from PySide6.QtWidgets import (
    QAbstractItemView,
    QDateEdit,
    QDialog,
    QFileDialog,
    QFrame,
    QGridLayout,
    QHeaderView,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QSpinBox,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from sqlalchemy import text

from app.database import engine


class DailyStockEditDialog(QDialog):
    def __init__(self, parent=None, row_data: dict | None = None):
        super().__init__(parent)

        self.row_data = row_data or {}

        self.setWindowTitle("Edit Daily Stock Entry")
        self.setMinimumWidth(620)

        self.sap_code_label = QLabel(str(self.row_data.get("sap_code") or "-"))
        self.description_label = QLabel(str(self.row_data.get("tyre_description") or "-"))

        self.production_qty_input = QSpinBox()
        self.production_qty_input.setRange(0, 999999999)
        self.production_qty_input.setValue(int(self.row_data.get("production_qty") or 0))

        self.fg_qty_input = QSpinBox()
        self.fg_qty_input.setRange(0, 999999999)
        self.fg_qty_input.setValue(int(self.row_data.get("fg_qty") or 0))

        self.qc_qty_input = QSpinBox()
        self.qc_qty_input.setRange(0, 999999999)
        self.qc_qty_input.setValue(int(self.row_data.get("qc_qty") or 0))

        self.scrap_qty_input = QSpinBox()
        self.scrap_qty_input.setRange(0, 999999999)
        self.scrap_qty_input.setValue(int(self.row_data.get("scrap_qty") or 0))

        self.blocked_qty_input = QSpinBox()
        self.blocked_qty_input.setRange(0, 999999999)
        self.blocked_qty_input.setValue(int(self.row_data.get("blocked_qty") or 0))

        self.note_input = QLineEdit()
        self.note_input.setPlaceholderText("Optional note...")
        self.note_input.setText(str(self.row_data.get("note") or ""))

        self.save_btn = QPushButton("Save Entry")
        self.save_btn.setObjectName("PrimaryButton")
        self.save_btn.clicked.connect(self.accept)

        self.cancel_btn = QPushButton("Cancel")
        self.cancel_btn.setObjectName("SecondaryButton")
        self.cancel_btn.clicked.connect(self.reject)

        self._apply_styles()
        self._build_ui()

    def _apply_styles(self) -> None:
        self.setStyleSheet(
            """
            QDialog {
                background: #f8fafc;
                font-family: "Segoe UI";
            }

            QFrame#Card {
                background: #ffffff;
                border: 1px solid #e2e8f0;
                border-radius: 16px;
            }

            QLabel#Title {
                color: #0f172a;
                font-size: 16pt;
                font-weight: 950;
            }

            QLabel#Hint {
                color: #64748b;
                font-size: 9.5pt;
                font-weight: 650;
            }

            QLabel#FieldLabel {
                color: #334155;
                font-size: 9pt;
                font-weight: 850;
            }

            QLabel#ReadonlyValue {
                background: #f1f5f9;
                color: #0f172a;
                border: 1px solid #e2e8f0;
                border-radius: 10px;
                padding: 10px 12px;
                font-size: 10pt;
                font-weight: 800;
            }

            QLineEdit, QSpinBox {
                background: #ffffff;
                color: #0f172a;
                border: 1px solid #cbd5e1;
                border-radius: 10px;
                padding: 9px 12px;
                font-size: 10pt;
                font-weight: 650;
                min-height: 24px;
            }

            QLineEdit:focus, QSpinBox:focus {
                border: 1px solid #2563eb;
            }

            QPushButton#PrimaryButton {
                background: #2563eb;
                color: white;
                border: none;
                border-radius: 10px;
                padding: 10px 20px;
                font-weight: 950;
                min-height: 26px;
            }

            QPushButton#PrimaryButton:hover {
                background: #1d4ed8;
            }

            QPushButton#SecondaryButton {
                background: #e2e8f0;
                color: #0f172a;
                border: none;
                border-radius: 10px;
                padding: 10px 20px;
                font-weight: 950;
                min-height: 26px;
            }

            QPushButton#SecondaryButton:hover {
                background: #cbd5e1;
            }
            """
        )

    def _build_ui(self) -> None:
        root = QVBoxLayout(self)
        root.setContentsMargins(18, 18, 18, 18)

        card = QFrame()
        card.setObjectName("Card")

        layout = QVBoxLayout(card)
        layout.setContentsMargins(20, 18, 20, 20)
        layout.setSpacing(14)

        title = QLabel("Daily Production / Stock Entry")
        title.setObjectName("Title")

        hint = QLabel("Edit the selected daily stock line. Available daily qty is calculated from FG + QC - Scrap - Blocked.")
        hint.setObjectName("Hint")
        hint.setWordWrap(True)

        layout.addWidget(title)
        layout.addWidget(hint)

        form = QGridLayout()
        form.setHorizontalSpacing(14)
        form.setVerticalSpacing(12)

        self._add_readonly(form, 0, "SAP Code", self.sap_code_label)
        self._add_readonly(form, 1, "Tyre Description", self.description_label)
        self._add_field(form, 2, "Production Qty", self.production_qty_input)
        self._add_field(form, 3, "FG Qty", self.fg_qty_input)
        self._add_field(form, 4, "QC Qty", self.qc_qty_input)
        self._add_field(form, 5, "Scrap Qty", self.scrap_qty_input)
        self._add_field(form, 6, "Blocked Qty", self.blocked_qty_input)
        self._add_field(form, 7, "Note", self.note_input)

        layout.addLayout(form)

        buttons = QHBoxLayout()
        buttons.addStretch()
        buttons.addWidget(self.cancel_btn)
        buttons.addWidget(self.save_btn)

        layout.addLayout(buttons)
        root.addWidget(card)

    def _add_field(self, grid: QGridLayout, row: int, label_text: str, widget: QWidget) -> None:
        label = QLabel(label_text)
        label.setObjectName("FieldLabel")
        grid.addWidget(label, row, 0)
        grid.addWidget(widget, row, 1)
        grid.setColumnStretch(1, 1)

    def _add_readonly(self, grid: QGridLayout, row: int, label_text: str, widget: QLabel) -> None:
        label = QLabel(label_text)
        label.setObjectName("FieldLabel")
        widget.setObjectName("ReadonlyValue")
        widget.setWordWrap(True)
        grid.addWidget(label, row, 0)
        grid.addWidget(widget, row, 1)

    def get_data(self) -> dict:
        return {
            "sap_code": self.row_data.get("sap_code"),
            "production_qty": self.production_qty_input.value(),
            "fg_qty": self.fg_qty_input.value(),
            "qc_qty": self.qc_qty_input.value(),
            "scrap_qty": self.scrap_qty_input.value(),
            "blocked_qty": self.blocked_qty_input.value(),
            "note": self.note_input.text().strip(),
        }


class DailyStockPage(QWidget):
    def __init__(self):
        super().__init__()

        self.selected_sap_code: str | None = None

        self.date_input = QDateEdit()
        self.date_input.setCalendarPopup(True)
        self.date_input.setDate(QDate.currentDate())
        self.date_input.dateChanged.connect(self.refresh)

        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search SAP code or tyre description...")
        self.search_input.textChanged.connect(self.refresh_table)

        self.import_btn = QPushButton("Import Excel")
        self.import_btn.setObjectName("PrimaryButton")
        self.import_btn.clicked.connect(self.import_excel)

        self.export_btn = QPushButton("Export Excel")
        self.export_btn.setObjectName("PrimaryButton")
        self.export_btn.clicked.connect(self.export_excel)

        self.edit_btn = QPushButton("Edit Selected")
        self.edit_btn.setObjectName("SecondaryButton")
        self.edit_btn.setEnabled(False)
        self.edit_btn.clicked.connect(self.edit_selected_entry)

        self.refresh_btn = QPushButton("Refresh")
        self.refresh_btn.setObjectName("SecondaryButton")
        self.refresh_btn.clicked.connect(self.refresh)

        self.table = QTableWidget(0, 8)
        self.table.setHorizontalHeaderLabels(
            [
                "SAP Code",
                "Tyre Description",
                "Production Qty",
                "FG",
                "QC",
                "Scrap",
                "Blocked",
                "Available",
            ]
        )

        self._setup_table()
        self._apply_styles()
        self._build_ui()
        self.refresh()

    def _apply_styles(self) -> None:
        self.setStyleSheet(
            """
            QFrame#ControlCard,
            QFrame#TableCard {
                background: #ffffff;
                border: 1px solid #e2e8f0;
                border-radius: 16px;
            }

            QLabel#PageTitle {
                color: #0f172a;
                font-size: 20pt;
                font-weight: 950;
            }

            QLabel#PageHint {
                color: #64748b;
                font-size: 9.5pt;
                font-weight: 650;
            }

            QLabel#FieldLabel {
                color: #334155;
                font-size: 9pt;
                font-weight: 850;
            }

            QLineEdit, QDateEdit {
                background: #ffffff;
                color: #0f172a;
                border: 1px solid #cbd5e1;
                border-radius: 10px;
                padding: 9px 12px;
                font-size: 10pt;
                font-weight: 650;
                min-height: 24px;
            }

            QLineEdit:focus, QDateEdit:focus {
                border: 1px solid #2563eb;
            }

            QPushButton#PrimaryButton {
                background: #2563eb;
                color: #ffffff;
                border: none;
                border-radius: 10px;
                padding: 10px 18px;
                font-weight: 950;
                min-height: 26px;
            }

            QPushButton#PrimaryButton:hover {
                background: #1d4ed8;
            }

            QPushButton#SecondaryButton {
                background: #e2e8f0;
                color: #0f172a;
                border: none;
                border-radius: 10px;
                padding: 10px 18px;
                font-weight: 950;
                min-height: 26px;
            }

            QPushButton#SecondaryButton:hover {
                background: #cbd5e1;
            }

            QPushButton#SecondaryButton:disabled {
                background: #f1f5f9;
                color: #94a3b8;
            }

            QTableWidget {
                background: #ffffff;
                color: #0f172a;
                border: 1px solid #e2e8f0;
                border-radius: 12px;
                gridline-color: #e2e8f0;
                alternate-background-color: #f8fafc;
                selection-background-color: #dbeafe;
                selection-color: #0f172a;
            }

            QTableWidget::item {
                padding: 8px 10px;
                border: none;
            }

            QHeaderView::section {
                background: #f1f5f9;
                color: #1e293b;
                border: none;
                border-right: 1px solid #e2e8f0;
                border-bottom: 1px solid #e2e8f0;
                padding: 10px;
                font-weight: 950;
            }
            """
        )

    def _build_ui(self) -> None:
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(16)

        root.addWidget(self._build_control_card())
        root.addWidget(self._build_table_card(), 1)

    def _build_control_card(self) -> QFrame:
        card = QFrame()
        card.setObjectName("ControlCard")

        layout = QVBoxLayout(card)
        layout.setContentsMargins(18, 16, 18, 18)
        layout.setSpacing(14)

        header = QHBoxLayout()

        title_box = QVBoxLayout()
        title_box.setSpacing(4)

        title = QLabel("Daily Stock")
        title.setObjectName("PageTitle")

        hint = QLabel("Select a date, import daily production from Excel, edit quantities, and export a professional daily stock report.")
        hint.setObjectName("PageHint")
        hint.setWordWrap(True)

        title_box.addWidget(title)
        title_box.addWidget(hint)

        header.addLayout(title_box, 1)
        header.addWidget(self.import_btn)
        header.addWidget(self.export_btn)
        header.addWidget(self.edit_btn)
        header.addWidget(self.refresh_btn)

        layout.addLayout(header)

        form = QGridLayout()
        form.setHorizontalSpacing(12)
        form.setVerticalSpacing(10)

        date_label = QLabel("Date")
        date_label.setObjectName("FieldLabel")

        search_label = QLabel("Search")
        search_label.setObjectName("FieldLabel")

        form.addWidget(date_label, 0, 0)
        form.addWidget(self.date_input, 0, 1)

        form.addWidget(search_label, 1, 0)
        form.addWidget(self.search_input, 1, 1, 1, 5)

        form.setColumnStretch(1, 1)

        layout.addLayout(form)

        return card

    def _build_table_card(self) -> QFrame:
        card = QFrame()
        card.setObjectName("TableCard")

        layout = QVBoxLayout(card)
        layout.setContentsMargins(18, 16, 18, 18)
        layout.setSpacing(14)

        title = QLabel("Daily Production List")
        title.setObjectName("PageTitle")

        hint = QLabel("Double-click a row to edit. Available = FG + QC - Scrap - Blocked.")
        hint.setObjectName("PageHint")
        hint.setWordWrap(True)

        layout.addWidget(title)
        layout.addWidget(hint)
        layout.addWidget(self.table, 1)

        return card

    def _setup_table(self) -> None:
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.table.verticalHeader().setVisible(False)
        self.table.verticalHeader().setDefaultSectionSize(48)
        self.table.setAlternatingRowColors(True)

        header = self.table.horizontalHeader()
        header.setStretchLastSection(False)

        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        for col in range(2, 8):
            header.setSectionResizeMode(col, QHeaderView.ResizeMode.Fixed)

        self.table.setColumnWidth(0, 150)
        self.table.setColumnWidth(2, 125)
        self.table.setColumnWidth(3, 90)
        self.table.setColumnWidth(4, 90)
        self.table.setColumnWidth(5, 90)
        self.table.setColumnWidth(6, 95)
        self.table.setColumnWidth(7, 105)

        self.table.itemSelectionChanged.connect(self.on_selection_changed)
        self.table.itemDoubleClicked.connect(self.edit_selected_entry)

    def selected_date(self) -> date:
        return self.date_input.date().toPython()

    def refresh(self) -> None:
        try:
            self.ensure_daily_stock_table()
            self.seed_date_from_final_stock()
            self.refresh_table()
        except Exception as exc:
            QMessageBox.critical(self, "Daily Stock Error", str(exc))

    def ensure_daily_stock_table(self) -> None:
        with engine.begin() as connection:
            connection.execute(
                text(
                    """
                    CREATE TABLE IF NOT EXISTS mpps_daily_stock_entries (
                        id SERIAL PRIMARY KEY,
                        stock_date DATE NOT NULL,
                        sap_code VARCHAR(100) NOT NULL,
                        tyre_description TEXT NOT NULL,
                        production_qty INTEGER NOT NULL DEFAULT 0,
                        fg_qty INTEGER NOT NULL DEFAULT 0,
                        qc_qty INTEGER NOT NULL DEFAULT 0,
                        scrap_qty INTEGER NOT NULL DEFAULT 0,
                        blocked_qty INTEGER NOT NULL DEFAULT 0,
                        note TEXT,
                        source_file TEXT,
                        created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                        updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                        UNIQUE(stock_date, sap_code)
                    );
                    """
                )
            )

    def seed_date_from_final_stock(self) -> None:
        stock_date = self.selected_date()

        with engine.begin() as connection:
            connection.execute(
                text(
                    """
                    INSERT INTO mpps_daily_stock_entries
                        (
                            stock_date,
                            sap_code,
                            tyre_description,
                            production_qty,
                            fg_qty,
                            qc_qty,
                            scrap_qty,
                            blocked_qty,
                            note,
                            updated_at
                        )
                    SELECT
                        :stock_date,
                        sap_code,
                        tyre_description,
                        0,
                        0,
                        0,
                        0,
                        0,
                        'Auto-created daily production line.',
                        CURRENT_TIMESTAMP
                    FROM mpps_sap_stock_items
                    WHERE is_active = TRUE
                    ON CONFLICT (stock_date, sap_code) DO NOTHING;
                    """
                ),
                {"stock_date": stock_date},
            )

    def refresh_table(self) -> None:
        self.selected_sap_code = None
        self.edit_btn.setEnabled(False)

        search = self.search_input.text().strip()
        stock_date = self.selected_date()

        params = {
            "stock_date": stock_date,
            "search": f"%{search}%",
        }

        where = "WHERE stock_date = :stock_date"

        if search:
            where += """
                AND (
                    sap_code ILIKE :search
                    OR tyre_description ILIKE :search
                )
            """

        sql = f"""
            SELECT
                sap_code,
                tyre_description,
                production_qty,
                fg_qty,
                qc_qty,
                scrap_qty,
                blocked_qty,
                (fg_qty + qc_qty - scrap_qty - blocked_qty) AS available_qty
            FROM mpps_daily_stock_entries
            {where}
            ORDER BY available_qty DESC, sap_code ASC;
        """

        with engine.begin() as connection:
            rows = connection.execute(text(sql), params).mappings().all()

        self.table.setRowCount(0)

        for row_index, row in enumerate(rows):
            self.table.insertRow(row_index)

            values = [
                row["sap_code"],
                row["tyre_description"],
                self._format_int(row["production_qty"]),
                self._format_int(row["fg_qty"]),
                self._format_int(row["qc_qty"]),
                self._format_int(row["scrap_qty"]),
                self._format_int(row["blocked_qty"]),
                self._format_int(row["available_qty"]),
            ]

            for col_index, value in enumerate(values):
                item = QTableWidgetItem(str(value))
                item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)

                if col_index in {0, 2, 3, 4, 5, 6, 7}:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

                if col_index == 0:
                    item.setData(Qt.ItemDataRole.UserRole, row["sap_code"])

                if col_index == 7:
                    self._apply_available_style(item, int(row["available_qty"] or 0))

                self.table.setItem(row_index, col_index, item)

        self.table.resizeRowsToContents()

    def import_excel(self) -> None:
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Import Daily Production Excel",
            "",
            "Excel Files (*.xlsx *.xlsm)",
        )

        if not file_path:
            return

        try:
            imported = self._import_excel_file(Path(file_path))
            self.refresh()

            QMessageBox.information(
                self,
                "Import Complete",
                f"Imported / updated {imported} daily production rows.",
            )

        except Exception as exc:
            QMessageBox.critical(self, "Import Failed", str(exc))

    def _import_excel_file(self, file_path: Path) -> int:
        stock_date = self.selected_date()
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.active

        headers = {}
        for col in range(1, sheet.max_column + 1):
            value = sheet.cell(row=1, column=col).value
            if value is not None:
                headers[self._normalise_header(str(value))] = col

        sap_col = self._find_header(headers, ["sap code", "sap", "sap_code", "material code", "code"])
        desc_col = self._find_header(headers, ["tyre description", "description", "item description"])
        production_col = self._find_header(headers, ["production qty", "produced qty", "daily production", "qty", "quantity"])
        fg_col = self._find_header(headers, ["fg", "fg qty", "fg stock"])
        qc_col = self._find_header(headers, ["qc", "qc qty", "qc stock"])
        scrap_col = self._find_header(headers, ["scrap", "scrap qty"])
        blocked_col = self._find_header(headers, ["blocked", "blocked qty"])
        note_col = self._find_header(headers, ["note", "remarks", "remark"])

        if sap_col is None:
            raise ValueError("Excel file must include a SAP Code column.")

        imported = 0

        with engine.begin() as connection:
            for row in range(2, sheet.max_row + 1):
                sap_code = self._cell_text(sheet, row, sap_col)

                if not sap_code:
                    continue

                existing_desc = connection.execute(
                    text(
                        """
                        SELECT tyre_description
                        FROM mpps_sap_stock_items
                        WHERE sap_code = :sap_code
                        LIMIT 1;
                        """
                    ),
                    {"sap_code": sap_code},
                ).scalar()

                excel_desc = self._cell_text(sheet, row, desc_col) if desc_col else ""
                tyre_description = excel_desc or existing_desc or "-"

                data = {
                    "stock_date": stock_date,
                    "sap_code": sap_code,
                    "tyre_description": tyre_description,
                    "production_qty": self._cell_int(sheet, row, production_col),
                    "fg_qty": self._cell_int(sheet, row, fg_col),
                    "qc_qty": self._cell_int(sheet, row, qc_col),
                    "scrap_qty": self._cell_int(sheet, row, scrap_col),
                    "blocked_qty": self._cell_int(sheet, row, blocked_col),
                    "note": self._cell_text(sheet, row, note_col) if note_col else "",
                    "source_file": str(file_path.name),
                }

                connection.execute(
                    text(
                        """
                        INSERT INTO mpps_daily_stock_entries
                            (
                                stock_date,
                                sap_code,
                                tyre_description,
                                production_qty,
                                fg_qty,
                                qc_qty,
                                scrap_qty,
                                blocked_qty,
                                note,
                                source_file,
                                updated_at
                            )
                        VALUES
                            (
                                :stock_date,
                                :sap_code,
                                :tyre_description,
                                :production_qty,
                                :fg_qty,
                                :qc_qty,
                                :scrap_qty,
                                :blocked_qty,
                                :note,
                                :source_file,
                                CURRENT_TIMESTAMP
                            )
                        ON CONFLICT (stock_date, sap_code)
                        DO UPDATE SET
                            tyre_description = EXCLUDED.tyre_description,
                            production_qty = EXCLUDED.production_qty,
                            fg_qty = EXCLUDED.fg_qty,
                            qc_qty = EXCLUDED.qc_qty,
                            scrap_qty = EXCLUDED.scrap_qty,
                            blocked_qty = EXCLUDED.blocked_qty,
                            note = EXCLUDED.note,
                            source_file = EXCLUDED.source_file,
                            updated_at = CURRENT_TIMESTAMP;
                        """
                    ),
                    data,
                )

                imported += 1

        return imported

    def export_excel(self) -> None:
        stock_date = self.selected_date()
        default_name = f"daily_stock_{stock_date.strftime('%Y_%m_%d')}.xlsx"

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export Daily Stock Excel",
            default_name,
            "Excel Files (*.xlsx)",
        )

        if not file_path:
            return

        if not file_path.lower().endswith(".xlsx"):
            file_path += ".xlsx"

        try:
            rows = self._fetch_export_rows()
            self._write_professional_excel(Path(file_path), rows)
            QMessageBox.information(self, "Export Complete", f"Excel report exported:\n{file_path}")

        except Exception as exc:
            QMessageBox.critical(self, "Export Failed", str(exc))

    def _fetch_export_rows(self) -> list[dict]:
        stock_date = self.selected_date()

        with engine.begin() as connection:
            rows = connection.execute(
                text(
                    """
                    SELECT
                        sap_code,
                        tyre_description,
                        production_qty,
                        fg_qty,
                        qc_qty,
                        scrap_qty,
                        blocked_qty,
                        (fg_qty + qc_qty - scrap_qty - blocked_qty) AS available_qty,
                        note
                    FROM mpps_daily_stock_entries
                    WHERE stock_date = :stock_date
                    ORDER BY available_qty DESC, sap_code ASC;
                    """
                ),
                {"stock_date": stock_date},
            ).mappings().all()

        return [dict(row) for row in rows]

    def _write_professional_excel(self, file_path: Path, rows: list[dict]) -> None:
        stock_date = self.selected_date()

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Daily Stock"

        title_fill = PatternFill("solid", fgColor="0F172A")
        header_fill = PatternFill("solid", fgColor="DBEAFE")
        total_fill = PatternFill("solid", fgColor="F1F5F9")
        white_font = Font(color="FFFFFF", bold=True, size=14)
        header_font = Font(color="0F172A", bold=True)
        border = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1"),
        )

        sheet.merge_cells("A1:I1")
        sheet["A1"] = "LAUGFS Rubber - Daily Stock / Production Report"
        sheet["A1"].fill = title_fill
        sheet["A1"].font = white_font
        sheet["A1"].alignment = Alignment(horizontal="center")

        sheet.merge_cells("A2:I2")
        sheet["A2"] = f"Date: {stock_date.strftime('%Y-%m-%d')}"
        sheet["A2"].font = Font(bold=True, color="334155")
        sheet["A2"].alignment = Alignment(horizontal="center")

        headers = [
            "SAP Code",
            "Tyre Description",
            "Production Qty",
            "FG",
            "QC",
            "Scrap",
            "Blocked",
            "Available",
            "Note",
        ]

        start_row = 4

        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=start_row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

        for idx, row in enumerate(rows, start_row + 1):
            values = [
                row.get("sap_code"),
                row.get("tyre_description"),
                int(row.get("production_qty") or 0),
                int(row.get("fg_qty") or 0),
                int(row.get("qc_qty") or 0),
                int(row.get("scrap_qty") or 0),
                int(row.get("blocked_qty") or 0),
                int(row.get("available_qty") or 0),
                row.get("note") or "",
            ]

            for col, value in enumerate(values, 1):
                cell = sheet.cell(row=idx, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(vertical="center")

                if col in {3, 4, 5, 6, 7, 8}:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        total_row = start_row + len(rows) + 2
        sheet.cell(row=total_row, column=1, value="TOTAL").font = header_font
        sheet.cell(row=total_row, column=1).fill = total_fill

        for col in range(3, 9):
            letter = get_column_letter(col)
            cell = sheet.cell(row=total_row, column=col, value=f"=SUM({letter}{start_row + 1}:{letter}{total_row - 2})")
            cell.font = header_font
            cell.fill = total_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

        widths = {
            "A": 16,
            "B": 48,
            "C": 16,
            "D": 12,
            "E": 12,
            "F": 12,
            "G": 12,
            "H": 14,
            "I": 34,
        }

        for col, width in widths.items():
            sheet.column_dimensions[col].width = width

        sheet.freeze_panes = "A5"
        sheet.auto_filter.ref = f"A4:I{max(start_row + 1, start_row + len(rows))}"

        workbook.save(file_path)

    def edit_selected_entry(self, *args) -> None:
        if not self.selected_sap_code:
            return

        row = self._get_selected_entry()

        if row is None:
            QMessageBox.warning(self, "Daily Stock", "Selected row was not found.")
            return

        dialog = DailyStockEditDialog(self, dict(row))

        if dialog.exec() != QDialog.DialogCode.Accepted:
            return

        data = dialog.get_data()
        data["stock_date"] = self.selected_date()

        with engine.begin() as connection:
            connection.execute(
                text(
                    """
                    UPDATE mpps_daily_stock_entries
                    SET
                        production_qty = :production_qty,
                        fg_qty = :fg_qty,
                        qc_qty = :qc_qty,
                        scrap_qty = :scrap_qty,
                        blocked_qty = :blocked_qty,
                        note = :note,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE stock_date = :stock_date
                      AND sap_code = :sap_code;
                    """
                ),
                data,
            )

        self.refresh()

    def _get_selected_entry(self):
        if not self.selected_sap_code:
            return None

        with engine.begin() as connection:
            return connection.execute(
                text(
                    """
                    SELECT *
                    FROM mpps_daily_stock_entries
                    WHERE stock_date = :stock_date
                      AND sap_code = :sap_code
                    LIMIT 1;
                    """
                ),
                {
                    "stock_date": self.selected_date(),
                    "sap_code": self.selected_sap_code,
                },
            ).mappings().first()

    def on_selection_changed(self) -> None:
        selected_items = self.table.selectedItems()

        if not selected_items:
            self.selected_sap_code = None
            self.edit_btn.setEnabled(False)
            return

        row = selected_items[0].row()
        sap_item = self.table.item(row, 0)

        if sap_item is None:
            self.selected_sap_code = None
            self.edit_btn.setEnabled(False)
            return

        self.selected_sap_code = sap_item.data(Qt.ItemDataRole.UserRole)
        self.edit_btn.setEnabled(bool(self.selected_sap_code))

    def _normalise_header(self, value: str) -> str:
        return " ".join(value.replace("_", " ").strip().lower().split())

    def _find_header(self, headers: dict[str, int], candidates: list[str]) -> int | None:
        normalised = {self._normalise_header(key): value for key, value in headers.items()}

        for candidate in candidates:
            key = self._normalise_header(candidate)

            if key in normalised:
                return normalised[key]

        return None

    def _cell_text(self, sheet, row: int, col: int | None) -> str:
        if col is None:
            return ""

        value = sheet.cell(row=row, column=col).value

        if value is None:
            return ""

        return str(value).strip()

    def _cell_int(self, sheet, row: int, col: int | None) -> int:
        if col is None:
            return 0

        value = sheet.cell(row=row, column=col).value

        if value is None or value == "":
            return 0

        try:
            return int(float(value))
        except Exception:
            return 0

    def _apply_available_style(self, item: QTableWidgetItem, value: int) -> None:
        if value > 0:
            item.setForeground(QColor("#166534"))
            item.setBackground(QColor("#dcfce7"))
        elif value == 0:
            item.setForeground(QColor("#92400e"))
            item.setBackground(QColor("#fef3c7"))
        else:
            item.setForeground(QColor("#991b1b"))
            item.setBackground(QColor("#fee2e2"))

    def _format_int(self, value) -> str:
        try:
            return f"{int(value or 0):,}"
        except Exception:
            return "0"
