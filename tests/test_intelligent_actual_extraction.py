import unittest
from datetime import date

from openpyxl import Workbook

from app.services.intelligent_excel_import_service import IntelligentExcelImportService


class IntelligentActualExtractionTests(unittest.TestCase):
    def test_prod_date_pair_is_day_night_verified_actual(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "PROD"
        formula_wb = Workbook()
        formula_ws = formula_wb.active
        formula_ws.title = "PROD"

        ws["B3"] = "SAP CODE"
        ws["C3"] = "Material Description"
        ws["BI3"] = date(2025, 11, 28)
        ws["BJ3"] = None
        ws["BK3"] = date(2025, 11, 29)
        ws["BL3"] = None
        ws["B4"] = "6000139"
        ws["C4"] = "TEST TYRE"
        ws["BI4"] = 10
        ws["BJ4"] = 8
        ws["BK4"] = 7
        ws["BL4"] = 6

        _, _, history, actual_dates = IntelligentExcelImportService()._extract_prod(
            wb,
            formula_wb,
            {"PRODUCTION_STOCK_SHIPMENTS": "PROD"},
            date(2025, 11, 30),
            {},
            [],
        )

        self.assertEqual(len(history), 2)
        self.assertEqual(len(actual_dates), 2)
        self.assertEqual(actual_dates[0]["source_day_column"], "BI")
        self.assertEqual(actual_dates[0]["source_night_column"], "BJ")
        self.assertEqual(history[0]["production_date"], "2025-11-28")
        self.assertEqual(history[0]["day_actual_qty"], 10)
        self.assertEqual(history[0]["night_actual_qty"], 8)
        self.assertEqual(history[0]["production_qty"], 18)
        self.assertEqual(history[0]["source_day_column"], "BI")
        self.assertEqual(history[0]["source_night_column"], "BJ")
        self.assertEqual(history[0]["source_semantics"], "VERIFIED_ACTUAL_PRODUCTION")


if __name__ == "__main__":
    unittest.main()
