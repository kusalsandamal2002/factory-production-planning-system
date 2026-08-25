from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from app.services.current_stock_service import CurrentStockService


class CurrentStockServiceTests(unittest.TestCase):
    def test_progress_uses_current_stock_over_shipment_caps_at_100_and_is_blank_without_shipment(self):
        self.assertEqual(CurrentStockService.progress_percent(100, 60), 60.0)
        self.assertEqual(CurrentStockService.progress_percent(100, 130), 100.0)
        self.assertIsNone(CurrentStockService.progress_percent(0, 10))
        self.assertIsNone(CurrentStockService.progress_percent(None, 10))

    def test_extracts_hr_to_hv_directly_from_prod(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "OVEN SHEET PLAN AUGUST 05-2026.xlsx"
            workbook = Workbook()
            ws = workbook.active
            ws.title = "PROD"
            ws["B3"] = "SAP CODE"
            ws["C3"] = "Material Description"
            ws["HR3"] = "Total To be Shipped"
            ws["HS3"] = "TOTAL STOCK"
            ws["HT3"] = "Balance to Produced"
            ws["HU3"] = "total Plan"
            ws["HV3"] = "total To be plan"
            ws["B4"] = 60000001
            ws["C4"] = "TEST TYRE"
            ws["HR4"] = 100
            ws["HS4"] = 60
            ws["HT4"] = 40
            ws["HU4"] = 25
            ws["HV4"] = 15
            workbook.save(path)
            workbook.close()

            rows = CurrentStockService.extract_workbook(path)

        self.assertEqual(len(rows), 1)
        row = rows[0]
        self.assertEqual(row["sap_code"], "60000001")
        self.assertEqual(row["total_to_be_shipped"], 100)
        self.assertEqual(row["current_stock"], 60)
        self.assertEqual(row["balance_to_produce"], 40)
        self.assertEqual(row["total_plan"], 25)
        self.assertEqual(row["total_to_be_plan"], 15)

    def test_direct_sheet_values_are_not_recalculated(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "source.xlsx"
            workbook = Workbook()
            ws = workbook.active
            ws.title = "PROD"
            ws["B4"] = "60000002"
            ws["C4"] = "DIRECT VALUES"
            ws["HR4"] = 50
            ws["HS4"] = 80
            ws["HT4"] = 777
            ws["HU4"] = 13
            ws["HV4"] = -9
            workbook.save(path)
            workbook.close()

            row = CurrentStockService.extract_workbook(path)[0]

        self.assertEqual(row["current_stock"], 80)
        self.assertEqual(row["balance_to_produce"], 777)
        self.assertEqual(row["total_to_be_plan"], -9)

    def test_operational_filters(self):
        rows = [
            {
                "sap_code": "A",
                "item_description": "NO SHIPMENT",
                "total_to_be_shipped": 0,
                "current_stock": 25,
                "balance_to_produce": 0,
            },
            {
                "sap_code": "B",
                "item_description": "COVERED",
                "total_to_be_shipped": 100,
                "current_stock": 120,
                "balance_to_produce": 0,
            },
            {
                "sap_code": "C",
                "item_description": "NEEDS PRODUCTION",
                "total_to_be_shipped": 100,
                "current_stock": 60,
                "balance_to_produce": 40,
            },
        ]

        self.assertEqual(
            [r["sap_code"] for r in CurrentStockService.filter_rows(
                rows, filter_mode=CurrentStockService.FILTER_SHIPMENT
            )],
            ["B", "C"],
        )
        self.assertEqual(
            [r["sap_code"] for r in CurrentStockService.filter_rows(
                rows, filter_mode=CurrentStockService.FILTER_COVERED
            )],
            ["B"],
        )
        self.assertEqual(
            [r["sap_code"] for r in CurrentStockService.filter_rows(
                rows, filter_mode=CurrentStockService.FILTER_TO_PRODUCE
            )],
            ["C"],
        )

    def test_search_and_filter_work_together(self):
        rows = [
            {
                "sap_code": "600001",
                "item_description": "ALPHA",
                "total_to_be_shipped": 50,
                "current_stock": 10,
                "balance_to_produce": 40,
            },
            {
                "sap_code": "600002",
                "item_description": "BETA",
                "total_to_be_shipped": 50,
                "current_stock": 10,
                "balance_to_produce": 40,
            },
        ]
        result = CurrentStockService.filter_rows(
            rows,
            query="BETA",
            filter_mode=CurrentStockService.FILTER_TO_PRODUCE,
        )
        self.assertEqual([r["sap_code"] for r in result], ["600002"])

    def test_summary_uses_visible_rows_only(self):
        rows = [
            {
                "total_to_be_shipped": 100,
                "current_stock": 60,
                "balance_to_produce": 40,
                "total_plan": 10,
                "total_to_be_plan": 30,
            },
            {
                "total_to_be_shipped": 50,
                "current_stock": 50,
                "balance_to_produce": 0,
                "total_plan": 50,
                "total_to_be_plan": 0,
            },
        ]
        result = CurrentStockService.summarize_rows(rows[:1])
        self.assertEqual(result["items"], 1)
        self.assertEqual(result["total_to_be_shipped"], 100)
        self.assertEqual(result["current_stock"], 60)
        self.assertEqual(result["balance_to_produce"], 40)


if __name__ == "__main__":
    unittest.main()
